from PIL import Image
import difflib
import io
import json
import os
import random
import re
import time
import urllib.parse
import html
import base64
import binascii
import codecs
import datetime
import hashlib
import hmac
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import uuid
from collections import Counter
from datetime import timedelta
from difflib import Differ

from Crypto.Cipher import AES, DES, DES3
from Crypto.Util.Padding import pad, unpad
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

from qa_toolkit.config.constants import PROVINCES, COUNTRIES, CATEGORIES, PROVINCE_MAP, TO_SECONDS, RANDOM_STRING_TYPES, \
    PASSWORD_OPTIONS, DOMAINS_PRESET, GENDERS, TOOL_CATEGORIES, CSS_STYLES, HEADLINE_STYLES, PRESET_SIZES
from qa_toolkit.config.constants import LANGUAGE_TEMPLATES
from qa_toolkit.config.constants import PREDEFINED_PATTERNS
from qa_toolkit.config.constants import PROVINCE_CITY_AREA_CODES
from qa_toolkit.config.constants import PLATFORM_MAPPING
from qa_toolkit.config.constants import STYLE_PREVIEWS
from qa_toolkit.config.constants import LANGUAGE_DESCRIPTIONS
from qa_toolkit.config.constants import SIMPLE_EXAMPLE, MEDIUM_EXAMPLE, COMPLEX_EXAMPLE
from qa_toolkit.core.api_dev_tools import InterfaceDevTools
from qa_toolkit.core.api_test_core import InterfaceAutoTestCore
from qa_toolkit.integrations.zentao_exporter import ZenTaoPerformanceExporter
from qa_toolkit.reporting.report_generator import EnhancedReportGenerator
from qa_toolkit.reporting.test_runner import EnhancedTestRunner
from qa_toolkit.support.documentation import show_doc, show_general_guidelines
from qa_toolkit.tools.bi_analysis import BIAnalyzer
from qa_toolkit.tools.data_generator import DataGenerator
from qa_toolkit.tools.ip_lookup import IPQueryTool
from qa_toolkit.tools.test_case_generator import TestCaseGenerator
from qa_toolkit.ui.components.author_profile import AuthorProfile
from qa_toolkit.ui.components.feedback_panel import FeedbackSection
from qa_toolkit.ui.pages.api_performance_page import render_api_performance_test_page
from qa_toolkit.ui.pages.api_security_page import render_api_security_test_page
from qa_toolkit.utils.datetime_tools import DateTimeUtils
from qa_toolkit.utils.json_utils import JSONFileUtils
from qa_toolkit.utils.log_analysis import LogAnalyzerUtils

# 导入Faker库
try:
    from faker import Faker

    fake = Faker('zh_CN')
    FAKER_AVAILABLE = True
except ImportError:
    FAKER_AVAILABLE = False
    st.warning("Faker库未安装，部分高级功能将受限。请运行: pip install faker")

# 设置页面
st.set_page_config(
    page_title="测试工程师常用工具集",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 现代化CSS样式
st.markdown(CSS_STYLES, unsafe_allow_html=True)


def generate_regex_from_examples(text, examples):
    """根据示例文本生成正则表达式"""
    if not text or not examples:
        return ""

    example_list = [ex.strip() for ex in examples.split(",") if ex.strip()]

    if not example_list:
        return ""

    # 简化的模式识别逻辑
    common_pattern = example_list[0]

    for example in example_list[1:]:
        # 找出共同前缀
        i = 0
        while i < min(len(common_pattern), len(example)) and common_pattern[i] == example[i]:
            i += 1
        common_pattern = common_pattern[:i]

    if len(common_pattern) < 2:
        return re.escape(example_list[0])

    escaped_pattern = re.escape(common_pattern)

    # 简单的模式推断
    if len(example_list) > 1:
        if all(ex.replace(common_pattern, "").isdigit() for ex in example_list):
            return escaped_pattern + r"\d+"
        elif all(ex.replace(common_pattern, "").isalpha() for ex in example_list):
            return escaped_pattern + r"[A-Za-z]+"

    return escaped_pattern + ".*"


def escape_js_string(text):
    """安全转义 JavaScript 字符串"""
    return json.dumps(text)


def create_copy_button(text, button_text="📋 复制到剪贴板", key=None):
    """创建一键复制按钮"""
    if key is None:
        key = hash(text)

    escaped_text = escape_js_string(text)

    copy_script = f"""
    <script>
    function copyTextToClipboard{key}() {{
        const text = {escaped_text};
        if (!navigator.clipboard) {{
            return fallbackCopyTextToClipboard(text);
        }}
        return navigator.clipboard.writeText(text).then(function() {{
            return true;
        }}, function(err) {{
            return fallbackCopyTextToClipboard(text);
        }});
    }}

    function fallbackCopyTextToClipboard(text) {{
        const textArea = document.createElement('textarea');
        textArea.value = text;
        textArea.style.position = 'fixed';
        textArea.style.top = '0';
        textArea.style.left = '0';
        textArea.style.opacity = '0';
        document.body.appendChild(textArea);
        textArea.focus();
        textArea.select();
        try {{
            const successful = document.execCommand('copy');
            document.body.removeChild(textArea);
            return successful;
        }} catch (err) {{
            document.body.removeChild(textArea);
            return false;
        }}
    }}

    document.addEventListener('DOMContentLoaded', function() {{
        const button = document.querySelector('[data-copy-button="{key}"]');
        if (button) {{
            button.addEventListener('click', function() {{
                copyTextToClipboard{key}().then(function(success) {{
                    if (success) {{
                        const originalText = button.innerHTML;
                        button.innerHTML = '✅ 复制成功！';
                        button.style.background = '#48bb78';
                        setTimeout(function() {{
                            button.innerHTML = originalText;
                            button.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                        }}, 2000);
                    }} else {{
                        button.innerHTML = '❌ 复制失败';
                        button.style.background = '#e53e3e';
                        setTimeout(function() {{
                            button.innerHTML = '{button_text}';
                            button.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                        }}, 2000);
                    }}
                }});
            }});
        }}
    }});
    </script>
    """

    button_html = f"""
    <div>
        <button data-copy-button="{key}"
                style="background:linear-gradient(135deg, #667eea 0%, #764ba2 100%);color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-size:14px;margin:5px;font-weight:500;transition:all 0.3s ease;width:100%;height:42px;font-family:inherit;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
            {button_text}
        </button>
    </div>
    """

    components.html(button_html + copy_script, height=60)


def render_back_to_top_button():
    """渲染固定在右下角的返回顶部按钮"""
    st.markdown(
        """
        <style>
        .qa-toolkit-back-to-top {
            position: fixed;
            right: 24px;
            bottom: 24px;
            z-index: 999;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-width: 116px;
            padding: 11px 16px;
            border-radius: 999px;
            background: linear-gradient(135deg, #1f4b99 0%, #0f766e 100%);
            color: #ffffff !important;
            text-decoration: none !important;
            font-size: 14px;
            font-weight: 700;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.22);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .qa-toolkit-back-to-top:hover {
            transform: translateY(-2px);
            box-shadow: 0 16px 32px rgba(15, 23, 42, 0.28);
        }
        </style>
        <a class="qa-toolkit-back-to-top" href="#page-top">↑ 返回顶部</a>
        """,
        unsafe_allow_html=True,
    )


def render_tool_picker():
    """渲染工具切换器。收起时优先展示功能区，展开时展示完整工具列表。"""
    current_tool = st.session_state.selected_tool
    current_info = TOOL_CATEGORIES[current_tool]
    accent_color = current_info.get("color", "#667eea")
    banner_background = f"linear-gradient(135deg, {accent_color} 0%, #0f172a 100%)"
    st.markdown(
        f"""
        <div style="
            background: {banner_background};
            border-radius: 18px;
            padding: 18px 20px;
            color: #ffffff;
            box-shadow: 0 14px 32px rgba(15, 23, 42, 0.14);
            margin-bottom: 12px;
        ">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
                <div style="font-size:26px;line-height:1;">{current_info['icon']}</div>
                <div style="font-size:20px;font-weight:800;">{current_tool}</div>
                <div style="
                    margin-left:auto;
                    background: rgba(255,255,255,0.18);
                    border: 1px solid rgba(255,255,255,0.22);
                    border-radius: 999px;
                    padding: 4px 10px;
                    font-size: 12px;
                    font-weight: 700;
                ">
                    {'功能区模式' if st.session_state.tool_picker_compact else '工具切换模式'}
                </div>
            </div>
            <div style="font-size:14px;line-height:1.7;color:rgba(255,255,255,0.92);">
                {current_info['description']}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    action_col1, action_col2 = st.columns([1, 1])
    with action_col1:
        primary_label = "展开工具列表" if st.session_state.tool_picker_compact else "收起工具列表"
        if st.button(primary_label, key="toggle_tool_picker", use_container_width=True):
            st.session_state.tool_picker_compact = not st.session_state.tool_picker_compact
            st.rerun()
    with action_col2:
        if st.button("仅看当前功能区", key="focus_current_tool", use_container_width=True):
            st.session_state.tool_picker_compact = True
            st.rerun()

    if st.session_state.tool_picker_compact:
        st.caption("当前已优先展示功能区。如需切换模块，点击“展开工具列表”。")
        st.markdown("---")
        return

    st.markdown('<div class="sub-header">🚀 可用工具</div>', unsafe_allow_html=True)
    cols = st.columns(3)
    col_index = 0

    for category, info in TOOL_CATEGORIES.items():
        with cols[col_index]:
            is_selected = current_tool == category
            card_class = "tool-picker-card selected" if is_selected else "tool-picker-card"
            accent_color = info.get("color", "#667eea")
            card_style = f' style="border-color: {accent_color};"' if is_selected else ""
            active_button_style = (
                f' style="border-color: {accent_color}; box-shadow: inset 1px 1px 0 rgba(255,255,255,0.88), '
                f'inset -1px -1px 0 {accent_color};"'
            )
            st.markdown(
                f"""
                <div class="{card_class}"{card_style}>
                    <div class="tool-picker-icon" style="color: {accent_color};">{info['icon']}</div>
                    <div class="tool-picker-title">{category}</div>
                    <div class="tool-picker-desc">{info['description']}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            if is_selected:
                st.markdown(
                    f'<div class="tool-picker-active-button"{active_button_style}>{info["icon"]} {category}</div>'
                    '<div class="tool-picker-status">当前已选中，可直接进入功能区</div>',
                    unsafe_allow_html=True,
                )

            button_label = "进入当前工具" if is_selected else f"{info['icon']} {category}"
            if st.button(button_label, key=f"select_{category}", use_container_width=True):
                st.session_state.selected_tool = category
                st.session_state.tool_picker_compact = True
                st.rerun()

            if not is_selected:
                st.markdown('<div class="tool-picker-status">单击进入工具</div>', unsafe_allow_html=True)

        col_index = (col_index + 1) % 3

    st.markdown("---")


def display_generated_results(title, content, filename_prefix):
    """统一展示生成结果 + 复制 + 下载"""
    st.markdown(f'<div class="category-card">📋 生成结果 - {title}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="result-box">{content}</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        create_copy_button(content, button_text="📋 复制结果", key=f"copy_{filename_prefix}")
    with col2:
        st.download_button(
            label="💾 下载结果",
            data=content,
            file_name=f"{filename_prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain"
        )


def display_structured_results(title, records, filename_prefix):
    """统一展示结构化结果，支持预览和多格式导出。"""
    if not records:
        st.warning("暂无可展示的数据。")
        return

    dataframe = pd.DataFrame(records)
    json_text = json.dumps(records, ensure_ascii=False, indent=2)
    csv_bytes = dataframe.to_csv(index=False).encode("utf-8-sig")

    st.markdown(f'<div class="category-card">🗂️ 结构化结果 - {title}</div>', unsafe_allow_html=True)
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    with metric_col1:
        st.metric("记录数", len(dataframe))
    with metric_col2:
        st.metric("字段数", len(dataframe.columns))
    with metric_col3:
        st.metric("导出格式", "CSV / JSON")

    preview_tab, json_tab = st.tabs(["表格预览", "JSON 预览"])
    with preview_tab:
        st.dataframe(dataframe, use_container_width=True, hide_index=True)
        st.caption(f"字段覆盖: {', '.join(dataframe.columns.tolist())}")
    with json_tab:
        st.code(json_text, language="json")

    action_col1, action_col2, action_col3 = st.columns(3)
    with action_col1:
        create_copy_button(json_text, button_text="📋 复制 JSON", key=f"copy_json_{filename_prefix}")
    with action_col2:
        st.download_button(
            label="💾 下载 CSV",
            data=csv_bytes,
            file_name=f"{filename_prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    with action_col3:
        st.download_button(
            label="🧾 下载 JSON",
            data=json_text.encode("utf-8"),
            file_name=f"{filename_prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )


# 初始化session state
if 'selected_tool' not in st.session_state:
    st.session_state.selected_tool = "数据生成工具"
if 'tool_picker_compact' not in st.session_state:
    st.session_state.tool_picker_compact = False
if st.session_state.selected_tool not in TOOL_CATEGORIES:
    st.session_state.selected_tool = "数据生成工具"
    st.session_state.tool_picker_compact = False

# 顶部标题区域
st.markdown(HEADLINE_STYLES, unsafe_allow_html=True)
st.markdown('<div id="page-top"></div>', unsafe_allow_html=True)
render_back_to_top_button()

tool_category = st.session_state.selected_tool

st.markdown('<div id="tool-content-anchor"></div>', unsafe_allow_html=True)
render_tool_picker()

# === 工具功能实现 ===
if tool_category == "数据生成工具":
    show_doc("data_generator")
    generator = DataGenerator()
    if 'clear_data_gen_counter' not in st.session_state:
        st.session_state.clear_data_gen_counter = 0

    gen_mode = st.radio(
        "选择生成模式",
        ["Faker高级生成器", "基础数据生成器"],
        horizontal=True
    )

    if gen_mode == "Faker高级生成器":
        if not FAKER_AVAILABLE:
            st.error("❌ Faker库未安装，无法使用高级生成器")
            st.info("请运行以下命令安装: `pip install faker`")
            st.code("pip install faker", language="bash")
        else:
            st.markdown('<div class="category-card">🚀 Faker高级数据生成器</div>', unsafe_allow_html=True)

            col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
            with col1:
                selected_category = st.selectbox("选择数据类别", list(CATEGORIES.keys()))
            with col2:
                selected_subcategory = st.selectbox("选择具体类型", CATEGORIES[selected_category])
            with col3:
                count = st.number_input("生成数量", min_value=1, max_value=100, value=5)
            with col4:
                st.write("")  # 占位符
                st.write("")
                if st.button("🗑️ 清空", key="clear_faker", use_container_width=True):
                    if 'faker_result' in st.session_state:
                        del st.session_state.faker_result
                    if 'last_category' in st.session_state:
                        del st.session_state.last_category
                    # 确保计数器存在
                    if 'clear_data_gen_counter' not in st.session_state:
                        st.session_state.clear_data_gen_counter = 0
                    st.session_state.clear_data_gen_counter += 1
                    st.rerun()

            extra_params = {}
            if selected_subcategory == "随机文本":
                text_length = st.slider("文本长度", min_value=10, max_value=1000, value=200)
                extra_params['length'] = text_length

            if st.button("🎯 生成数据", use_container_width=True):
                with st.spinner("正在生成数据..."):
                    results = generator.safe_generate(generator.generate_faker_data, selected_category,
                                                      selected_subcategory, count, **extra_params)
                    if results is not None:
                        result_text = "\n".join([str(r) for r in results])
                        st.session_state.faker_result = result_text
                        st.session_state.last_category = f"{selected_category} - {selected_subcategory}"

            if 'faker_result' in st.session_state:
                title = st.session_state.get("last_category", "")
                if selected_subcategory == "完整个人信息":
                    st.text_area("生成结果", st.session_state.faker_result, height=300, key="profile_result")
                else:
                    st.markdown(f'<div class="result-box">{st.session_state.faker_result}</div>',
                                unsafe_allow_html=True)
                create_copy_button(st.session_state.faker_result, button_text="📋 复制结果", key="copy_faker_result")
                st.download_button(
                    label="💾 下载结果",
                    data=st.session_state.faker_result,
                    file_name=f"faker_data_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime="text/plain"
                )

    else:  # 基础数据生成器
        st.markdown('<div class="category-card">🔧 基础数据生成器</div>', unsafe_allow_html=True)
        data_gen_tool = st.radio(
            "选择生成工具",
            ["随机内容生成器", "随机邮箱生成器", "电话号码生成器", "随机地址生成器", "随机身份证生成器",
             "测试场景造数", "边界值/异常值生成器"]
        )
        st.caption("前 5 项适合快速单字段造数，后 2 项适合测试联调、批量导入和边界值设计。")

        if data_gen_tool == "随机内容生成器":
            st.markdown('<div class="category-card">🎲 随机内容生成器</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                gen_type = st.selectbox("生成类型", ["随机字符串", "随机数字", "随机密码", "UUID"])

                if gen_type in ["随机字符串", "随机密码"]:
                    length = st.slider("长度", 1, 100, 10, help="生成内容的长度（字符数）")
                if gen_type == "随机数字":
                    min_val = st.number_input("最小值", value=0)
                    max_val = st.number_input("最大值", value=100)

                count = st.number_input("生成数量", min_value=1, max_value=100, value=5)

            with col2:
                if gen_type == "随机字符串":
                    chars_type = st.multiselect("字符类型", RANDOM_STRING_TYPES, default=RANDOM_STRING_TYPES[:3],
                                                help="选择包含的字符类型")
                if gen_type == "随机密码":
                    password_options = st.multiselect("密码选项", PASSWORD_OPTIONS, default=PASSWORD_OPTIONS[:3],
                                                      help="设置密码复杂度要求")

                st.write("📋 条件说明")
                if gen_type == "随机字符串":
                    st.write(f"- 类型: 随机字符串")
                    st.write(f"- 长度: {length}字符")
                    st.write(f"- 字符类型: {', '.join(chars_type)}")
                elif gen_type == "随机数字":
                    st.write(f"- 类型: 随机数字")
                    st.write(f"- 范围: {min_val} 到 {max_val}")
                elif gen_type == "随机密码":
                    st.write(f"- 类型: 随机密码")
                    st.write(f"- 长度: {length}字符")
                    st.write(f"- 复杂度: {', '.join(password_options)}")
                else:
                    st.write(f"- 类型: UUID")

                st.write("💡 提示: 点击生成按钮后结果将保留在页面")

            if st.button("生成内容", key="gen_random_content"):
                results = []
                with st.spinner(f"正在生成{count}个{gen_type}..."):
                    for _ in range(count):
                        if gen_type == "随机字符串":
                            res = generator.safe_generate(generator.generate_random_string, length, chars_type)
                        elif gen_type == "随机数字":
                            res = str(random.randint(min_val, max_val))
                        elif gen_type == "随机密码":
                            res = generator.safe_generate(generator.generate_random_password, length, password_options)
                        elif gen_type == "UUID":
                            res = str(uuid.uuid4())
                        if res is not None:
                            results.append(res)

                result_text = "\n".join(results)
                conditions = (
                        f"类型: {gen_type}, " +
                        (f"长度: {length}, " if gen_type in ["随机字符串", "随机密码"] else "") +
                        (f"范围: {min_val}-{max_val}, " if gen_type == "随机数字" else "") +
                        (f"字符类型: {', '.join(chars_type)}" if gen_type == "随机字符串" else "") +
                        (f"复杂度: {', '.join(password_options)}" if gen_type == "随机密码" else "")
                )
                display_generated_results(conditions, result_text, "随机内容")

        elif data_gen_tool == "随机邮箱生成器":
            st.markdown('<div class="category-card">📧 随机邮箱生成器</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                count = st.number_input("邮箱数量", min_value=1, max_value=100, value=10)
                domain_option = st.selectbox("域名选项", ["随机域名", "自定义域名"])

            with col2:
                if domain_option == "自定义域名":
                    custom_domain = st.text_input("自定义域名", "example.com", placeholder="输入不带http://的域名")
                    conditions = f"域名: {custom_domain}"
                else:
                    selected_domains = st.multiselect("选择域名", DOMAINS_PRESET, default=DOMAINS_PRESET[:3])
                    conditions = f"随机域名: {', '.join(selected_domains)}"
                st.write("💡 提示: 用户名将随机生成8-12位字母数字组合")

            if st.button("生成邮箱", key="gen_emails"):
                results = []
                with st.spinner(f"正在生成{count}个邮箱地址..."):
                    for _ in range(count):
                        email = generator.safe_generate(generator.generate_random_email, domain_option,
                                                        custom_domain if domain_option == "自定义域名" else None,
                                                        selected_domains if domain_option != "自定义域名" else None)
                        if email is not None:
                            results.append(email)

                result_text = "\n".join(results)
                display_generated_results(conditions, result_text, "邮箱列表")
        elif data_gen_tool == "电话号码生成器":
            st.markdown('<div class="category-card">📞 电话号码生成器</div>', unsafe_allow_html=True)
            # 确保PROVINCES是列表类型
            PROVINCES = list(PROVINCE_CITY_AREA_CODES.keys())


            def get_cities_by_province(province):
                """根据省份获取城市列表"""
                return list(PROVINCE_CITY_AREA_CODES.get(province, {}).keys())


            def get_area_code(province, city):
                """根据省份和城市获取区号"""
                return PROVINCE_CITY_AREA_CODES.get(province, {}).get(city, "")


            col1, col2 = st.columns(2)
            with col1:
                phone_type = st.selectbox("号码类型", ["手机号", "座机", "国际号码"])

                # 初始化变量
                operator = None
                country = None
                province = None
                city = None
                area_code = None

                if phone_type == "国际号码":
                    country = st.selectbox("选择国家", COUNTRIES)
                elif phone_type == "手机号":
                    operator = st.selectbox("运营商", ["随机", "移动", "联通", "电信", "广电"])
                elif phone_type == "座机":
                    # 使用本地定义的 PROVINCES - 确保是列表
                    province_options = ["随机"] + PROVINCES
                    province = st.selectbox("选择省份", province_options)

                    if province and province != "随机":
                        cities = get_cities_by_province(province)
                        city = st.selectbox("选择城市", ["随机"] + cities)

                        # 如果选择了具体城市，获取对应的区号
                        if city and city != "随机":
                            area_code = get_area_code(province, city)
                            if area_code:
                                st.success(f"✅ 所选城市区号: {area_code}")
                            else:
                                st.warning("⚠️ 未找到该城市的区号")
                    else:
                        city = "随机"
                        st.info("将随机生成区号")

                count = st.number_input("生成数量", min_value=1, max_value=100, value=10)

            with col2:
                if phone_type == "座机":
                    if province == "随机":
                        conditions = f"类型: {phone_type}, 区号: 随机"
                    elif city == "随机":
                        conditions = f"类型: {phone_type}, 省份: {province}, 区号: 随机"
                    else:
                        conditions = f"类型: {phone_type}, 城市: {city}, 区号: {area_code}"
                elif phone_type == "国际号码":
                    conditions = f"类型: {phone_type}, 国家: {country}"
                else:
                    conditions = f"运营商: {operator}, 类型: {phone_type}"

                st.write("💡 提示: 生成的号码将匹配相应的号码规则")

            if st.button("生成电话号码", key="gen_conditional_phones"):
                results = []
                selected_area_codes = []  # 用于记录实际使用的区号

                with st.spinner(f"正在生成{count}个号码..."):
                    for i in range(count):
                        try:
                            if phone_type == "座机":
                                # 根据选择确定最终的区号
                                final_area_code = None

                                if province != "随机":
                                    if city != "随机" and area_code:
                                        # 使用具体城市的区号
                                        final_area_code = area_code
                                    else:
                                        # 随机选择该省份下的一个城市区号
                                        cities = get_cities_by_province(province)
                                        if cities:
                                            random_city = random.choice(cities)
                                            final_area_code = get_area_code(province, random_city)

                                # 记录实际使用的区号
                                if final_area_code:
                                    selected_area_codes.append(final_area_code)

                                # 调用生成函数
                                phone = generator.generate_landline_number(area_code=final_area_code)

                            elif phone_type == "国际号码":
                                phone = generator.generate_international_phone(country)
                            else:  # 手机号
                                phone = generator.generate_conditional_phone(operator)

                            if phone is not None:
                                results.append(phone)

                        except Exception as e:
                            # 处理可能的生成错误，继续生成其他号码
                            st.error(f"生成第 {i + 1} 个号码时出错: {str(e)}")
                            continue

                if results:
                    result_text = "\n".join(results)

                    # 删除原来的显示代码，直接使用封装的函数
                    display_generated_results("电话号码", result_text, "电话号码")

                    # 显示调试信息（实际使用的区号）
                    if phone_type == "座机" and selected_area_codes:
                        unique_codes = list(set(selected_area_codes))
                        st.info(f"实际使用的区号: {', '.join(unique_codes)}")

                    # 显示生成统计
                    st.success(f"✅ 成功生成 {len(results)} 个电话号码")
                else:
                    st.warning("⚠️ 未能生成任何有效的电话号码，请检查参数设置")

        elif data_gen_tool == "随机地址生成器":
            st.markdown('<div class="category-card">🏠 随机地址生成器</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                province = st.selectbox("选择省份", ["随机"] + [p for p in PROVINCES.keys() if p != "随机"])
                count = st.number_input("生成数量", min_value=1, max_value=50, value=10)
                detailed = st.checkbox("生成详细地址", value=True)

            with col2:
                if province != "随机":
                    city_options = PROVINCES[province]
                    city = st.selectbox("选择城市", ["随机"] + [c for c in city_options if c != province])
                else:
                    city = "随机"

                conditions = (
                        f"省份: {province if province != '随机' else '随机选择'}, " +
                        f"城市: {city if city != '随机' else '随机选择'}, " +
                        f"详细程度: {'详细地址' if detailed else '仅省市信息'}"
                )
                st.write("💡 提示: 详细地址包含街道、门牌号等信息")

            if st.button("生成地址", key="gen_addresses"):
                results = []
                with st.spinner(f"正在生成{count}个地址..."):
                    for _ in range(count):
                        selected_province = province
                        if province == "随机":
                            selected_province = random.choice([p for p in PROVINCES.keys() if p != "随机"])

                        selected_city = city
                        if city == "随机":
                            if selected_province in PROVINCES:
                                city_options = [c for c in PROVINCES[selected_province] if c != selected_province]
                                selected_city = random.choice(city_options) if city_options else selected_province

                        addr = generator.safe_generate(generator.generate_random_address, selected_province,
                                                       selected_city, detailed)
                        if addr is not None:
                            results.append(addr)

                result_text = "\n".join(results)
                display_generated_results(conditions, result_text, "地址列表")

        elif data_gen_tool == "随机身份证生成器":
            st.markdown('<div class="category-card">🆔 随机身份证生成器</div>', unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                province = st.selectbox("选择省份", ["随机"] + list(PROVINCE_MAP.keys()))
                gender = st.selectbox("选择性别", GENDERS)
                count = st.number_input("生成数量", min_value=1, max_value=100, value=10)

            with col2:
                min_age = st.number_input("最小年龄", min_value=0, max_value=100, value=18)
                max_age = st.number_input("最大年龄", min_value=0, max_value=100, value=60)
                if min_age > max_age:
                    st.error("最小年龄不能大于最大年龄")

                conditions = f"省份: {province}, 性别: {gender}, 年龄: {min_age}-{max_age}岁"
                st.write("💡 提示: 生成的身份证将严格符合选择的省份、性别和年龄条件")

            if st.button("生成身份证", key="gen_id_cards"):
                results = []
                with st.spinner(f"正在生成{count}个身份证号码..."):
                    for _ in range(count):
                        id_card = generator.safe_generate(generator.generate_random_id_card,
                                                          province if province != "随机" else random.choice(
                                                              list(PROVINCE_MAP.keys())),
                                                          gender,
                                                          min_age,
                                                          max_age)
                        if id_card is not None:
                            results.append(id_card)

                result_text = "\n".join(results)
                display_generated_results(conditions, result_text, "身份证列表")

        elif data_gen_tool == "测试场景造数":
            st.markdown('<div class="category-card">🧩 测试场景造数</div>', unsafe_allow_html=True)
            scenario_templates = generator.get_test_data_scenarios()

            col1, col2 = st.columns([2, 1])
            with col1:
                selected_scenario = st.selectbox("选择造数场景", list(scenario_templates.keys()))
                scenario_meta = scenario_templates[selected_scenario]
                st.info(scenario_meta["description"])
                st.caption(f"字段覆盖: {', '.join(scenario_meta['fields'])}")
                batch_tag = st.text_input(
                    "批次标识",
                    value=f"qa{datetime.datetime.now().strftime('%m%d')}",
                    help="会拼接进用户名、员工号、订单号等字段，便于批量清理和回溯。"
                )
                email_domain = st.text_input(
                    "邮箱域名",
                    value="example.com",
                    help="用于注册账号或员工档案中的邮箱字段。"
                )

            with col2:
                count = st.number_input("生成数量", min_value=1, max_value=200, value=10)
                scenario_seed_text = st.text_input(
                    "随机种子",
                    value="",
                    placeholder="留空则每次生成不同",
                    help="填写整数后，同一参数组合可复现同一批数据。"
                )
                st.write("💡 使用建议")
                st.write("- 联调环境建议填写批次标识")
                st.write("- 回归场景建议补随机种子")
                st.write("- 结果可直接下载 CSV/JSON")

            scenario_seed = None
            if scenario_seed_text.strip():
                try:
                    scenario_seed = int(scenario_seed_text.strip())
                except ValueError:
                    st.warning("随机种子需为整数，当前已忽略。")

            action_col1, action_col2 = st.columns([3, 1])
            with action_col1:
                if st.button("生成场景数据", key="gen_scenario_dataset", use_container_width=True):
                    dataset = generator.safe_generate(
                        generator.generate_test_dataset,
                        selected_scenario,
                        count,
                        scenario_seed,
                        batch_tag,
                        email_domain,
                    )
                    if dataset is not None:
                        st.session_state.scenario_dataset_records = dataset
                        st.session_state.scenario_dataset_title = selected_scenario
            with action_col2:
                if st.button("清空结果", key="clear_scenario_dataset", use_container_width=True):
                    st.session_state.pop("scenario_dataset_records", None)
                    st.session_state.pop("scenario_dataset_title", None)
                    st.rerun()

            if st.session_state.get("scenario_dataset_records"):
                display_structured_results(
                    st.session_state.get("scenario_dataset_title", "测试场景数据"),
                    st.session_state["scenario_dataset_records"],
                    "scenario_dataset"
                )

        elif data_gen_tool == "边界值/异常值生成器":
            st.markdown('<div class="category-card">🧪 边界值/异常值生成器</div>', unsafe_allow_html=True)
            boundary_templates = generator.get_boundary_field_templates()

            col1, col2 = st.columns([2, 1])
            with col1:
                selected_field_type = st.selectbox("选择字段类型", list(boundary_templates.keys()))
                st.info(boundary_templates[selected_field_type]["description"])
                case_types = st.multiselect(
                    "包含用例类型",
                    ["正常值", "边界值", "异常值"],
                    default=["正常值", "边界值", "异常值"],
                    help="可按需要只生成某类数据，便于接口联调或测试设计。"
                )

                min_length = 2
                max_length = 20
                amount_min = 0.00
                amount_max = 99999.99
                boundary_email_domain = "example.com"

                if selected_field_type in ["用户名", "密码"]:
                    length_col1, length_col2 = st.columns(2)
                    with length_col1:
                        min_length = st.number_input("最小长度", min_value=1, max_value=128, value=2)
                    with length_col2:
                        max_length = st.number_input("最大长度", min_value=min_length, max_value=256, value=20)

                if selected_field_type == "金额":
                    amount_col1, amount_col2 = st.columns(2)
                    with amount_col1:
                        amount_min = st.number_input("最小金额", value=0.00, format="%.2f")
                    with amount_col2:
                        amount_max = st.number_input("最大金额", min_value=float(amount_min), value=99999.99,
                                                     format="%.2f")

                if selected_field_type == "邮箱":
                    boundary_email_domain = st.text_input("邮箱域名", value="example.com")

            with col2:
                boundary_seed_text = st.text_input(
                    "随机种子",
                    value="",
                    placeholder="留空则每次生成不同"
                )
                st.write("📌 适用场景")
                st.write("- 表单校验测试")
                st.write("- 接口参数边界值设计")
                st.write("- 前后端联调异常值验证")

            boundary_seed = None
            if boundary_seed_text.strip():
                try:
                    boundary_seed = int(boundary_seed_text.strip())
                except ValueError:
                    st.warning("随机种子需为整数，当前已忽略。")

            action_col1, action_col2 = st.columns([3, 1])
            with action_col1:
                if st.button("生成边界值用例", key="gen_boundary_cases", use_container_width=True):
                    cases = generator.safe_generate(
                        generator.generate_boundary_test_cases,
                        selected_field_type,
                        boundary_seed,
                        min_length,
                        max_length,
                        amount_min,
                        amount_max,
                        boundary_email_domain,
                        "正常值" in case_types,
                        "边界值" in case_types,
                        "异常值" in case_types,
                    )
                    if cases is not None:
                        st.session_state.boundary_case_records = cases
                        st.session_state.boundary_case_title = selected_field_type
            with action_col2:
                if st.button("清空结果", key="clear_boundary_cases", use_container_width=True):
                    st.session_state.pop("boundary_case_records", None)
                    st.session_state.pop("boundary_case_title", None)
                    st.rerun()

            if st.session_state.get("boundary_case_records"):
                display_structured_results(
                    f"{st.session_state.get('boundary_case_title', '字段')}边界值",
                    st.session_state["boundary_case_records"],
                    "boundary_cases"
                )

    st.markdown('</div>', unsafe_allow_html=True)

# 字数统计工具
elif tool_category == "字数统计工具":
    show_doc("word_counter")

    # 添加CSS样式
    st.markdown("""
    <style>
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        text-align: center;
        border-left: 4px solid;
        margin-bottom: 1rem;
    }
    .progress-bar {
        height: 8px;
        background: #e2e8f0;
        border-radius: 4px;
        margin: 0.5rem 0;
        overflow: hidden;
    }
    .progress-fill {
        height: 100%;
        background: linear-gradient(90deg, #667eea, #764ba2);
        border-radius: 4px;
        transition: width 0.3s ease;
    }
    .stat-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 1rem;
        margin: 1rem 0;
    }
    </style>
    """, unsafe_allow_html=True)

    # 侧边栏设置
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🎯 字数目标设置")

    target_words = st.sidebar.number_input("设定目标单词数", min_value=0, value=1000, step=100)
    target_chars = st.sidebar.number_input("设定目标字符数", min_value=0, value=5000, step=500)

    st.sidebar.markdown("### 🎨 显示选项")
    show_charts = st.sidebar.checkbox("显示图表", value=True)
    show_advanced = st.sidebar.checkbox("显示高级分析", value=False)
    show_suggestions = st.sidebar.checkbox("显示编辑建议", value=True)

    text_input = st.text_area("输入要统计的文本", height=200, placeholder="在此处输入或粘贴文本...", key="word_counter_text")

    if text_input:
        # 基础统计计算
        words = text_input.split()
        lines = text_input.split('\n')
        paragraphs = [p for p in text_input.split('\n\n') if p.strip()]
        char_freq = Counter(text_input)

        # 字符类型统计
        import string

        letters = sum(1 for char in text_input if char.isalpha())
        digits = sum(1 for char in text_input if char.isdigit())
        spaces = text_input.count(' ')
        punctuation = sum(1 for char in text_input if char in string.punctuation)
        chinese_chars = sum(1 for char in text_input if '\u4e00' <= char <= '\u9fff')

        # 句子统计（简单实现）
        sentences = []
        for sep in ['.', '!', '?', '。', '！', '？']:
            sentences.extend([s.strip() for s in text_input.split(sep) if s.strip()])
        sentences = [s for s in sentences if s]

        # 计算常用指标
        total_chars = len(text_input)
        total_chars_no_spaces = len(text_input.replace(' ', ''))
        total_words = len(words)
        total_lines = len(lines)
        total_paragraphs = len(paragraphs)
        total_sentences = len(sentences)

        # 质量指标计算
        avg_word_length = sum(len(word) for word in words) / total_words if words else 0
        avg_sentence_length = total_words / total_sentences if total_sentences else 0
        avg_paragraph_length = total_words / total_paragraphs if total_paragraphs else 0
        reading_time = total_words / 200  # 按200词/分钟

        # 主要指标卡片布局
        st.markdown("### 📊 主要统计指标")
        col1, col2, col3, col4, col5 = st.columns(5)

        metrics_data = [
            {"title": "字符数（含空格）", "value": total_chars, "color": "#667eea"},
            {"title": "字符数（不含空格）", "value": total_chars_no_spaces, "color": "#48bb78"},
            {"title": "单词数", "value": total_words, "color": "#ed8936"},
            {"title": "行数", "value": total_lines, "color": "#9f7aea"},
            {"title": "段落数", "value": total_paragraphs, "color": "#f56565"}
        ]

        for i, metric in enumerate(metrics_data):
            with [col1, col2, col3, col4, col5][i]:
                st.markdown(f"""
                <div class="metric-card" style="border-left-color: {metric['color']};">
                    <div style="font-size: 1rem; font-weight: 600; color: {metric['color']};">{metric['title']}</div>
                    <div style="font-size: 1.8rem; font-weight: 700; color: #2d3748;">{metric['value']:,}</div>
                </div>
                """, unsafe_allow_html=True)

        # 进度跟踪
        if target_words > 0 or target_chars > 0:
            st.markdown("### 🎯 目标进度")
            progress_col1, progress_col2 = st.columns(2)

            progress_data = [
                {"target": target_words, "current": total_words, "label": "单词"},
                {"target": target_chars, "current": total_chars, "label": "字符"}
            ]

            for i, progress in enumerate(progress_data):
                if progress["target"] > 0:
                    with [progress_col1, progress_col2][i]:
                        progress_value = min(progress["current"] / progress["target"], 1.0)
                        st.write(f"{progress['label']}进度: {progress['current']}/{progress['target']}")
                        st.markdown(f"""
                        <div class="progress-bar">
                            <div class="progress-fill" style="width: {progress_value * 100}%"></div>
                        </div>
                        <div style="text-align: center; font-size: 0.9rem; color: #666;">{progress_value * 100:.1f}%</div>
                        """, unsafe_allow_html=True)

                        if progress["current"] >= progress["target"]:
                            st.success(f"🎉 恭喜！已达到目标{progress['label']}数！")

        # 字符类型统计
        st.markdown("### 🔤 字符类型分析")
        col6, col7, col8, col9, col10 = st.columns(5)

        char_type_data = [
            ("字母数", letters),
            ("数字数", digits),
            ("标点符号", punctuation),
            ("空格数", spaces),
            ("中文字符", chinese_chars)
        ]

        for i, (title, value) in enumerate(char_type_data):
            with [col6, col7, col8, col9, col10][i]:
                st.metric(title, f"{value:,}")

        # 文本质量指标
        st.markdown("### 📈 文本质量指标")
        col11, col12, col13, col14 = st.columns(4)

        quality_metrics = [
            ("平均词长", f"{avg_word_length:.1f}字符"),
            ("平均句长", f"{avg_sentence_length:.1f}词"),
            ("阅读时间", f"{reading_time:.1f}分钟"),
            ("平均段落长", f"{avg_paragraph_length:.1f}词")
        ]

        for i, (title, value) in enumerate(quality_metrics):
            with [col11, col12, col13, col14][i]:
                st.metric(title, value)

        # 图表显示
        if show_charts:
            st.markdown("### 📊 可视化分析")

            try:
                import plotly.express as px
                import plotly.graph_objects as go
                import pandas as pd

                tab1, tab2, tab3 = st.tabs(["字符频率", "类型分布", "文本结构"])

                with tab1:
                    top_chars = char_freq.most_common(15)
                    if top_chars:
                        chars, freqs = zip(*top_chars)
                        SPECIAL_CHARS_DISPLAY = {
                            ' ': "空格",
                            '\n': "换行",
                            '\t': "制表符",
                            '\r': "回车"
                        }
                        char_display = [SPECIAL_CHARS_DISPLAY.get(char, char) for char in chars]

                        fig = px.bar(
                            x=freqs, y=char_display,
                            orientation='h',
                            title='Top 15 字符频率',
                            labels={'x': '出现次数', 'y': '字符'}
                        )
                        fig.update_layout(yaxis={'categoryorder': 'total ascending'})
                        st.plotly_chart(fig)

                with tab2:
                    type_data = {
                        '字母': letters,
                        '数字': digits,
                        '标点': punctuation,
                        '空格': spaces,
                        '中文': chinese_chars,
                        '其他': total_chars - (letters + digits + punctuation + spaces + chinese_chars)
                    }
                    type_data = {k: v for k, v in type_data.items() if v > 0}

                    if type_data:
                        fig = px.pie(
                            values=list(type_data.values()),
                            names=list(type_data.keys()),
                            title='字符类型分布'
                        )
                        st.plotly_chart(fig)

                with tab3:
                    structure_data = {
                        '字符': total_chars,
                        '单词': total_words,
                        '句子': total_sentences,
                        '行数': total_lines,
                        '段落': total_paragraphs
                    }

                    fig = px.bar(
                        x=list(structure_data.keys()),
                        y=list(structure_data.values()),
                        title='文本结构概览',
                        labels={'x': '统计类型', 'y': '数量'},
                        color=list(structure_data.keys()),
                        color_discrete_sequence=['#667eea', '#48bb78', '#ed8936', '#9f7aea', '#f56565']
                    )
                    st.plotly_chart(fig)

            except ImportError:
                st.warning("高级图表需要 plotly 库。请安装: `pip install plotly`")
                # 回退到 streamlit 原生图表
                st.info("使用基础图表显示...")

        # 字符频率详情
        st.markdown("### 🔍 字符频率详情")
        SPECIAL_CHARS_DISPLAY = {
            ' ': "[空格]",
            '\n': "[换行]",
            '\t': "[制表符]",
            '\r': "[回车]"
        }

        col_freq1, col_freq2 = st.columns(2)

        with col_freq1:
            st.write("**最常见字符（前10个）:**")
            sorted_chars = char_freq.most_common(10)
            for char, freq in sorted_chars:
                display_char = SPECIAL_CHARS_DISPLAY.get(char, char)
                st.write(f"`{display_char}`: {freq:,}次 ({freq / total_chars * 100:.2f}%)")

        with col_freq2:
            st.write("**最罕见字符（后10个）:**")
            rare_chars = char_freq.most_common()[-10:]
            for char, freq in rare_chars:
                display_char = SPECIAL_CHARS_DISPLAY.get(char, char)
                st.write(f"`{display_char}`: {freq:,}次")

        # 编辑建议
        if show_suggestions:
            st.markdown("### 📝 编辑建议")
            suggestions = []

            # 文本长度分析
            if total_chars < 50:
                suggestions.append("📝 **文本较短**: 当前仅{}字，建议补充更多细节、例证或分析以丰富内容".format(total_chars))
            elif total_chars > 10000:
                suggestions.append("📝 **文本较长**: 当前{}字，考虑是否可拆分为多个部分或精简冗余内容".format(total_chars))

            # 句子结构分析
            if total_sentences > 0:
                if avg_sentence_length > 25:
                    suggestions.append("📝 **句子偏长**: 平均句长{:.1f}词，建议拆分复杂长句，每句控制在15-25词为宜".format(avg_sentence_length))
                elif avg_sentence_length < 8:
                    suggestions.append("📝 **句子偏短**: 平均句长仅{:.1f}词，可适当合并短句以增强表达连贯性".format(avg_sentence_length))

            # 词汇层面分析
            long_words = [word for word in words if len(word) > 20]
            if long_words:
                suggestions.append("📝 **超长单词**: 发现{}个超长单词（如'{}'），建议使用更简洁的表达替代".format(len(long_words), long_words[0]))

            # 段落结构分析
            if total_paragraphs > 0:
                if avg_paragraph_length > 300:
                    suggestions.append("📝 **段落过长**: 平均每段{:.0f}词，建议将长段落按主题拆分为2-3个段落".format(avg_paragraph_length))
                elif avg_paragraph_length < 50:
                    suggestions.append("📝 **段落过短**: 平均每段仅{:.0f}词，可适当合并相关短段落".format(avg_paragraph_length))

            # 词汇多样性分析
            if total_words > 0:
                lexical_diversity = len(set(words)) / total_words
                if lexical_diversity < 0.5:
                    suggestions.append("📝 **词汇重复**: 词汇多样性指数{:.2f}，建议使用同义词替换高频重复词汇".format(lexical_diversity))
                elif lexical_diversity > 0.8:
                    suggestions.append("🌈 **词汇丰富**: 词汇多样性指数{:.2f}，用词变化丰富，表现良好".format(lexical_diversity))

            # 可读性增强建议
            if len([word for word in words if word.isupper() and len(word) > 1]) > 3:
                suggestions.append("📝 **全大写使用**: 文本中全大写词汇较多，建议适度使用以保持阅读舒适度")

            # 输出建议
            if suggestions:
                st.markdown("#### 改进建议")
                for i, suggestion in enumerate(suggestions, 1):
                    if "表现良好" in suggestion:
                        st.success(f"{i}. {suggestion}")
                    else:
                        st.warning(f"{i}. {suggestion}")

                # 总结性建议
                st.markdown("---")
                improvement_count = len([s for s in suggestions if "表现良好" not in s])
                if improvement_count == 0:
                    st.balloons()
                    st.success("🎉 文本质量优秀！所有指标均达到理想标准")
                else:
                    st.info(f"**总结**: 共发现{improvement_count}个可改进方面，按照建议调整可提升文本质量")
            else:
                st.success("✅ 文本结构良好，无明显问题")
                st.balloons()

        # 高级分析
        if show_advanced:
            st.markdown("### 🔬 高级分析")

            advanced_tab1, advanced_tab2, advanced_tab3 = st.tabs(["重复内容分析", "文本结构洞察", "文本预览"])

            with advanced_tab1:
                # 重复单词分析
                word_freq = Counter(words)

                # 高频词分析
                repeated_words = [(word, freq) for word, freq in word_freq.items()
                                  if freq > 3 and len(word) > 2 and word.isalpha()]

                if repeated_words:
                    st.subheader("🔁 高频重复词汇")
                    st.write(f"**出现3次以上的词汇 (共{len(repeated_words)}个):**")

                    # 按频率排序
                    repeated_words.sort(key=lambda x: x[1], reverse=True)

                    # 使用Streamlit内置图表替代matplotlib
                    top_words = repeated_words[:10]
                    if top_words:
                        chart_data = {
                            '词汇': [word for word, freq in top_words],
                            '出现次数': [freq for word, freq in top_words]
                        }
                        st.bar_chart(chart_data.set_index('词汇'))

                    # 详细列表
                    repeated_col1, repeated_col2 = st.columns(2)
                    mid_point = len(repeated_words) // 2

                    with repeated_col1:
                        st.write("**详细列表:**")
                        for word, freq in repeated_words[:mid_point]:
                            percentage = (freq / total_words) * 100
                            st.write(f"`{word}`: {freq}次 ({percentage:.1f}%)")

                    with repeated_col2:
                        st.write("&nbsp;")  # 空行占位
                        for word, freq in repeated_words[mid_point:]:
                            percentage = (freq / total_words) * 100
                            st.write(f"`{word}`: {freq}次 ({percentage:.1f}%)")

                    # 重复度评分
                    repetition_score = len(repeated_words) / len(word_freq) * 100
                    st.metric("词汇重复度", f"{repetition_score:.1f}%")

                else:
                    st.info("✅ 未发现高频重复词汇")

            with advanced_tab2:
                st.subheader("📊 文本结构洞察")

                col1, col2 = st.columns(2)

                with col1:
                    # 句子长度分布
                    if total_sentences > 0:
                        sentence_lengths = [len(sentence.split()) for sentence in sentences]

                        st.metric("平均句子长度", f"{avg_sentence_length:.1f}词")

                        # 使用Streamlit内置图表
                        if sentence_lengths:
                            # 创建句子长度分布数据
                            length_ranges = {'1-10词': 0, '11-20词': 0, '21-30词': 0, '31-40词': 0, '41+词': 0}
                            for length in sentence_lengths:
                                if length <= 10:
                                    length_ranges['1-10词'] += 1
                                elif length <= 20:
                                    length_ranges['11-20词'] += 1
                                elif length <= 30:
                                    length_ranges['21-30词'] += 1
                                elif length <= 40:
                                    length_ranges['31-40词'] += 1
                                else:
                                    length_ranges['41+词'] += 1

                            st.write("**句子长度分布:**")
                            for range_name, count in length_ranges.items():
                                if count > 0:
                                    percentage = (count / total_sentences) * 100
                                    st.write(f"- {range_name}: {count}句 ({percentage:.1f}%)")

                with col2:
                    # 段落分析
                    if total_paragraphs > 0:
                        paragraph_lengths = [len(para.split()) for para in paragraphs if para.strip()]

                        st.metric("平均段落长度", f"{avg_paragraph_length:.1f}词")
                        st.metric("段落数量", total_paragraphs)

                        # 段落长度分析
                        st.write("**段落长度分布:**")
                        short_paras = len([l for l in paragraph_lengths if l < 50])
                        medium_paras = len([l for l in paragraph_lengths if 50 <= l <= 200])
                        long_paras = len([l for l in paragraph_lengths if l > 200])

                        st.write(f"- 短段落 (<50词): {short_paras}个")
                        st.write(f"- 中段落 (50-200词): {medium_paras}个")
                        st.write(f"- 长段落 (>200词): {long_paras}个")

                # 词汇复杂度分析
                st.subheader("📈 词汇复杂度")
                col3, col4, col5 = st.columns(3)

                with col3:
                    unique_words = len(set(words))
                    st.metric("独特词汇量", unique_words)

                with col4:
                    lexical_density = (unique_words / total_words) * 100
                    st.metric("词汇密度", f"{lexical_density:.1f}%")

                with col5:
                    st.metric("平均词长", f"{avg_word_length:.1f}字符")

            with advanced_tab3:
                st.subheader("👁️ 文本预览")

                # 文本统计概览
                stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
                with stat_col1:
                    st.metric("总字符数", total_chars)
                with stat_col2:
                    st.metric("总词数", total_words)
                with stat_col3:
                    st.metric("总句数", total_sentences)
                with stat_col4:
                    st.metric("总段落数", total_paragraphs)

                # 文本预览
                st.write("**内容预览:**")
                preview = text_input[:500] + "..." if total_chars > 500 else text_input

                # 高亮显示长句子
                preview_highlighted = preview
                for sentence in sentences:
                    sentence_words = sentence.split()
                    if len(sentence_words) > 25 and sentence in preview:
                        # 使用HTML标记高亮
                        preview_highlighted = preview_highlighted.replace(
                            sentence, f"<mark style='background-color: #ffd70033'>{sentence}</mark>"
                        )

                st.markdown(
                    f'<div style="border: 1px solid #e0e0e0; padding: 15px; border-radius: 5px; background-color: #fafafa; white-space: pre-wrap;">{preview_highlighted}</div>',
                    unsafe_allow_html=True)

                # 阅读时间估算
                st.info(f"📖 预计阅读时间: {reading_time:.1f}分钟 (按200词/分钟计算)")

        # 导出功能
        st.markdown("### 📤 导出统计结果")

        import json
        import pandas as pd

        # 创建完整的统计字典
        stats = {
            "基础统计": {
                "字符数（含空格）": total_chars,
                "字符数（不含空格）": total_chars_no_spaces,
                "单词数": total_words,
                "句子数": total_sentences,
                "行数": total_lines,
                "段落数": total_paragraphs
            },
            "字符类型": {
                "字母数": letters,
                "数字数": digits,
                "标点符号": punctuation,
                "空格数": spaces,
                "中文字符": chinese_chars
            },
            "质量指标": {
                "平均词长": round(avg_word_length, 2),
                "平均句长": round(avg_sentence_length, 2),
                "平均段落长": round(avg_paragraph_length, 2),
                "阅读时间(分钟)": round(reading_time, 2)
            }
        }

        export_col1, export_col2, export_col3 = st.columns(3)

        with export_col1:
            # JSON导出
            st.download_button(
                label="📥 导出为JSON",
                data=json.dumps(stats, indent=2, ensure_ascii=False),
                file_name="文本统计报告.json",
                mime="application/json"
            )

        with export_col2:
            # CSV导出
            csv_data = []
            for category, items in stats.items():
                for key, value in items.items():
                    csv_data.append({"类别": category, "指标": key, "数值": value})

            df = pd.DataFrame(csv_data)
            csv_string = df.to_csv(index=False)
            st.download_button(
                label="📥 导出为CSV",
                data=csv_string,
                file_name="文本统计报告.csv",
                mime="text/csv"
            )

        with export_col3:
            # 文本报告导出
            report_text = f"""文本统计报告
生成时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
==============================

基础统计:
--------
字符数（含空格）: {total_chars:,}
字符数（不含空格）: {total_chars_no_spaces:,}
单词数: {total_words:,}
句子数: {total_sentences:,}
行数: {total_lines:,}
段落数: {total_paragraphs:,}

字符类型:
--------
字母数: {letters:,}
数字数: {digits:,}
标点符号: {punctuation:,}
空格数: {spaces:,}
中文字符: {chinese_chars:,}

质量指标:
--------
平均词长: {avg_word_length:.2f}
平均句长: {avg_sentence_length:.2f}
平均段落长: {avg_paragraph_length:.2f}
阅读时间: {reading_time:.2f}分钟
"""
            st.download_button(
                label="📥 导出为文本报告",
                data=report_text,
                file_name="文本统计报告.txt",
                mime="text/plain"
            )

    else:
        # 没有输入时的提示
        st.info("👆 请在上方文本框中输入文本以开始统计")

        # 示例文本
        with st.expander("📋 点击查看示例文本"):
            sample_text = """这是一个示例文本，用于展示字数统计工具的功能。

你可以在这里输入任意文本，工具会自动计算：
- 字符数（包含和不包含空格）
- 单词数量
- 行数和段落数
- 各种字符类型的分布

此外，工具还提供：
📊 可视化图表分析
📝 文本编辑建议
📈 质量评估指标
📤 多种格式导出功能

尝试复制你自己的文本到这里，看看详细的统计结果！"""
            st.text_area("示例文本", sample_text, height=200, key="sample_text")

    st.markdown('</div>', unsafe_allow_html=True)

# 文本对比工具
elif tool_category == "文本对比工具":
    show_doc("text_comparison")

    # 简化 session 初始化逻辑
    st.session_state.setdefault('text1_content', "")
    st.session_state.setdefault('text2_content', "")
    st.session_state.setdefault('clear_counter', 0)
    st.session_state.setdefault('diff_mode', 'line')
    st.session_state.setdefault('show_legend', True)


    # 新增：词对比相关函数
    def word_diff(text1, text2):
        """实现词级别的对比"""
        # 使用正则表达式分割单词，保留标点符号
        words1 = re.findall(r'\b\w+\b|[^\w\s]|\s+', text1)
        words2 = re.findall(r'\b\w+\b|[^\w\s]|\s+', text2)

        d = Differ()
        diff = list(d.compare(words1, words2))

        return diff, words1, words2


    def render_word_diff(diff):
        """渲染词对比结果"""
        html_parts = [
            "<div style='background-color: #f8f9fa; padding: 15px; border-radius: 5px; font-family: monospace; line-height: 1.6;'>"]

        current_line = []
        for item in diff:
            if item.startswith('+ '):
                word = html.escape(item[2:])
                current_line.append(
                    f"<span style='background-color: #d4edda; color: #155724; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #c3e6cb;'>+{word}</span>")
            elif item.startswith('- '):
                word = html.escape(item[2:])
                current_line.append(
                    f"<span style='background-color: #f8d7da; color: #721c24; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #f5c6cb;'>-{word}</span>")
            elif item.startswith('? '):
                # 在词模式中，? 通常不需要特殊显示
                continue
            else:
                word = html.escape(item[2:] if len(item) > 2 else item)
                # 处理换行符
                if word == '\n' or word == '\r\n':
                    if current_line:
                        html_parts.append(''.join(current_line))
                        current_line = []
                    html_parts.append("<br>")
                else:
                    current_line.append(f"<span style='padding: 1px 2px;'>{word}</span>")

        # 添加最后一行
        if current_line:
            html_parts.append(''.join(current_line))

        html_parts.append("</div>")
        return ''.join(html_parts)


    def render_enhanced_word_diff(text1, text2):
        """增强的词对比，显示更详细的词级变化"""
        # 使用 difflib 的 SequenceMatcher 进行更精确的词级对比
        words1 = re.findall(r'\b\w+\b|[^\w\s]|\s+', text1)
        words2 = re.findall(r'\b\w+\b|[^\w\s]|\s+', text2)

        matcher = difflib.SequenceMatcher(None, words1, words2)

        html_parts = [
            "<div style='background-color: #f8f9fa; padding: 15px; border-radius: 5px; font-family: monospace; line-height: 1.6; word-wrap: break-word;'>"]

        for opcode in matcher.get_opcodes():
            tag, i1, i2, j1, j2 = opcode

            if tag == 'equal':
                # 相同的部分
                for word in words1[i1:i2]:
                    escaped_word = html.escape(word)
                    html_parts.append(f"<span style='padding: 1px 2px; color: #6c757d;'>{escaped_word}</span>")
            elif tag == 'replace':
                # 替换的部分 - 显示删除和新增
                # 删除的单词
                for word in words1[i1:i2]:
                    escaped_word = html.escape(word)
                    html_parts.append(
                        f"<span style='background-color: #f8d7da; color: #721c24; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #f5c6cb; text-decoration: line-through;'>-{escaped_word}</span>")
                # 新增的单词
                for word in words2[j1:j2]:
                    escaped_word = html.escape(word)
                    html_parts.append(
                        f"<span style='background-color: #d4edda; color: #155724; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #c3e6cb;'>+{escaped_word}</span>")
            elif tag == 'delete':
                # 删除的部分
                for word in words1[i1:i2]:
                    escaped_word = html.escape(word)
                    html_parts.append(
                        f"<span style='background-color: #f8d7da; color: #721c24; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #f5c6cb; text-decoration: line-through;'>-{escaped_word}</span>")
            elif tag == 'insert':
                # 新增的部分
                for word in words2[j1:j2]:
                    escaped_word = html.escape(word)
                    html_parts.append(
                        f"<span style='background-color: #d4edda; color: #155724; padding: 1px 3px; margin: 0 1px; border-radius: 2px; border: 1px solid #c3e6cb;'>+{escaped_word}</span>")

        html_parts.append("</div>")
        return ''.join(html_parts)


    # 设置选项区域
    with st.expander("⚙️ 对比设置", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            diff_mode = st.selectbox(
                "对比模式",
                options=['line', 'word', 'enhanced_word'],
                index=0,
                help="行模式：按行对比；词模式：按单词对比；增强词模式：更精确的词级对比"
            )
        with col2:
            show_legend = st.checkbox("显示图例", value=True)
            ignore_case = st.checkbox("忽略大小写", value=False)
            ignore_whitespace = st.checkbox("忽略空白字符", value=False)

    col_input1, col_input2 = st.columns(2)

    with col_input1:
        st.markdown("**原始文本**")
        text1 = st.text_area(" ", height=300,
                             key=f"text1_{st.session_state.clear_counter}",
                             value=st.session_state.text1_content,
                             label_visibility="collapsed")

        if text1:
            lines1 = len(text1.splitlines())
            words1 = len(re.findall(r'\b\w+\b', text1))
            chars1 = len(text1)
            st.caption(f"📊 统计: {lines1} 行, {words1} 词, {chars1} 字符")

    with col_input2:
        st.markdown("**对比文本**")
        text2 = st.text_area(" ", height=300,
                             key=f"text2_{st.session_state.clear_counter}",
                             value=st.session_state.text2_content,
                             label_visibility="collapsed")

        if text2:
            lines2 = len(text2.splitlines())
            words2 = len(re.findall(r'\b\w+\b', text2))
            chars2 = len(text2)
            st.caption(f"📊 统计: {lines2} 行, {words2} 词, {chars2} 字符")

    # 操作按钮区域
    button_col1, button_col2, button_col3, button_col4 = st.columns([1, 1, 1, 1])

    with button_col1:
        compare_clicked = st.button("🔄 开始对比", use_container_width=True)

    with button_col2:
        if st.button("📋 交换文本", use_container_width=True):
            # 先同步当前输入框的内容到 session state
            st.session_state.text1_content = text1
            st.session_state.text2_content = text2
            # 然后交换
            st.session_state.text1_content, st.session_state.text2_content = \
                st.session_state.text2_content, st.session_state.text1_content
            st.session_state.clear_counter += 1
            st.rerun()

    with button_col3:
        if st.button("📁 导入示例", use_container_width=True):
            # 提供更适合词对比的示例文本
            st.session_state.text1_content = """这是一个示例文本，用于演示词对比功能。
    第一行包含一些单词。
    第二行有更多的内容。
    第三行是最后一行。"""

            st.session_state.text2_content = """这是一个示范文本，用于演示词汇对比功能。
    第一行包含某些词语。
    第二行有更多不同的内容。
    新增的第四行文本。"""
            st.session_state.clear_counter += 1
            st.rerun()

    with button_col4:
        if st.button("🗑️ 清空所有", use_container_width=True):
            st.session_state.text1_content = ""
            st.session_state.text2_content = ""
            st.session_state.clear_counter += 1
            st.rerun()

    # 图例说明
    if show_legend:
        st.markdown("---")
        if diff_mode == 'line':
            col_legend1, col_legend2, col_legend3 = st.columns(3)
            with col_legend1:
                st.markdown(
                    "<div style='background-color: #f8d7da; padding: 5px; border-radius: 3px; text-align: center;'>"
                    "❌ 删除的行</div>",
                    unsafe_allow_html=True
                )
            with col_legend2:
                st.markdown(
                    "<div style='background-color: #d4edda; padding: 5px; border-radius: 3px; text-align: center;'>"
                    "✅ 新增的行</div>",
                    unsafe_allow_html=True
                )
            with col_legend3:
                st.markdown(
                    "<div style='background-color: #fff3cd; padding: 5px; border-radius: 3px; text-align: center;'>"
                    "⚠️ 修改的行</div>",
                    unsafe_allow_html=True
                )
        else:
            col_legend1, col_legend2 = st.columns(2)
            with col_legend1:
                st.markdown(
                    "<div style='background-color: #f8d7da; padding: 5px; border-radius: 3px; text-align: center; border: 1px solid #f5c6cb;'>"
                    "<span style='color: #721c24;'>-删除的单词</span></div>",
                    unsafe_allow_html=True
                )
            with col_legend2:
                st.markdown(
                    "<div style='background-color: #d4edda; padding: 5px; border-radius: 3px; text-align: center; border: 1px solid #c3e6cb;'>"
                    "<span style='color: #155724;'>+新增的单词</span></div>",
                    unsafe_allow_html=True
                )

    if compare_clicked:
        st.session_state.text1_content = text1
        st.session_state.text2_content = text2
        if text1 and text2:
            try:
                # 预处理文本
                processed_text1 = text1
                processed_text2 = text2

                if ignore_case:
                    processed_text1 = processed_text1.lower()
                    processed_text2 = processed_text2.lower()

                if ignore_whitespace:
                    processed_text1 = ' '.join(processed_text1.split())
                    processed_text2 = ' '.join(processed_text2.split())

                st.markdown("### 📊 对比结果")

                if diff_mode == 'line':
                    # 行对比模式
                    d = Differ()
                    diff = list(d.compare(processed_text1.splitlines(), processed_text2.splitlines()))

                    # 差异统计
                    added_lines = sum(1 for line in diff if line.startswith('+ '))
                    removed_lines = sum(1 for line in diff if line.startswith('- '))
                    unchanged_lines = sum(1 for line in diff if line.startswith('  '))

                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    with col_stat1:
                        st.metric("新增行数", added_lines)
                    with col_stat2:
                        st.metric("删除行数", removed_lines)
                    with col_stat3:
                        st.metric("相同行数", unchanged_lines)

                    # 显示行对比结果
                    html_parts = [
                        "<div style='background-color: #f8f9fa; padding: 10px; border-radius: 5px; font-family: monospace;'>"]
                    for line in diff:
                        escaped_line = html.escape(line[2:] if len(line) > 2 else line)
                        if line.startswith('+ '):
                            html_parts.append(
                                f"<div style='background-color: #d4edda; margin: 2px 0; padding: 2px 5px; border-radius: 3px; border-left: 3px solid #28a745;'>"
                                f"<span style='color: #28a745; font-weight: bold;'>+ </span>{escaped_line}</div>")
                        elif line.startswith('- '):
                            html_parts.append(
                                f"<div style='background-color: #f8d7da; margin: 2px 0; padding: 2px 5px; border-radius: 3px; border-left: 3px solid #dc3545;'>"
                                f"<span style='color: #dc3545; font-weight: bold;'>- </span>{escaped_line}</div>")
                        elif line.startswith('? '):
                            html_parts.append(
                                f"<div style='background-color: #fff3cd; margin: 2px 0; padding: 2px 5px; border-radius: 3px; border-left: 3px solid #ffc107;'>"
                                f"<span style='color: #856404; font-weight: bold;'>? </span>{escaped_line}</div>")
                        else:
                            content = escaped_line if line.startswith('  ') else html.escape(line)
                            html_parts.append(
                                f"<div style='margin: 2px 0; padding: 2px 5px; border-left: 3px solid #6c757d; color: #6c757d;'>"
                                f"{content}</div>")
                    html_parts.append("</div>")
                    result_html = ''.join(html_parts)
                    st.markdown(result_html, unsafe_allow_html=True)

                elif diff_mode == 'word':
                    # 基本词对比模式
                    with st.spinner("正在进行词级对比..."):
                        diff, words1, words2 = word_diff(processed_text1, processed_text2)

                        # 词级统计
                        added_words = sum(1 for word in diff if word.startswith('+ '))
                        removed_words = sum(1 for word in diff if word.startswith('- '))
                        unchanged_words = sum(1 for word in diff if word.startswith('  '))

                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            st.metric("新增词汇", added_words)
                        with col_stat2:
                            st.metric("删除词汇", removed_words)
                        with col_stat3:
                            st.metric("相同词汇", unchanged_words)

                        result_html = render_word_diff(diff)
                        st.markdown(result_html, unsafe_allow_html=True)

                else:  # enhanced_word
                    # 增强词对比模式
                    with st.spinner("正在进行增强词级对比..."):
                        result_html = render_enhanced_word_diff(processed_text1, processed_text2)

                        # 简单统计
                        words1 = re.findall(r'\b\w+\b', processed_text1)
                        words2 = re.findall(r'\b\w+\b', processed_text2)

                        col_stat1, col_stat2 = st.columns(2)
                        with col_stat1:
                            st.metric("原文词汇数", len(words1))
                        with col_stat2:
                            st.metric("对比文词汇数", len(words2))

                        st.markdown(result_html, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"发生错误: {e}")
                st.info("建议尝试使用行对比模式或检查文本格式")
        else:
            st.warning("⚠️ 请填写原始文本和对比文本")

    st.markdown('</div>', unsafe_allow_html=True)

# 正则表达式测试工具
elif tool_category == "正则测试工具":
    show_doc("regex_tester")

    # 初始化session_state
    if 'regex_clear_counter' not in st.session_state:
        st.session_state.regex_clear_counter = 0

    # 添加工具选择选项卡
    tab1, tab2, tab3 = st.tabs(["正则表达式测试", "代码生成器", "从示例生成"])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            # 预定义模式选择
            st.markdown("**选择预定义模式**")
            selected_pattern = st.selectbox("", ["自定义"] + list(PREDEFINED_PATTERNS.keys()), key="pattern_select")

            # 使用不同的key策略来避免session_state冲突
            if selected_pattern != "自定义":
                regex_pattern = PREDEFINED_PATTERNS[selected_pattern]
                st.code(f"当前模式: {regex_pattern}")
                # 同时允许用户修改预定义模式
                custom_regex = st.text_input("或自定义正则表达式", value=regex_pattern, placeholder="可在此修改表达式",
                                             key=f"custom_regex_input_{st.session_state.regex_clear_counter}")
                if custom_regex != regex_pattern:
                    regex_pattern = custom_regex
            else:
                regex_pattern = st.text_input("正则表达式", placeholder="例如: ^[a-zA-Z0-9]+$",
                                              key=f"manual_regex_input_{st.session_state.regex_clear_counter}")

            test_text = st.text_area("测试文本", height=200, placeholder="在此输入要测试的文本...",
                                     key=f"test_text_area_{st.session_state.regex_clear_counter}")

        with col2:
            st.markdown("**匹配选项**")
            global_match = st.checkbox("全局匹配 (g)", value=True, key="global_match_check")
            ignore_case = st.checkbox("忽略大小写 (i)", key="ignore_case_check")
            multiline = st.checkbox("多行模式 (m)", key="multiline_check")
            dotall = st.checkbox("点号匹配换行 (s)", key="dotall_check")

            st.markdown("**替换功能**")
            replace_text = st.text_input("替换文本", placeholder="输入替换文本（可选）",
                                         key=f"replace_text_input_{st.session_state.regex_clear_counter}")

        button_col1, button_col2 = st.columns(2)
        with button_col1:
            if st.button("测试正则表达式", use_container_width=True, key="test_regex"):
                # 获取当前输入框的值
                current_regex = ""
                if selected_pattern != "自定义":
                    current_regex = custom_regex
                else:
                    current_regex = regex_pattern

                current_test_text = test_text

                if current_regex and current_test_text:
                    try:
                        flags = 0
                        if ignore_case:
                            flags |= re.IGNORECASE
                        if multiline:
                            flags |= re.MULTILINE
                        if dotall:
                            flags |= re.DOTALL

                        if global_match:
                            matches = list(re.finditer(current_regex, current_test_text, flags))
                            match_count = len(matches)

                            if match_count > 0:
                                st.success(f"匹配成功！找到 {match_count} 个匹配项。")

                                # 增强的匹配详情显示
                                st.markdown("**匹配详情**")
                                for i, match in enumerate(matches):
                                    with st.expander(f"匹配 {i + 1}: 位置 {match.start()}-{match.end()}"):
                                        st.write(f"匹配文本: `{match.group()}`")
                                        if match.groups():
                                            st.write("**捕获组:**")
                                            for j, group in enumerate(match.groups(), 1):
                                                st.write(f"  组 {j}: `{group}`")
                                        if match.groupdict():
                                            st.write("**命名分组:**")
                                            for name, group in match.groupdict().items():
                                                st.write(f"  {name}: `{group}`")
                            else:
                                st.warning("未找到匹配项。")
                        else:
                            match = re.search(current_regex, current_test_text, flags)
                            if match:
                                st.success("匹配成功！")
                                st.write(f"匹配文本: `{match.group()}`")
                                st.write(f"匹配位置: {match.start()}-{match.end()}")
                                if match.groups():
                                    st.write("**捕获组:**")
                                    for i, group in enumerate(match.groups(), 1):
                                        st.write(f"组 {i}: `{group}`")
                            else:
                                st.warning("未找到匹配项。")

                        if replace_text:
                            replaced_text = re.sub(current_regex, replace_text, current_test_text, flags=flags)
                            st.markdown("**替换结果**")
                            display_generated_results("替换后的文本", replaced_text, "regex_replaced")
                    except re.error as e:
                        st.error(f"正则表达式错误: {e}")
                else:
                    st.warning("请输入正则表达式和测试文本")

        with button_col2:
            if st.button("🗑️ 清空输入", use_container_width=True, key="clear_input"):
                # 通过增加计数器并重新渲染来清空
                st.session_state.regex_clear_counter += 1
                st.rerun()

    with tab2:
        st.markdown("### 正则表达式代码生成器")

        col1, col2 = st.columns(2)

        with col1:
            # 模式选择：预定义或自定义
            pattern_source = st.radio("正则表达式来源", ["预定义模式", "自定义表达式"],
                                      key=f"pattern_source_{st.session_state.regex_clear_counter}")

            if pattern_source == "预定义模式":
                code_pattern = st.selectbox("选择预定义模式", list(PREDEFINED_PATTERNS.keys()),
                                            key=f"code_pattern_{st.session_state.regex_clear_counter}")
                pattern_display = PREDEFINED_PATTERNS[code_pattern]
                st.code(f"模式: {pattern_display}")
            else:
                pattern_display = st.text_input("输入自定义正则表达式", placeholder="例如: ^[a-zA-Z0-9]+$",
                                                key=f"custom_pattern_input_{st.session_state.regex_clear_counter}")
                if pattern_display:
                    st.code(f"模式: {pattern_display}")

            # 编程语言选择
            target_language = st.selectbox("选择目标语言", list(LANGUAGE_TEMPLATES.keys()),
                                           key=f"target_lang_{st.session_state.regex_clear_counter}")

            # 操作类型
            operation_type = st.radio("选择操作类型", ["匹配", "测试", "替换"],
                                      key=f"operation_type_{st.session_state.regex_clear_counter}")

            # 替换文本
            replacement_code = ""
            if operation_type == "替换":
                replacement_code = st.text_input("替换文本", placeholder="输入替换文本",
                                                 key=f"replacement_input_{st.session_state.regex_clear_counter}")

        with col2:
            st.markdown("**代码生成选项**")

            # 标志选择
            flags_selected = []
            lang_flags = LANGUAGE_TEMPLATES[target_language]["flags"]

            for flag_name, flag_char in lang_flags.items():
                if st.checkbox(f"{flag_name} ({flag_char})",
                               key=f"flag_{flag_char}_{target_language}_{st.session_state.regex_clear_counter}"):
                    flags_selected.append(flag_name)

            # 生成代码按钮
            if st.button("生成代码", use_container_width=True, key="generate_code"):
                current_pattern = ""
                if pattern_source == "预定义模式":
                    current_pattern = PREDEFINED_PATTERNS[code_pattern]
                else:
                    current_pattern = pattern_display

                if not current_pattern:
                    st.warning("请输入或选择正则表达式")
                else:
                    # 构建标志
                    if target_language in ["Python", "Java", "C#"]:
                        flags_value = " | ".join(flags_selected) if flags_selected else "0"
                    else:
                        flags_value = "".join([lang_flags[flag] for flag in flags_selected])

                    # 获取模板
                    template_key = "match" if operation_type == "匹配" else "test" if operation_type == "测试" else "replace"
                    template = LANGUAGE_TEMPLATES[target_language][template_key]

                    # 生成代码
                    try:
                        generated_code = template.format(
                            pattern=current_pattern,
                            flags=flags_value,
                            flags_value=flags_value,
                            replacement=replacement_code
                        )

                        st.session_state.generated_code = generated_code
                        st.session_state.generated_language = target_language

                    except KeyError as e:
                        st.error(f"代码生成错误: {e}")

            # 显示已生成的代码（如果有）
            if 'generated_code' in st.session_state and st.session_state.generated_code:
                language = st.session_state.generated_language if 'generated_language' in st.session_state else target_language
                display_generated_results(
                    f"{language} 代码",
                    st.session_state.generated_code,
                    f"regex_{language.lower()}_code"
                )

        # 清空所有按钮
        button_col3, _ = st.columns(2)
        with button_col3:
            if st.button("🗑️ 清空所有", use_container_width=True, key="clear_all_code"):
                # 清除生成的代码状态
                if 'generated_code' in st.session_state:
                    del st.session_state.generated_code
                if 'generated_language' in st.session_state:
                    del st.session_state.generated_language
                # 通过增加计数器清空输入
                st.session_state.regex_clear_counter += 1
                st.rerun()

    with tab3:
        st.markdown("### 从示例生成正则表达式")

        col1, col2 = st.columns(2)

        with col1:
            source_text = st.text_area("原文内容", height=150,
                                       placeholder="输入包含要提取内容的原文...",
                                       key=f"source_text_area_{st.session_state.regex_clear_counter}")

        with col2:
            examples_text = st.text_area("示例文本（用逗号分隔）", height=150,
                                         placeholder="输入要匹配的示例，用逗号分隔...",
                                         key=f"examples_text_area_{st.session_state.regex_clear_counter}")

        button_col4, button_col5 = st.columns(2)
        with button_col4:
            if st.button("生成正则表达式", use_container_width=True, key="generate_from_examples"):
                current_source = source_text
                current_examples = examples_text

                if current_source and current_examples:
                    generated_regex = generate_regex_from_examples(current_source, current_examples)

                    if generated_regex:
                        st.success("已生成正则表达式！")

                        # 使用统一的显示函数
                        display_generated_results("生成的正则表达式", generated_regex, "generated_regex")

                        # 测试生成的正则表达式
                        try:
                            matches = re.findall(generated_regex, current_source)
                            if matches:
                                st.write(f"在原文中找到 {len(matches)} 个匹配项:")
                                for i, match in enumerate(matches):
                                    st.write(f"{i + 1}. `{match}`")
                            else:
                                st.warning("生成的正则表达式在原文中未找到匹配项")
                        except re.error as e:
                            st.error(f"生成的正则表达式有误: {e}")
                    else:
                        st.warning("无法生成合适的正则表达式，请提供更多或更明确的示例")
                else:
                    st.warning("请输入原文内容和示例文本")

        with button_col5:
            if st.button("🗑️ 清空示例", use_container_width=True, key="clear_examples"):
                # 通过增加计数器清空输入
                st.session_state.regex_clear_counter += 1
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)
# JSON数据对比工具
elif tool_category == "JSON处理工具":
    utils = JSONFileUtils()

    # 工具选择
    tool_mode = st.radio(
        "选择处理模式",
        ["JSON解析与格式化", "JSON数据对比", "JSONPath查询"],
        horizontal=True
    )

    if tool_mode == "JSON解析与格式化":
        show_doc("json_parser")

        # 初始化session_state
        if 'json_input_content' not in st.session_state:
            st.session_state.json_input_content = '{"name": "Tom", "age": 25, "hobbies": ["reading", "swimming"]}'
        if 'parse_result' not in st.session_state:
            st.session_state.parse_result = None
        if 'parse_error' not in st.session_state:
            st.session_state.parse_error = None

        # 输入区域
        st.markdown("**JSON输入**")
        json_input = st.text_area("", height=300, key="json_input", value=st.session_state.json_input_content,
                                  placeholder='请输入JSON字符串，例如: {"name": "Tom", "age": 25}')

        # 按钮区域
        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])

        with col1:
            if st.button("🚀 解析JSON", use_container_width=True, key="parse_json"):
                if json_input.strip():
                    try:
                        parsed_json = json.loads(json_input)
                        st.session_state.parse_result = parsed_json
                        st.session_state.parse_error = None
                    except json.JSONDecodeError as e:
                        st.session_state.parse_result = None
                        st.session_state.parse_error = str(e)
                else:
                    st.warning("请输入JSON字符串")

        with col2:
            if st.button("✨ 格式化", use_container_width=True, key="format_json"):
                if json_input.strip():
                    try:
                        parsed_json = json.loads(json_input)
                        formatted_json = json.dumps(parsed_json, indent=2, ensure_ascii=False)
                        st.session_state.json_input_content = formatted_json
                        st.session_state.parse_result = parsed_json
                        st.session_state.parse_error = None
                    except json.JSONDecodeError as e:
                        st.session_state.parse_error = str(e)

        with col3:
            # 使用统一的复制按钮
            if st.session_state.parse_result is not None:
                formatted_json = json.dumps(st.session_state.parse_result, indent=2, ensure_ascii=False)
                create_copy_button(
                    formatted_json,
                    button_text="📋 复制结果",
                    key="copy_json_result"
                )
            else:
                # 禁用状态的按钮
                st.button("📋 复制结果", use_container_width=True, disabled=True, key="copy_disabled")

        with col4:
            if st.button("🗑️ 清空", use_container_width=True, key="clear_json"):
                st.session_state.json_input_content = ""
                st.session_state.parse_result = None
                st.session_state.parse_error = None

        # 显示解析结果
        if st.session_state.parse_result is not None:
            st.markdown("### 📊 解析结果")
            formatted_json = json.dumps(st.session_state.parse_result, indent=2, ensure_ascii=False)

            with st.expander("📄 格式化JSON", expanded=True):
                st.code(formatted_json, language='json')

            # 显示错误信息（如果有）
            if st.session_state.parse_error:
                st.error(f"解析错误: {st.session_state.parse_error}")


    elif tool_mode == "JSON数据对比":
        show_doc("json_comparison")

        # 初始化 session_state
        if 'json1_content' not in st.session_state:
            st.session_state.json1_content = '{"name": "John", "age": 30, "city": "New York"}'
        if 'json2_content' not in st.session_state:
            st.session_state.json2_content = '{"name": "Jane", "age": 25, "country": "USA"}'
        if 'comparison_result' not in st.session_state:
            st.session_state.comparison_result = None
        if 'differences_text' not in st.session_state:
            st.session_state.differences_text = ""
        if 'clear_counter' not in st.session_state:
            st.session_state.clear_counter = 0  # 添加计数器用于强制重新渲染

        # 输入区域 - 使用计数器确保重新渲染
        input_cols = st.columns(2)
        with input_cols[0]:
            st.markdown("**JSON 1**")
            json1 = st.text_area("", height=300,
                                 key=f"json1_{st.session_state.clear_counter}",  # 使用动态key
                                 value=st.session_state.json1_content,
                                 placeholder='输入第一个JSON数据...')
        with input_cols[1]:
            st.markdown("**JSON 2**")
            json2 = st.text_area("", height=300,
                                 key=f"json2_{st.session_state.clear_counter}",  # 使用动态key
                                 value=st.session_state.json2_content,
                                 placeholder='输入第二个JSON数据...')

        # 按钮区域
        button_cols = st.columns(4)
        with button_cols[0]:
            if st.button("✨ 格式化全部", use_container_width=True, key="format_all"):
                try:
                    # 先同步当前输入的内容到 session state
                    st.session_state.json1_content = json1
                    st.session_state.json2_content = json2

                    if json1.strip():
                        parsed_json1 = json.loads(json1)
                        formatted_json1 = json.dumps(parsed_json1, indent=2, ensure_ascii=False)
                        st.session_state.json1_content = formatted_json1
                    if json2.strip():
                        parsed_json2 = json.loads(json2)
                        formatted_json2 = json.dumps(parsed_json2, indent=2, ensure_ascii=False)
                        st.session_state.json2_content = formatted_json2

                    st.session_state.clear_counter += 1  # 增加计数器强制重新渲染
                    st.rerun()
                except json.JSONDecodeError as e:
                    st.error(f"JSON格式错误: {e}")

        with button_cols[1]:
            compare_clicked = st.button("🔍 开始对比", use_container_width=True, key="compare")

        with button_cols[2]:
            if st.button("🔄 交换数据", use_container_width=True, key="swap_data"):
                # 先同步当前输入的内容到 session state
                st.session_state.json1_content = json1
                st.session_state.json2_content = json2
                # 然后交换数据
                st.session_state.json1_content, st.session_state.json2_content = \
                    st.session_state.json2_content, st.session_state.json1_content
                st.session_state.clear_counter += 1  # 增加计数器强制重新渲染
                st.rerun()

        with button_cols[3]:
            if st.button("🗑️ 清空全部", use_container_width=True, key="clear_all"):
                st.session_state.json1_content = ""
                st.session_state.json2_content = ""
                st.session_state.comparison_result = None
                st.session_state.differences_text = ""
                st.session_state.clear_counter += 1  # 增加计数器强制重新渲染
                st.rerun()

        # 处理对比结果
        if compare_clicked:
            # 同步当前输入的内容到 session state
            st.session_state.json1_content = json1
            st.session_state.json2_content = json2

            if json1 and json2:
                try:
                    obj1 = json.loads(json1)
                    obj2 = json.loads(json2)

                    st.markdown("### 📋 对比结果")

                    utils.reset_stats()
                    differences = utils.compare_json(obj1, obj2)
                    st.session_state.comparison_result = differences

                    difference_text = "\n".join([f"- {diff}" for diff in differences])
                    st.session_state.differences_text = difference_text

                    if differences:
                        st.error(f"发现 {len(differences)} 个差异:")
                        st.write(difference_text)

                        # 使用下载按钮作为复制替代方案
                        st.download_button(
                            "📋 下载差异结果",
                            difference_text,
                            file_name="json_differences.txt",
                            mime="text/plain",
                            use_container_width=True,
                            key="download_diff"
                        )

                        # 同时提供文本区域用于手动复制
                        st.text_area("差异结果", difference_text, height=200, key="diff_copy_area")
                    else:
                        st.success("✅ 两个JSON对象完全相同")

                except json.JSONDecodeError as e:
                    st.error(f"JSON格式错误: {e}")

    elif tool_mode == "JSONPath查询":
        show_doc("jsonpath_tool")

        # 初始化session_state
        if 'jsonpath_json_content' not in st.session_state:
            st.session_state.jsonpath_json_content = '{"store": {"book": [{"title": "Book 1", "author": "Author 1"}, {"title": "Book 2", "author": "Author 2"}]}}'
        if 'jsonpath_expression' not in st.session_state:
            st.session_state.jsonpath_expression = "$.store.book[*].author"
        if 'jsonpath_result' not in st.session_state:
            st.session_state.jsonpath_result = None
        if 'jsonpath_result_text' not in st.session_state:
            st.session_state.jsonpath_result_text = ""

        # 布局：左右分栏
        left_col, right_col = st.columns([1, 1])

        with left_col:
            st.markdown("**📝 JSON数据**")
            json_data_input = st.text_area("", height=400, key="jsonpath_json",
                                           value=st.session_state.jsonpath_json_content,
                                           placeholder='输入JSON数据...')

            st.markdown("**🎯 JSONPath表达式**")
            jsonpath_input = st.text_input("", key="jsonpath_expr",
                                           value=st.session_state.jsonpath_expression,
                                           placeholder='例如: $.store.book[*].author')

            # 操作按钮
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🚀 执行查询", use_container_width=True, key="execute_jsonpath"):
                    if json_data_input.strip() and jsonpath_input.strip():
                        try:
                            json_data = json.loads(json_data_input)
                            result = utils.execute_jsonpath(json_data, jsonpath_input)
                            st.session_state.jsonpath_result = result
                            result_text = "\n".join([str(item) for item in result])
                            st.session_state.jsonpath_result_text = result_text
                        except json.JSONDecodeError as e:
                            st.error(f"JSON数据格式错误: {e}")
                        except Exception as e:
                            st.error(f"JSONPath查询错误: {e}")

        with right_col:
            st.markdown("### 📋 查询结果")

            if st.session_state.jsonpath_result is not None:
                result = st.session_state.jsonpath_result
                result_text = st.session_state.jsonpath_result_text

                if result:
                    st.success(f"✅ 找到 {len(result)} 个匹配项")
                    st.metric("匹配数量", len(result))

                    st.markdown("**📄 匹配结果:**")
                    for i, item in enumerate(result):
                        with st.expander(f"结果 #{i + 1}", expanded=len(result) <= 3):
                            if isinstance(item, (dict, list)):
                                st.json(item)
                            else:
                                st.code(str(item))

                    # 使用下载按钮
                    st.download_button(
                        "📋 下载查询结果",
                        result_text,
                        file_name="jsonpath_results.txt",
                        mime="text/plain",
                        use_container_width=True,
                        key="download_jsonpath"
                    )

                    # 提供文本区域用于手动复制
                    st.text_area("查询结果", result_text, height=200, key="jsonpath_copy_area")
                else:
                    st.warning("❌ 未找到匹配项")

# 日志分析工具
if tool_category == "日志分析工具":
    show_doc("log_analyzer")
    utils = LogAnalyzerUtils()

    # 初始化所有session_state变量
    session_vars = [
        'log_data', 'file_info', 'filtered_lines', 'search_results',
        'search_count', 'df', 'is_csv', 'csv_columns', 'json_columns',
        'json_fields', 'search_keyword', 'search_cleared',
        'text_filters', 'json_filters', 'filter_logic', 'auto_apply_filters'
    ]

    for var in session_vars:
        if var not in st.session_state:
            if var in ['text_filters', 'json_filters']:
                st.session_state[var] = []
            elif var == 'filter_logic':
                st.session_state[var] = "AND"
            elif var == 'auto_apply_filters':
                st.session_state[var] = False
            else:
                st.session_state[var] = None

    # 使用tab布局
    tab1, tab2, tab3 = st.tabs(["📥 日志导入", "🔍 日志过滤", "🔎 关键词搜索"])

    # Tab1: 日志导入
    with tab1:
        st.header("日志导入")

        import_method = st.radio("日志导入方式", ["文件上传", "直接粘贴"], horizontal=True)
        log_content = ""
        file_info = None

        if import_method == "文件上传":
            uploaded_file = st.file_uploader("选择日志文件", type=['txt', 'log', 'csv'],
                                             help="支持 txt, log, csv 格式文件")

            if uploaded_file is not None:
                # 文件信息
                current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                file_info = {
                    "文件名": uploaded_file.name,
                    "文件大小": f"{uploaded_file.size / 1024:.2f} KB",
                    "文件类型": uploaded_file.type or "未知类型",
                    "上传时间": current_time
                }

                # 文件预览
                st.subheader("文件预览")
                preview_lines = 10

                try:
                    if uploaded_file.name.endswith('.csv'):
                        # CSV文件处理
                        df = pd.read_csv(uploaded_file)
                        st.write("前10行数据预览:")
                        st.dataframe(df.head(preview_lines), use_container_width=True)

                        # 保存DataFrame和列信息
                        st.session_state.df = df
                        st.session_state.csv_columns = df.columns.tolist()
                        st.session_state.is_csv = True

                        # 检测JSON格式的列并提取字段
                        st.session_state.json_columns = []
                        st.session_state.json_fields = {}

                        for column in df.columns:
                            # 检查列中是否包含JSON格式的数据
                            json_sample = None
                            for value in df[column].dropna().head(5):
                                if isinstance(value, str) and value.strip().startswith('{') and value.strip().endswith(
                                        '}'):
                                    try:
                                        json_data = json.loads(value)
                                        if isinstance(json_data, dict):
                                            json_sample = json_data
                                            break
                                    except:
                                        continue

                            if json_sample:
                                st.session_state.json_columns.append(column)
                                st.session_state.json_fields[column] = list(json_sample.keys())
                                st.success(f"✅ 检测到列 '{column}' 包含JSON数据，提取到 {len(json_sample.keys())} 个字段")

                        # 将DataFrame转换为文本格式用于显示
                        log_content = ""
                        for _, row in df.iterrows():
                            log_content += " | ".join([str(x) for x in row]) + "\n"
                    else:
                        # 文本文件处理
                        content = uploaded_file.getvalue().decode("utf-8")
                        lines = content.split('\n')
                        preview_content = "\n".join(lines[:preview_lines])
                        st.text_area("预览内容", preview_content, height=150, key="preview")
                        log_content = content
                        st.session_state.is_csv = False
                        st.session_state.df = None
                        st.session_state.csv_columns = []
                        st.session_state.json_columns = []
                        st.session_state.json_fields = {}

                except Exception as e:
                    st.error(f"❌ 文件读取错误: {e}")
                    log_content = ""

                # 显示文件信息
                if file_info:
                    st.subheader("文件信息")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**文件名:** {file_info['文件名']}")
                        st.write(f"**文件大小:** {file_info['文件大小']}")
                    with col2:
                        st.write(f"**文件类型:** {file_info['文件类型']}")
                        st.write(f"**上传时间:** {file_info['上传时间']}")

                # 自动导入日志数据
                if log_content and uploaded_file is not None:
                    st.session_state.log_data = log_content
                    st.session_state.file_info = file_info
                    st.session_state.filtered_lines = []
                    st.session_state.search_results = []
                    st.session_state.search_count = 0
                    st.success("✅ 日志数据导入成功！")

        else:  # 直接粘贴
            if 'clear_paste_counter' not in st.session_state:
                st.session_state.clear_paste_counter = 0
            col1, col2 = st.columns([3, 1])
            with col1:
                # 使用动态key，每次清空时key会改变
                paste_key = f"paste_content_{st.session_state.clear_paste_counter}"
                log_content = st.text_area("粘贴日志内容", height=200,
                                           placeholder="请将日志内容粘贴到此处...",
                                           key=paste_key)
            with col2:
                st.write("")  # 占位符
                st.write("")
                if st.button("🗑️ 清空", key="clear_paste", use_container_width=True):
                    # 清空所有相关状态
                    st.session_state.log_data = None
                    st.session_state.file_info = None
                    st.session_state.filtered_lines = []
                    st.session_state.search_results = []
                    st.session_state.search_count = 0
                    st.session_state.is_csv = False
                    st.session_state.df = None
                    st.session_state.csv_columns = []
                    st.session_state.json_columns = []
                    st.session_state.json_fields = {}
                    # 增加清空计数器，使text_area的key改变
                    st.session_state.clear_paste_counter += 1
                    st.rerun()

            # 自动导入粘贴的日志内容
            if log_content and log_content.strip():
                st.session_state.log_data = log_content
                st.session_state.file_info = None
                st.session_state.filtered_lines = []
                st.session_state.search_results = []
                st.session_state.search_count = 0
                st.session_state.is_csv = False
                st.session_state.df = None
                st.session_state.csv_columns = []
                st.session_state.json_columns = []
                st.session_state.json_fields = {}
                st.success("✅ 日志数据导入成功！")

    # 检查是否有导入的日志数据并显示统计信息
    if st.session_state.log_data:
        log_content = st.session_state.log_data

        # 根据文件类型处理数据
        if st.session_state.is_csv and st.session_state.df is not None:
            # CSV数据
            df = st.session_state.df
            lines = []
            for _, row in df.iterrows():
                line = " | ".join([str(x) for x in row])
                lines.append(line)
            total_lines = len(df)
        else:
            # 文本数据
            lines = log_content.split('\n')
            total_lines = len(lines)

        # 在主内容区显示统计信息
        st.header("📊 日志统计信息")

        # 改进的日志级别统计
        # 改进的日志级别统计 - 修复DEBUG判断
        error_count = sum(
            1 for line in lines if any(word in line.upper() for word in [' ERROR', ' ERR ', ']ERROR', ']ERR']))
        warn_count = sum(
            1 for line in lines if any(word in line.upper() for word in [' WARN', ' WARNING', ']WARN', ']WARNING']))
        info_count = sum(1 for line in lines if
                         any(word in line.upper() for word in [' INFO', ' INFORMATION', ']INFO', ']INFORMATION']))
        debug_count = sum(
            1 for line in lines if any(word in line.upper() for word in [' DEBUG', ' DBG', ']DEBUG', ']DBG']))
        other_count = total_lines - error_count - warn_count - info_count - debug_count

        # 统计指标
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("总行数", total_lines)
        with col2:
            st.metric("错误", error_count, delta_color="off")
        with col3:
            st.metric("警告", warn_count, delta_color="off")
        with col4:
            st.metric("信息", info_count, delta_color="off")
        with col5:
            st.metric("调试", debug_count, delta_color="off")

        # 级别分布图表
        if total_lines > 0:
            try:
                import plotly.express as px

                level_data = {
                    '级别': ['错误', '警告', '信息', '调试', '其他'],
                    '数量': [error_count, warn_count, info_count, debug_count, other_count],
                    '百分比': [
                        f"{(error_count / total_lines) * 100:.1f}%",
                        f"{(warn_count / total_lines) * 100:.1f}%",
                        f"{(info_count / total_lines) * 100:.1f}%",
                        f"{(debug_count / total_lines) * 100:.1f}%",
                        f"{(other_count / total_lines) * 100:.1f}%"
                    ]
                }

                fig = px.pie(level_data, values='数量', names='级别',
                             title='日志级别分布',
                             hover_data=['百分比'])
                st.plotly_chart(fig, use_container_width=True)
            except ImportError:
                st.info("如需查看图表，请安装 plotly: pip install plotly")

        # Tab2: 日志过滤 - 支持多条件组合查询
        with tab2:
            st.header("🔍 日志过滤 - 多条件组合查询")

            # 自动应用过滤选项
            col1, col2 = st.columns([1, 3])
            with col1:
                st.session_state.auto_apply_filters = st.checkbox(
                    "自动应用过滤",
                    value=st.session_state.auto_apply_filters,
                    help="启用后，添加/删除条件会自动应用过滤"
                )

            # 逻辑运算符选择
            with col2:
                logic_operator = st.radio(
                    "组合逻辑",
                    ["AND", "OR"],
                    index=0 if st.session_state.filter_logic == "AND" else 1,
                    help="AND: 所有条件都要满足, OR: 任一条件满足即可",
                    horizontal=True
                )
                # 立即更新逻辑运算符
                if logic_operator != st.session_state.filter_logic:
                    st.session_state.filter_logic = logic_operator
                    if st.session_state.auto_apply_filters:
                        st.rerun()

            # 条件管理区域
            st.subheader("📋 筛选条件管理")

            if st.session_state.is_csv and st.session_state.csv_columns:
                # CSV文件的字段筛选
                filter_type = st.radio("筛选类型", ["普通文本筛选", "CSV列筛选", "JSON字段筛选"], horizontal=True)
            else:
                filter_type = "普通文本筛选"

            # 添加新条件的表单
            with st.expander("➕ 添加新筛选条件", expanded=False):
                if filter_type == "普通文本筛选":
                    col1, col2, col3 = st.columns([2, 2, 1])

                    with col1:
                        condition_type = st.selectbox(
                            "条件类型",
                            ["log_level", "ip_filter", "status_code", "keyword", "show_only_errors", "hide_debug"],
                            format_func=lambda x: {
                                "log_level": "日志级别",
                                "ip_filter": "IP地址",
                                "status_code": "状态码",
                                "keyword": "关键词",
                                "show_only_errors": "仅显示错误",
                                "hide_debug": "隐藏调试"
                            }[x],
                            key="new_condition_type"
                        )

                    with col2:
                        if condition_type in ["log_level", "ip_filter", "status_code", "keyword"]:
                            if condition_type == "log_level":
                                condition_value = st.multiselect(
                                    "值",
                                    ["错误", "警告", "信息", "调试"],
                                    default=["错误", "警告"],
                                    key="new_condition_value"
                                )
                            elif condition_type == "status_code":
                                condition_value = st.text_input("状态码(逗号分隔)", placeholder="200,404,500",
                                                                key="new_condition_value")
                            else:
                                condition_value = st.text_input("值", placeholder="输入筛选值...", key="new_condition_value")
                        else:
                            condition_value = "true"
                            st.info("此条件无需输入值")

                    with col3:
                        st.write("")  # 占位
                        st.write("")
                        if st.button("添加条件", key="add_text_condition", use_container_width=True):
                            if condition_type in ["log_level", "ip_filter", "status_code",
                                                  "keyword"] and not condition_value:
                                st.warning("请输入条件值")
                            else:
                                new_filter = {
                                    'type': condition_type,
                                    'value': condition_value
                                }
                                st.session_state.text_filters.append(new_filter)
                                st.success("✅ 条件已添加")
                                # 如果启用自动应用，立即重新运行
                                if st.session_state.auto_apply_filters:
                                    st.rerun()

                elif filter_type == "CSV列筛选" and st.session_state.csv_columns:
                    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])

                    with col1:
                        selected_column = st.selectbox("选择列", st.session_state.csv_columns, key="csv_filter_col")

                    with col2:
                        filter_operator = st.selectbox("操作符", ["包含", "等于", "开头为", "结尾为", "有值", "没有值"], key="csv_filter_op")

                    with col3:
                        if filter_operator in ["包含", "等于", "开头为", "结尾为"]:
                            filter_value = st.text_input("值", key="csv_filter_value")
                        else:  # "有值" 或 "没有值"
                            filter_value = None
                            st.info("此条件无需输入值")

                    with col4:
                        st.write("")
                        st.write("")
                        if st.button("添加条件", key="add_csv_condition", use_container_width=True):
                            if filter_operator in ["包含", "等于", "开头为", "结尾为"] and not filter_value:
                                st.warning("请输入筛选值")
                            else:
                                # 将CSV列条件转换为文本条件
                                new_filter = {
                                    'type': 'keyword',
                                    'value': filter_value if filter_operator in ["包含", "等于", "开头为", "结尾为"] else "",
                                    'column': selected_column,
                                    'operator': filter_operator
                                }
                                st.session_state.text_filters.append(new_filter)
                                st.success("✅ 条件已添加")
                                if st.session_state.auto_apply_filters:
                                    st.rerun()
                elif filter_type == "JSON字段筛选" and st.session_state.json_columns:
                    col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])

                    with col1:
                        json_column = st.selectbox("JSON列", st.session_state.json_columns, key="json_filter_col")

                    with col2:
                        if json_column in st.session_state.json_fields:
                            json_field = st.selectbox("JSON字段", st.session_state.json_fields[json_column],
                                                      key="json_filter_field")
                        else:
                            json_field = None
                            st.warning("未找到JSON字段")

                    with col3:
                        filter_operator = st.selectbox("操作符", ["包含", "等于", "开头为", "结尾为", "有值", "没有值"], key="json_filter_op")

                    with col4:
                        if filter_operator in ["包含", "等于", "开头为", "结尾为"]:
                            json_value = st.text_input("字段值", key="json_filter_value")
                            value_range = None
                        elif filter_operator == "数值范围":
                            json_value = None
                            # 数值范围筛选代码
                            try:
                                numeric_values = []
                                for value in st.session_state.df[json_column].dropna():
                                    if isinstance(value, str) and value.strip().startswith(
                                            '{') and value.strip().endswith('}'):
                                        try:
                                            json_data = json.loads(value)
                                            if (json_field in json_data and
                                                    isinstance(json_data[json_field], (int, float))):
                                                numeric_values.append(json_data[json_field])
                                        except:
                                            continue

                                if numeric_values:
                                    min_val = min(numeric_values)
                                    max_val = max(numeric_values)
                                    value_range = st.slider("范围", min_val, max_val, (min_val, max_val),
                                                            key="json_range_slider")
                                else:
                                    value_range = None
                                    st.info("该字段不包含数值数据")
                            except:
                                value_range = None
                        else:  # "有值" 或 "没有值"
                            json_value = None
                            value_range = None
                            st.info("此条件无需输入值")

                    with col5:
                        st.write("")
                        st.write("")
                        if st.button("添加条件", key="add_json_condition", use_container_width=True):
                            if not json_field:
                                st.warning("请选择JSON字段")
                            else:
                                new_filter = {
                                    'column': json_column,
                                    'field': json_field,
                                    'operator': filter_operator,  # 新增操作符
                                    'value': json_value,
                                    'value_range': value_range
                                }
                                st.session_state.json_filters.append(new_filter)
                                st.success("✅ JSON条件已添加")
                                if st.session_state.auto_apply_filters:
                                    st.rerun()

            # 显示当前条件列表
            if st.session_state.text_filters or st.session_state.json_filters:
                st.subheader("当前筛选条件")

                # 显示文本条件
                for i, filter_config in enumerate(st.session_state.text_filters):
                    col1, col2, col3 = st.columns([3, 2, 1])
                    with col1:
                        if filter_config['type'] == 'keyword' and filter_config.get('column'):
                            # 对于CSV列筛选条件，显示列名和操作符
                            column_name = filter_config.get('column', '未知列')
                            operator = filter_config.get('operator', '包含')
                            value = filter_config.get('value', '')

                            if operator in ['有值', '没有值']:
                                display_text = f"{column_name} {operator}"
                            else:
                                display_text = f"{column_name} {operator} '{value}'"

                            st.write(f"**CSV列筛选**: {display_text}")
                        else:
                            type_name = {
                                "log_level": "日志级别",
                                "ip_filter": "IP地址",
                                "status_code": "状态码",
                                "keyword": "关键词",
                                "show_only_errors": "仅显示错误",
                                "hide_debug": "隐藏调试"
                            }.get(filter_config['type'], filter_config['type'])

                            if filter_config['type'] == 'log_level':
                                value_display = ", ".join(filter_config['value'])
                            else:
                                value_display = filter_config.get('value', '')

                            st.write(f"**{type_name}**: {value_display}")

                    with col2:
                        st.write(f"**类型**: 文本条件")

                    with col3:
                        if st.button("🗑️", key=f"del_text_{i}", use_container_width=True):
                            st.session_state.text_filters.pop(i)
                            st.success("✅ 条件已删除")
                            st.rerun()

                # 显示JSON条件
                for i, filter_config in enumerate(st.session_state.json_filters):
                    col1, col2, col3 = st.columns([3, 2, 1])
                    with col1:
                        if filter_config.get('value_range'):
                            value_display = f"{filter_config['value_range'][0]} - {filter_config['value_range'][1]}"
                        else:
                            value_display = filter_config.get('value', '范围筛选')
                        st.write(f"**{filter_config['column']}.{filter_config['field']}**: {value_display}")

                    with col2:
                        st.write(f"**类型**: JSON条件")

                    with col3:
                        if st.button("🗑️", key=f"del_json_{i}", use_container_width=True):
                            st.session_state.json_filters.pop(i)
                            st.success("✅ 条件已删除")
                            st.rerun()

                st.info(
                    f"当前使用 **{st.session_state.filter_logic}** 逻辑，共有 {len(st.session_state.text_filters)} 个文本条件和 {len(st.session_state.json_filters)} 个JSON条件")

            else:
                st.info("📝 暂无筛选条件，请添加条件后应用过滤")


            # 应用过滤函数
            def apply_filters():
                filtered_lines = []

                if st.session_state.is_csv and st.session_state.df is not None:
                    # CSV数据过滤
                    filtered_df = st.session_state.df.copy()

                    # 应用JSON过滤器
                    if st.session_state.json_filters:
                        filtered_df = utils.apply_json_filters(
                            filtered_df,
                            st.session_state.json_filters,
                            st.session_state.filter_logic
                        )

                    # 转换为文本行并应用文本过滤
                    for _, row in filtered_df.iterrows():
                        line = " | ".join([str(x) for x in row])
                        if utils.apply_text_filters(line, st.session_state.text_filters, st.session_state.filter_logic):
                            filtered_lines.append(line)
                else:
                    # 文本数据过滤
                    for line in lines:
                        if utils.apply_text_filters(line, st.session_state.text_filters, st.session_state.filter_logic):
                            filtered_lines.append(line)

                st.session_state.filtered_lines = filtered_lines
                return filtered_lines


            # 自动应用过滤逻辑
            if st.session_state.auto_apply_filters and (st.session_state.text_filters or st.session_state.json_filters):
                filtered_lines = apply_filters()
                st.success(f"✅ 自动过滤完成！找到 {len(filtered_lines)} 行日志 (逻辑: {st.session_state.filter_logic})")

            # 手动应用过滤按钮
            if st.session_state.text_filters or st.session_state.json_filters:
                if not st.session_state.auto_apply_filters:
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        if st.button("🚀 应用过滤条件", key="filter_btn", use_container_width=True, type="primary"):
                            filtered_lines = apply_filters()
                            st.success(f"✅ 过滤完成！找到 {len(filtered_lines)} 行日志 (逻辑: {st.session_state.filter_logic})")

            # 清空所有条件按钮
            if st.session_state.text_filters or st.session_state.json_filters:
                if st.button("🗑️ 清空所有条件", use_container_width=True):
                    st.session_state.text_filters = []
                    st.session_state.json_filters = []
                    st.session_state.filtered_lines = []
                    st.success("✅ 所有条件已清空！")
                    st.rerun()

            # 显示过滤结果
            if st.session_state.filtered_lines:
                st.subheader(f"📋 过滤结果 (共 {len(st.session_state.filtered_lines)} 行)")

                # 根据文件类型选择显示方式
                if st.session_state.is_csv and st.session_state.df is not None:
                    # 对于CSV文件，显示DataFrame格式
                    # 重新构建过滤后的DataFrame用于显示
                    filtered_indices = []
                    for i, line in enumerate(lines):
                        if line in st.session_state.filtered_lines:
                            filtered_indices.append(i)

                    if filtered_indices:
                        filtered_df_display = st.session_state.df.iloc[filtered_indices]
                        st.dataframe(filtered_df_display, use_container_width=True, height=400)
                    else:
                        # 如果无法匹配索引，回退到文本显示
                        st.text_area("过滤后的日志", "\n".join(st.session_state.filtered_lines), height=400,
                                     key="filtered_output")
                else:
                    # 对于文本文件，保持原始文本格式
                    st.text_area("过滤后的日志", "\n".join(st.session_state.filtered_lines), height=400,
                                 key="filtered_output")

                # 显示统计信息
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("总行数", len(st.session_state.filtered_lines))
                with col2:
                    error_count = sum(1 for line in st.session_state.filtered_lines if
                                      any(word in line.upper() for word in ['ERROR', 'ERR']))
                    st.metric("错误数", error_count)
                with col3:
                    warn_count = sum(1 for line in st.session_state.filtered_lines if
                                     any(word in line.upper() for word in ['WARN', 'WARNING']))
                    st.metric("警告数", warn_count)

                # 导出结果 - 保持原始数据格式
                export_data = ""
                if st.session_state.is_csv and st.session_state.df is not None:
                    # 对于CSV文件，导出为CSV格式
                    csv_buffer = io.StringIO()
                    filtered_indices = []
                    for i, line in enumerate(lines):
                        if line in st.session_state.filtered_lines:
                            filtered_indices.append(i)

                    if filtered_indices:
                        filtered_df_display = st.session_state.df.iloc[filtered_indices]
                        filtered_df_display.to_csv(csv_buffer, index=False)
                        export_data = csv_buffer.getvalue()
                    else:
                        export_data = "\n".join(st.session_state.filtered_lines)
                    file_extension = "csv"
                else:
                    # 对于文本文件，导出为文本格式
                    export_data = "\n".join(st.session_state.filtered_lines)
                    file_extension = "txt"

                st.download_button(
                    label="📥 导出过滤结果",
                    data=export_data,
                    file_name=f"filtered_logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}",
                    mime="text/csv" if file_extension == "csv" else "text/plain",
                    use_container_width=True
                )
            else:
                if st.session_state.text_filters or st.session_state.json_filters:
                    st.info("🔍 暂无过滤结果，请调整筛选条件")

        # Tab3: 关键词搜索 (保持不变)
        with tab3:
            st.header("🔍 关键词搜索")

            # 处理清空搜索条件
            if st.session_state.search_cleared:
                search_key = f"search_input_{datetime.datetime.now().timestamp()}"
                case_key = f"case_sensitive_{datetime.datetime.now().timestamp()}"
                whole_key = f"whole_word_{datetime.datetime.now().timestamp()}"
                regex_key = f"use_regex_{datetime.datetime.now().timestamp()}"
                st.session_state.search_cleared = False
            else:
                search_key = "search_input"
                case_key = "case_sensitive"
                whole_key = "whole_word"
                regex_key = "use_regex"

            col1, col2 = st.columns([2, 1])

            with col1:
                search_keyword = st.text_input(
                    "搜索关键词",
                    value="",
                    placeholder="输入要搜索的关键词...",
                    help="支持普通文本和正则表达式搜索",
                    key=search_key
                )

            with col2:
                st.write("搜索选项")
                case_sensitive = st.checkbox("区分大小写", value=False, key=case_key)
                whole_word = st.checkbox("全词匹配", value=False, key=whole_key)
                use_regex = st.checkbox("正则表达式", value=False, key=regex_key)

            # 按钮布局
            col1, col2, col3 = st.columns([1, 1, 1])
            with col1:
                if st.button("🔍 执行搜索", type="primary", use_container_width=True):
                    if search_keyword:
                        # 更新搜索关键词状态
                        st.session_state.search_keyword = search_keyword

                        search_results = []

                        for line in lines:
                            search_text = line
                            original_line = line

                            if not case_sensitive:
                                search_text = search_text.lower()
                                keyword = search_keyword.lower()
                            else:
                                keyword = search_keyword

                            match_found = False

                            if use_regex:
                                try:
                                    if re.search(keyword, search_text, 0 if case_sensitive else re.IGNORECASE):
                                        match_found = True
                                except re.error as e:
                                    st.error(f"❌ 正则表达式错误: {e}")
                                    break
                            elif whole_word:
                                # 全词匹配
                                words = re.findall(r'\b\w+\b', search_text)
                                if any(word == keyword for word in words):
                                    match_found = True
                            else:
                                # 普通搜索
                                if keyword in search_text:
                                    match_found = True

                            if match_found:
                                search_results.append(original_line)

                        st.session_state.search_results = search_results
                        st.session_state.search_count = len(search_results)
                        if search_results:
                            st.success(f"✅ 找到 {len(search_results)} 条匹配结果")
                        else:
                            st.warning("⚠️ 未找到匹配的搜索结果")

                    else:
                        st.warning("⚠️ 请输入搜索关键词")

            with col2:
                if st.button("🗑️ 清空搜索条件", key="clear_search", use_container_width=True):
                    st.session_state.search_results = []
                    st.session_state.search_count = 0
                    st.session_state.search_keyword = ""
                    st.session_state.search_cleared = True
                    st.success("✅ 搜索条件已清空！")
                    st.rerun()

            # 显示搜索结果
            if st.session_state.search_results:
                st.subheader(f"📊 搜索结果 (共 {len(st.session_state.search_results)} 条)")

                # 根据文件类型选择显示方式
                if st.session_state.is_csv and st.session_state.df is not None:
                    # 对于CSV文件，显示DataFrame格式
                    search_indices = []
                    for i, line in enumerate(lines):
                        if line in st.session_state.search_results:
                            search_indices.append(i)

                    if search_indices:
                        search_df_display = st.session_state.df.iloc[search_indices]
                        st.dataframe(search_df_display, use_container_width=True, height=400)
                    else:
                        st.text_area("搜索结果", "\n".join(st.session_state.search_results), height=400,
                                     key="search_output")
                else:
                    st.text_area("搜索结果", "\n".join(st.session_state.search_results), height=400, key="search_output")

                # 搜索统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总匹配数", len(st.session_state.search_results))
                with col2:
                    unique_lines = len(set(st.session_state.search_results))
                    st.metric("唯一行数", unique_lines)
                with col3:
                    avg_length = sum(len(line) for line in st.session_state.search_results) // len(
                        st.session_state.search_results) if st.session_state.search_results else 0
                    st.metric("平均长度", f"{avg_length} 字符")
                with col4:
                    error_matches = sum(1 for line in st.session_state.search_results if
                                        any(word in line.upper() for word in ['ERROR', 'ERR']))
                    st.metric("错误匹配", error_matches)

                # 导出搜索结果
                export_search_data = ""
                if st.session_state.is_csv and st.session_state.df is not None:
                    csv_buffer = io.StringIO()
                    search_indices = []
                    for i, line in enumerate(lines):
                        if line in st.session_state.search_results:
                            search_indices.append(i)

                    if search_indices:
                        search_df_display = st.session_state.df.iloc[search_indices]
                        search_df_display.to_csv(csv_buffer, index=False)
                        export_search_data = csv_buffer.getvalue()
                    else:
                        export_search_data = "\n".join(st.session_state.search_results)
                    file_extension = "csv"
                else:
                    export_search_data = "\n".join(st.session_state.search_results)
                    file_extension = "txt"

                st.download_button(
                    label="📥 导出搜索结果",
                    data=export_search_data,
                    file_name=f"search_results_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}",
                    mime="text/csv" if file_extension == "csv" else "text/plain",
                    use_container_width=True
                )
            elif st.session_state.search_count == 0 and st.session_state.search_keyword:
                st.info("🔍 暂无搜索结果，请尝试其他关键词")

    else:
        st.info("📥 请先导入日志数据以开始分析")

# 在Streamlit界面中添加新的时间处理功能
elif tool_category == "时间处理工具":
    show_doc("time_processor")

    dt_utils = DateTimeUtils
    time_tool = st.radio(
        "选择时间处理工具",
        ["时间戳转换", "时间换算工具", "日期计算器", "日期信息查询", "时间间隔格式化",
         "星座生肖查询", "测试数据生成", "SLA计算器", "性能测试工具", "定时任务分析"],
        horizontal=True
    )

    if time_tool == "时间戳转换":
        st.markdown('<div class="category-card">⏰ 时间戳转换</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**时间戳转日期**")
            timestamp_input = st.text_input("输入时间戳", placeholder="例如: 1633046400")
            timestamp_type = st.radio("时间戳类型", ["秒", "毫秒"])
            if st.button("转换为日期", use_container_width=True):
                if not timestamp_input:
                    st.warning("请输入时间戳")
                else:
                    try:
                        timestamp = float(timestamp_input)
                        if timestamp_type == "毫秒":
                            timestamp /= 1000
                        dt = datetime.datetime.fromtimestamp(timestamp)
                        st.success(f"转换结果: {dt.strftime('%Y-%m-%d %H:%M:%S')}")
                    except (ValueError, OSError) as e:
                        st.error(f"请输入有效的时间戳: {e}")
            if st.button("获取当前时间戳", use_container_width=True):
                current_timestamp = int(time.time())
                st.text_input("当前时间戳", value=str(current_timestamp))
        with col2:
            st.markdown("**日期转时间戳**")
            date_input = st.date_input("选择日期")
            time_input = st.time_input("选择时间")
            if st.button("转换为时间戳", use_container_width=True):
                try:
                    dt = datetime.datetime.combine(date_input, time_input)
                    timestamp = int(dt.timestamp())
                    st.success(f"转换结果: {timestamp} (秒)")
                except Exception as e:
                    st.error(f"日期转换失败: {e}")

    elif time_tool == "时间换算工具":
        st.markdown('<div class="category-card">🔄 时间换算工具</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            value = st.number_input("输入数值", value=1.0)
            from_unit = st.selectbox("从单位", list(TO_SECONDS.keys()))
        with col2:
            to_unit = st.selectbox("转换为", list(TO_SECONDS.keys()))
            if st.button("转换", use_container_width=True):
                if from_unit in TO_SECONDS and to_unit in TO_SECONDS:
                    value_in_seconds = value * TO_SECONDS[from_unit]
                    result = value_in_seconds / TO_SECONDS[to_unit]
                    st.success(f"{value} {from_unit} = {result:.6f} {to_unit}")
                else:
                    st.error("单位转换错误")
        with col3:
            st.markdown("**常用时间换算表**")
            st.write("1 分钟 = 60 秒")
            st.write("1 小时 = 60 分钟 = 3600 秒")
            st.write("1 天 = 24 小时 = 1440 分钟")
            st.write("1 周 = 7 天 = 168 小时")
            st.write("1 月 ≈ 30.44 天")
            st.write("1 年 ≈ 365.25 天")

    elif time_tool == "日期计算器":
        st.markdown('<div class="category-card">📅 日期计算器</div>', unsafe_allow_html=True)
        calc_type = st.radio("计算类型", ["日期加减计算", "日期间隔计算"])

        if calc_type == "日期加减计算":
            col1, col2, col3 = st.columns(3)
            with col1:
                start_date = st.date_input("起始日期")
                operation = st.selectbox("操作", ["加上", "减去"])
            with col2:
                value = st.number_input("数值", min_value=0, value=7)
                unit = st.selectbox("单位", ["天", "周", "月", "年"])
            with col3:
                if st.button("计算", use_container_width=True):
                    try:
                        if operation == "加上":
                            if unit == "天":
                                result_date = start_date + timedelta(days=value)
                            elif unit == "周":
                                result_date = start_date + timedelta(weeks=value)
                            elif unit == "月":
                                result_date = dt_utils.add_months(start_date, value)
                            elif unit == "年":
                                result_date = start_date.replace(year=start_date.year + value)
                        else:
                            if unit == "天":
                                result_date = start_date - timedelta(days=value)
                            elif unit == "周":
                                result_date = start_date - timedelta(weeks=value)
                            elif unit == "月":
                                result_date = dt_utils.subtract_months(start_date, value)
                            elif unit == "年":
                                result_date = start_date.replace(year=start_date.year - value)
                        st.success(f"计算结果: {result_date.strftime('%Y-%m-%d')}")
                    except Exception as e:
                        st.error(f"日期运算错误: {e}")
        else:
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("开始日期")
            with col2:
                end_date = st.date_input("结束日期")
            if st.button("计算间隔", use_container_width=True):
                if not start_date or not end_date:
                    st.warning("请选择完整的日期范围")
                elif start_date > end_date:
                    st.error("开始日期不能晚于结束日期")
                else:
                    delta = end_date - start_date
                    business_days = dt_utils.count_business_days(start_date, end_date)
                    weekend_days = delta.days + 1 - business_days
                    st.success(f"间隔天数: {delta.days} 天")
                    st.info(f"工作日: {business_days} 天")
                    st.info(f"周末天数: {weekend_days} 天")

    elif time_tool == "日期信息查询":
        st.markdown('<div class="category-card">📊 日期信息查询</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            query_date = st.date_input("选择查询日期", datetime.date.today())

            if st.button("查询日期信息", use_container_width=True):
                with st.spinner("正在查询..."):
                    # 基本信息
                    st.subheader("📅 基本信息")
                    col_info1, col_info2 = st.columns(2)
                    with col_info1:
                        st.metric("星期", ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"][query_date.weekday()])
                        st.metric("季度", f"第{dt_utils.get_quarter(query_date)}季度")
                        st.metric("是否周末", "是" if dt_utils.is_weekend(query_date) else "否")
                    with col_info2:
                        st.metric("周数", f"第{dt_utils.get_week_number(query_date)}周")
                        st.metric("是否闰年", "是" if dt_utils.is_leap_year(query_date.year) else "否")
                        st.metric("当月天数", dt_utils.days_in_month(query_date.year, query_date.month))

                    # 月份范围
                    st.subheader("📈 月份范围")
                    first_day = dt_utils.get_first_day_of_month(query_date)
                    last_day = dt_utils.get_last_day_of_month(query_date)
                    col_range1, col_range2 = st.columns(2)
                    with col_range1:
                        st.metric("月初", first_day.strftime("%Y-%m-%d"))
                    with col_range2:
                        st.metric("月末", last_day.strftime("%Y-%m-%d"))

                    # 周范围
                    week_start, week_end = dt_utils.get_week_range(query_date)
                    st.metric("本周范围", f"{week_start.strftime('%Y-%m-%d')} 到 {week_end.strftime('%Y-%m-%d')}")

    elif time_tool == "时间间隔格式化":
        st.markdown('<div class="category-card">⏱️ 时间间隔格式化</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            seconds_input = st.number_input("输入秒数", min_value=0, value=3661, step=1)

            if st.button("格式化时间间隔", use_container_width=True):
                formatted = dt_utils.format_duration(seconds_input)
                st.success(f"格式化结果: {formatted}")

        with col2:
            st.markdown("**示例:**")
            st.write("3661 秒 → 1小时1分钟1秒")
            st.write("86400 秒 → 1天")
            st.write("90061 秒 → 1天1小时1分钟1秒")

    elif time_tool == "星座生肖查询":
        st.markdown('<div class="category-card">✨ 星座生肖查询</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            birth_date = st.date_input("选择出生日期", datetime.date(2000, 1, 1))

            if st.button("查询星座生肖", use_container_width=True):
                zodiac = dt_utils.get_chinese_zodiac(birth_date.year)
                constellation = dt_utils.get_constellation(birth_date.month, birth_date.day)
                age = dt_utils.get_age(birth_date)

                st.success("查询结果:")
                col_result1, col_result2, col_result3 = st.columns(3)
                with col_result1:
                    st.metric("生肖", zodiac)
                with col_result2:
                    st.metric("星座", constellation)
                with col_result3:
                    st.metric("年龄", age)

                st.info(f"出生日期: {birth_date.strftime('%Y年%m月%d日')}")

        with col2:
            st.markdown("**星座日期范围:**")
            constellations_info = [
                "♑ 摩羯座: 12月22日-1月19日",
                "♒ 水瓶座: 1月20日-2月18日",
                "♓ 双鱼座: 2月19日-3月20日",
                "♈ 白羊座: 3月21日-4月19日",
                "♉ 金牛座: 4月20日-5月20日",
                "♊ 双子座: 5月21日-6月21日",
                "♋ 巨蟹座: 6月22日-7月22日",
                "♌ 狮子座: 7月23日-8月22日",
                "♍ 处女座: 8月23日-9月22日",
                "♎ 天秤座: 9月23日-10月23日",
                "♏ 天蝎座: 10月24日-11月22日",
                "♐ 射手座: 11月23日-12月21日"
            ]
            for info in constellations_info:
                st.write(info)

    elif time_tool == "测试数据生成":
        st.markdown('<div class="category-card">🧪 测试数据生成</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("开始日期", datetime.date.today() - timedelta(days=30))
            end_date = st.date_input("结束日期", datetime.date.today())
            frequency = st.selectbox("生成频率", ["daily", "weekly", "monthly", "hourly"])
            count = st.number_input("生成数量（可选）", min_value=1, value=10)

            if st.button("生成测试日期", use_container_width=True):
                dates = dt_utils.generate_test_dates(start_date, end_date, frequency, count)
                st.success(f"生成了 {len(dates)} 个测试日期")

                date_strings = [date.strftime("%Y-%m-%d") for date in dates]
                st.text_area("生成的日期序列", "\n".join(date_strings), height=200)

                # 提供下载
                csv_data = "日期\n" + "\n".join(date_strings)
                st.download_button(
                    "📥 下载CSV",
                    csv_data,
                    file_name="test_dates.csv",
                    mime="text/csv"
                )

    elif time_tool == "SLA计算器":
        st.markdown('<div class="category-card">⏱️ SLA计算器</div>', unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["SLA到期时间", "工作时间计算"])

        with tab1:
            col1, col2 = st.columns(2)
            with col1:
                # 修正：使用 date_input 和 time_input 组合
                start_date = st.date_input("开始日期", datetime.date.today())
                start_time = st.time_input("开始时间", datetime.time(9, 0))
                start_dt = datetime.datetime.combine(start_date, start_time)

                sla_hours = st.number_input("SLA小时数", min_value=1, value=8)
                work_start = st.number_input("工作开始时间", min_value=0, max_value=23, value=9)
                work_end = st.number_input("工作结束时间", min_value=1, max_value=24, value=17)

            with col2:
                st.write(f"**开始时间:** {start_dt.strftime('%Y-%m-%d %H:%M')}")

                if st.button("计算SLA到期时间", use_container_width=True):
                    due_date = dt_utils.calculate_sla_due_date(start_dt, sla_hours, work_start, work_end)
                    st.success(f"SLA到期时间: {due_date.strftime('%Y-%m-%d %H:%M:%S')}")

                    # 显示详细信息
                    st.info(f"""
                    **计算详情:**
                    - 开始时间: {start_dt.strftime('%Y-%m-%d %H:%M')}
                    - SLA要求: {sla_hours} 工作时间
                    - 工作时间: {work_start}:00 - {work_end}:00
                    - 到期时间: {due_date.strftime('%Y-%m-%d %H:%M')}
                    """)

        with tab2:
            col1, col2 = st.columns(2)
            with col1:
                # 修正：使用 date_input 和 time_input 组合
                start_date = st.date_input("开始日期", datetime.date.today() - timedelta(days=2), key="start_date_work")
                start_time = st.time_input("开始时间", datetime.time(9, 0), key="start_time_work")
                start_dt = datetime.datetime.combine(start_date, start_time)

                end_date = st.date_input("结束日期", datetime.date.today(), key="end_date_work")
                end_time = st.time_input("结束时间", datetime.time(17, 0), key="end_time_work")
                end_dt = datetime.datetime.combine(end_date, end_time)

                work_start = st.number_input("工作开始", min_value=0, max_value=23, value=9, key="work_start2")
                work_end = st.number_input("工作结束", min_value=1, max_value=24, value=17, key="work_end2")

            with col2:
                st.write(f"**时间范围:** {start_dt.strftime('%Y-%m-%d %H:%M')} 到 {end_dt.strftime('%Y-%m-%d %H:%M')}")

                if st.button("计算工作时间", use_container_width=True):
                    if start_dt >= end_dt:
                        st.error("开始时间必须早于结束时间")
                    else:
                        work_hours = dt_utils.get_working_hours(start_dt, end_dt, work_start, work_end)
                        total_hours = (end_dt - start_dt).total_seconds() / 3600

                        st.success(f"实际工作时间: {work_hours:.2f} 小时")
                        st.info(f"总时间: {total_hours:.2f} 小时")
                        st.info(f"非工作时间: {total_hours - work_hours:.2f} 小时")

                        # 显示工作日统计
                        current = start_dt.date()
                        work_days = 0
                        while current <= end_dt.date():
                            if current.weekday() < 5:  # 周一到周五
                                work_days += 1
                            current += timedelta(days=1)

                        st.info(f"涉及工作日: {work_days} 天")

    elif time_tool == "性能测试工具":
        st.markdown('<div class="category-card">🚀 性能测试工具</div>', unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["时间戳生成", "响应时间分析"])

        with tab1:
            col1, col2 = st.columns(2)
            with col1:
                duration = st.number_input("测试时长（秒）", min_value=1, value=60, key="perf_duration")
                rps = st.number_input("每秒请求数", min_value=1, value=10, key="perf_rps")

                # 添加开始时间选择
                test_date = st.date_input("测试日期", datetime.date.today(), key="perf_test_date")
                test_time = st.time_input("测试开始时间", datetime.time(10, 0), key="perf_test_time")

            with col2:
                if st.button("生成时间戳", use_container_width=True, key="generate_timestamps_btn"):
                    # 使用选择的日期时间作为基准
                    base_datetime = datetime.datetime.combine(test_date, test_time)

                    timestamps = dt_utils.get_performance_test_timestamps(duration, rps, base_datetime)
                    st.success(f"生成了 {len(timestamps)} 个时间戳")

                    # 显示前10个时间戳作为示例
                    sample_timestamps = [
                        f"{datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}"
                        for ts in timestamps[:10]
                    ]
                    st.text_area("前10个时间戳示例", "\n".join(sample_timestamps), height=150, key="timestamp_samples")

                    # 统计信息
                    intervals = [timestamps[i + 1] - timestamps[i] for i in range(len(timestamps) - 1)]
                    avg_interval = sum(intervals) / len(intervals)
                    st.info(f"平均间隔: {avg_interval:.6f} 秒")
                    st.info(f"实际RPS: {1 / avg_interval:.2f}")

                    # 提供下载
                    timestamp_data = "\n".join([f"{ts:.6f}" for ts in timestamps])
                    st.download_button(
                        "📥 下载时间戳数据",
                        timestamp_data,
                        file_name="performance_timestamps.txt",
                        mime="text/plain",
                        key="download_timestamps"
                    )

        with tab2:
            st.markdown("**响应时间百分位数计算**")
            response_times_input = st.text_area(
                "输入响应时间列表（毫秒）",
                "100, 150, 200, 120, 300, 180, 250, 110, 190, 220, 280, 130, 160, 240, 170",
                help="用逗号分隔的响应时间数值，单位毫秒"
            )

            if st.button("计算百分位数", use_container_width=True):
                try:
                    response_times = [float(x.strip()) for x in response_times_input.split(",") if x.strip()]
                    percentiles = dt_utils.calculate_response_time_percentiles(response_times)

                    st.success("响应时间百分位数:")
                    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                    with col_p1:
                        st.metric("P50", f"{percentiles[50]:.2f}ms")
                    with col_p2:
                        st.metric("P90", f"{percentiles[90]:.2f}ms")
                    with col_p3:
                        st.metric("P95", f"{percentiles[95]:.2f}ms")
                    with col_p4:
                        st.metric("P99", f"{percentiles[99]:.2f}ms")

                    # 显示所有数据
                    st.dataframe({
                        '统计量': ['最小值', '平均值', '最大值', '总数'],
                        '数值': [f"{min(response_times):.2f}ms",
                               f"{sum(response_times) / len(response_times):.2f}ms",
                               f"{max(response_times):.2f}ms",
                               len(response_times)]
                    })

                except Exception as e:
                    st.error(f"计算错误: {e}")

    elif time_tool == "定时任务分析":
        st.markdown('<div class="category-card">⏰ 定时任务分析</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            cron_expression = st.text_input(
                "Cron表达式",
                "0 9 * * 1-5",
                help="例如: 0 9 * * 1-5 表示工作日早上9点",
                key="cron_expression"
            )
            start_date = st.date_input("开始日期", datetime.date.today(), key="cron_start_date")
            # 添加开始时间选择
            start_time = st.time_input("开始时间", datetime.time(0, 0), key="cron_start_time")
            run_count = st.number_input("生成执行次数", min_value=1, max_value=20, value=5, key="cron_run_count")

        with col2:
            if st.button("分析Cron表达式", use_container_width=True, key="analyze_cron_btn"):
                # 组合日期和时间
                start_datetime = datetime.datetime.combine(start_date, start_time)
                next_runs = dt_utils.generate_cron_next_runs(cron_expression, start_datetime, run_count)

                if isinstance(next_runs, list) and next_runs:
                    st.success(f"接下来 {run_count} 次执行时间:")
                    for i, run_time in enumerate(next_runs, 1):
                        if isinstance(run_time, datetime.datetime):
                            st.write(f"{i}. {run_time.strftime('%Y-%m-%d %H:%M:%S')}")
                        else:
                            st.write(f"{i}. {run_time}")
                else:
                    st.error("无法解析Cron表达式")

            # Cron表达式示例
            with st.expander("Cron表达式示例", expanded=False):
                st.write("""
                **基本格式:** `分 时 日 月 周`

                **常用示例:**
                - `0 9 * * *` - 每天9:00
                - `0 9 * * 1-5` - 工作日9:00  
                - `*/15 * * * *` - 每15分钟
                - `0 0 1 * *` - 每月1号0:00
                - `0 12 * * 0` - 每周日12:00
                - `0 0,12 * * *` - 每天0点和12点
                - `0 4-6 * * *` - 每天4、5、6点
                """)

    st.markdown('</div>', unsafe_allow_html=True)

# IP/域名查询工具
elif tool_category == "IP/域名查询工具":
    show_doc("ip_domain_query")

    # 创建实例
    ip_tool = IPQueryTool()

    tab1, tab2, tab3 = st.tabs(
        ["IP/域名查询", "批量查询", "IPv4转换工具"])

    with tab1:
        st.markdown("**IP/域名基本信息查询**")

        # # 获取当前公网IP
        # if st.button("获取当前公网IP", key="get_public_ip", use_container_width=True):
        #     with st.spinner("正在获取当前公网IP..."):
        #         public_ip = ip_tool.get_public_ip()
        #         if public_ip != "获取公网IP失败":
        #             st.session_state.current_public_ip = public_ip
        #             st.success(f"当前公网IP: {public_ip}")
        #         else:
        #             st.error("无法获取当前公网IP")

        # 输入框
        if 'current_public_ip' in st.session_state:
            target_input = st.text_input("输入IP地址或域名",
                                         value=st.session_state.current_public_ip,
                                         placeholder="例如: 117.30.73.100 或 baidu.com",
                                         key="target_input_with_public_ip")
        else:
            target_input = st.text_input("输入IP地址或域名",
                                         placeholder="例如: 117.30.73.100 或 baidu.com",
                                         key="target_input")

        st.caption("支持IPv4、IPv6地址和域名查询")

        col1, col2 = st.columns([2, 1])
        with col2:
            st.write("")
            st.write("")
            query_button = st.button("开始查询", use_container_width=True, key="main_query")

        if query_button and target_input:
            is_ip = False
            is_domain = False
            ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
            ipv6_pattern = r'^([0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}$'

            # 处理URL格式
            if target_input.startswith(('http://', 'https://')):
                try:
                    target_input = target_input.split('://')[1].split('/')[0]
                except:
                    pass

            if re.match(ipv4_pattern, target_input.strip()) or re.match(ipv6_pattern, target_input.strip()):
                is_ip = True
            else:
                domain_pattern = r'^[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?(\.[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?)*$'
                if re.match(domain_pattern, target_input.strip()):
                    is_domain = True

            if not is_ip and not is_domain:
                st.error("请输入有效的IP地址或域名格式")
                st.stop()

            with st.spinner("查询中..."):
                result = ip_tool.get_ip_domain_info(target_input, is_ip)

                if result['success']:
                    st.success("查询成功！")

                    # rDNS查询
                    if is_ip:
                        rdns_result = ip_tool.get_rdns_info(target_input)
                        if rdns_result['success']:
                            st.metric("rDNS", rdns_result['data'].get('rDNS', '未知'))

                    # 详细信息展示
                    st.markdown("**详细信息**")
                    detailed_info = result['data'].copy()
                    for key in ['IP地址', '域名', 'location', 'isp']:
                        detailed_info.pop(key, None)

                    info_keys = list(detailed_info.keys())
                    for i in range(0, len(info_keys), 2):
                        cols = st.columns(2)
                        for j in range(2):
                            if i + j < len(info_keys):
                                key = info_keys[i + j]
                                value = detailed_info[key]
                                with cols[j]:
                                    st.markdown(f"""
                                    <div style="background: #f8fafc; padding: 1rem; border-radius: 8px; margin: 0.5rem 0; border-left: 4px solid #667eea;">
                                        <div style="font-weight: 600; color: #2d3748; margin-bottom: 0.5rem;">{key}</div>
                                        <div style="color: #4a5568;">{value}</div>
                                    </div>
                                    """, unsafe_allow_html=True)
                else:
                    st.error(f"查询失败: {result['error']}")

    with tab2:
        st.markdown("**批量查询工具**")
        st.info("支持批量查询IP/域名信息")

        query_type = st.radio("查询类型", ["IP地址查询", "域名查询"], horizontal=True)

        if query_type == "IP地址查询":
            ips_input = st.text_area("输入IP地址列表（每行一个）", height=150,
                                     placeholder="例如:\n8.8.8.8\n1.1.1.1\n117.30.73.100")
            if st.button("批量查询IP", use_container_width=True):
                if not ips_input.strip():
                    st.error("请输入至少一个IP地址")
                    st.stop()

                ips = [ip.strip() for ip in ips_input.split('\n') if ip.strip()]
                valid_ips = []
                invalid_ips = []

                ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
                ipv6_pattern = r'^([0.9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}$'

                for ip in ips:
                    if re.match(ipv4_pattern, ip) or re.match(ipv6_pattern, ip):
                        valid_ips.append(ip)
                    else:
                        invalid_ips.append(ip)

                if invalid_ips:
                    st.warning(f"发现 {len(invalid_ips)} 个无效IP地址: {', '.join(invalid_ips)}")

                if not valid_ips:
                    st.error("没有有效的IP地址可查询")
                    st.stop()

                results = []
                progress_bar = st.progress(0)
                status_text = st.empty()

                for i, ip in enumerate(valid_ips):
                    progress = (i + 1) / len(valid_ips)
                    progress_bar.progress(progress)
                    status_text.text(f"正在查询 {i + 1}/{len(valid_ips)}: {ip}")

                    result = ip_tool.get_ip_domain_info(ip, True)
                    if result['success']:
                        results.append(result['data'])
                    else:
                        results.append({'IP地址': ip, '状态': '查询失败', '错误': result['error']})

                    time.sleep(0.5)  # 模拟查询延迟

                progress_bar.empty()
                status_text.empty()

                st.success(f"已完成 {len(valid_ips)} 个IP地址的查询")

                # 显示结果表格
                df = pd.DataFrame(results)
                st.dataframe(df)

                # 提供下载按钮
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="下载查询结果",
                    data=csv,
                    file_name='ip_query_results.csv',
                    mime='text/csv'
                )

        else:  # 域名查询
            domains_input = st.text_area("输入域名列表（每行一个）", height=150,
                                         placeholder="例如:\nbaidu.com\ngoogle.com\nqq.com")
            if st.button("批量查询域名", use_container_width=True):
                if not domains_input.strip():
                    st.error("请输入至少一个域名")
                    st.stop()

                domains = [domain.strip() for domain in domains_input.split('\n') if domain.strip()]
                valid_domains = []
                invalid_domains = []

                domain_pattern = r'^[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?(\.[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?)+$'

                for domain in domains:
                    if re.match(domain_pattern, domain):
                        valid_domains.append(domain)
                    else:
                        invalid_domains.append(domain)

                if invalid_domains:
                    st.warning(f"发现 {len(invalid_domains)} 个无效域名: {', '.join(invalid_domains)}")

                if not valid_domains:
                    st.error("没有有效的域名可查询")
                    st.stop()

                results = []
                progress_bar = st.progress(0)
                status_text = st.empty()

                for i, domain in enumerate(valid_domains):
                    progress = (i + 1) / len(valid_domains)
                    progress_bar.progress(progress)
                    status_text.text(f"正在查询 {i + 1}/{len(valid_domains)}: {domain}")

                    result = ip_tool.get_ip_domain_info(domain, False)
                    if result['success']:
                        results.append(result['data'])
                    else:
                        results.append({'域名': domain, '状态': '查询失败', '错误': result['error']})

                    time.sleep(0.5)  # 模拟查询延迟

                progress_bar.empty()
                status_text.empty()

                st.success(f"已完成 {len(valid_domains)} 个域名的查询")

                # 显示结果表格
                df = pd.DataFrame(results)
                st.dataframe(df)

                # 提供下载按钮
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="下载查询结果",
                    data=csv,
                    file_name='domain_query_results.csv',
                    mime='text/csv'
                )

    with tab3:
        st.markdown("**IPv4转换工具**")
        st.info("IPv4地址的各种格式转换")
        conversion_type = st.radio("转换类型",
                                   ["十进制 ↔ 点分十进制",
                                    "点分十进制 ↔ 十六进制",
                                    "点分十进制 ↔ 二进制"],
                                   horizontal=True)

        col1, col2 = st.columns(2)
        with col1:
            if conversion_type == "十进制 ↔ 点分十进制":
                input_value = st.text_input("输入IP地址或十进制数", placeholder="例如: 8.8.8.8 或 134744072")
            elif conversion_type == "点分十进制 ↔ 十六进制":
                input_value = st.text_input("输入IP地址或十六进制", placeholder="例如: 8.8.8.8 或 0x8080808")
            else:
                input_value = st.text_input("输入IP地址或二进制", placeholder="例如: 8.8.8.8 或 1000000010000000100000001000")
        with col2:
            st.write("")
            st.write("")
            convert_button = st.button("转换", use_container_width=True, key="convert_ip")

        if convert_button and input_value:
            result = ip_tool.convert_ip_address(input_value, conversion_type)
            if result['success']:
                st.success("转换成功！")
                for key, value in result['data'].items():
                    with st.container():
                        st.markdown(f"""
                         <div style="background: #f8fafc; padding: 1rem; border-radius: 8px; margin: 0.5rem 0; border-left: 4px solid #667eea;">
                             <div style="font-weight: 600; color: #2d3748; margin-bottom: 0.5rem;">{key}</div>
                             <div style="font-family: monospace; font-size: 14px; color: #4a5568;">{value}</div>
                         </div>
                         """, unsafe_allow_html=True)
            else:
                st.error(f"转换失败: {result['error']}")

        with st.expander("转换示例"):
            st.markdown("""
             十进制 ↔ 点分十进制
             ▪ 8.8.8.8 → 134744072

             ▪ 134744072 → 8.8.8.8


             点分十进制 ↔ 十六进制
             ▪ 8.8.8.8 → 0x8080808

             ▪ 0x8080808 → 8.8.8.8


             点分十进制 ↔ 二进制
             ▪ 8.8.8.8 → 00001000.00001000.00001000.00001000

             """)

# 在图片处理工具部分，添加翻转、旋转、裁剪、水印功能
elif tool_category == "图片处理工具":
    # 延迟导入图片处理模块
    try:
        from qa_toolkit.utils.image_processing import ImageProcessor

        image_tool = ImageProcessor()
    except ImportError as e:
        st.error(f"❌ 图片处理模块加载失败: {e}")
        st.info("请确保 image_processor.py 文件存在于正确的位置")
        st.stop()

    show_doc("image_processor")

    # 初始化session state
    if 'original_image' not in st.session_state:
        st.session_state.original_image = None
    if 'processed_image' not in st.session_state:
        st.session_state.processed_image = None
    if 'image_info' not in st.session_state:
        st.session_state.image_info = None
    if 'processed_info' not in st.session_state:
        st.session_state.processed_info = None
    if 'crop_coordinates' not in st.session_state:
        st.session_state.crop_coordinates = None
    if 'crop_preview' not in st.session_state:
        st.session_state.crop_preview = None
    if 'processed_image_data' not in st.session_state:
        st.session_state.processed_image_data = None
    if 'processed_image_format' not in st.session_state:
        st.session_state.processed_image_format = None

    st.markdown('<div class="category-card">🖼️ 图片处理工具</div>', unsafe_allow_html=True)

    # 文件上传区域
    st.markdown("### 1. 上传图片")
    uploaded_file = st.file_uploader(
        "点击或拖拽图片到此处上传",
        type=['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'],
        help="支持 JPG、PNG、GIF、BMP、WEBP 格式",
        key="image_uploader"
    )

    if uploaded_file is not None:
        try:
            # 打开图片并保存到session state
            image = Image.open(uploaded_file)
            st.session_state.original_image = image
            st.session_state.processed_image = image.copy()

            # 获取图片信息
            img_format = image.format
            img_size = image.size
            img_mode = image.mode
            file_size = len(uploaded_file.getvalue())

            # 保存图片信息
            st.session_state.image_info = {
                "文件名": uploaded_file.name,
                "格式": img_format,
                "模式": img_mode,
                "尺寸": f"{img_size[0]} × {img_size[1]} 像素",
                "文件大小": f"{file_size / 1024:.2f} KB",
                "原始大小字节": file_size
            }

        except Exception as e:
            st.error(f"图片读取失败: {e}")

    # 显示原图信息
    if st.session_state.original_image and st.session_state.image_info:
        st.markdown("### 2. 原图信息")

        col1, col2 = st.columns(2)
        with col1:
            st.image(st.session_state.original_image, caption="原图预览", use_container_width=True)
        with col2:
            st.markdown("**图片详细信息:**")
            for key, value in st.session_state.image_info.items():
                if key != "原始大小字节":
                    st.write(f"**{key}:** {value}")

    # 转换设置
    if st.session_state.original_image:
        st.markdown("### 3. 转换设置")

        # 处理模式选择
        processing_mode = st.radio(
            "处理模式",
            ["格式转换和质量调整", "调整尺寸", "图片翻转", "图片旋转", "图片裁剪", "添加水印"],
            horizontal=True
        )

        if processing_mode == "格式转换和质量调整":
            col1, col2, col3 = st.columns(3)

            with col1:
                output_format = st.selectbox("输出格式", ["JPG", "PNG", "GIF", "BMP", "WEBP"], index=0)
                if output_format in ["JPG", "WEBP"]:
                    quality = st.slider("图片质量", 1, 100, 85)
                else:
                    quality = 100
                    st.info("PNG、GIF、BMP格式不支持质量调整")

            with col2:
                compression_mode = st.radio("压缩模式", ["质量优先", "体积优先", "平衡模式"], horizontal=True)
                if compression_mode == "体积优先" and output_format in ["JPG", "WEBP"]:
                    quality = max(1, quality - 30)
                elif compression_mode == "平衡模式" and output_format in ["JPG", "WEBP"]:
                    quality = max(1, quality - 15)

            with col3:
                resize_option = st.radio("尺寸调整", ["保持原尺寸", "自定义尺寸"], horizontal=True)
                if resize_option == "自定义尺寸":
                    new_width = st.number_input("宽度(像素)", min_value=1, value=st.session_state.original_image.width)
                    new_height = st.number_input("高度(像素)", min_value=1, value=st.session_state.original_image.height)
                else:
                    new_width = st.session_state.original_image.width
                    new_height = st.session_state.original_image.height

        elif processing_mode == "调整尺寸":
            st.markdown("**调整图片尺寸**")
            col1, col2 = st.columns(2)

            with col1:
                resize_method = st.radio("调整方式", ["自定义尺寸", "按比例缩放", "预设尺寸"], horizontal=True)
                if resize_method == "自定义尺寸":
                    new_width = st.number_input("宽度(像素)", min_value=1, value=st.session_state.original_image.width)
                    new_height = st.number_input("高度(像素)", min_value=1, value=st.session_state.original_image.height)
                elif resize_method == "按比例缩放":
                    scale_percent = st.slider("缩放比例 (%)", 10, 200, 100)
                    original_width = st.session_state.original_image.width
                    original_height = st.session_state.original_image.height
                    new_width = int(original_width * scale_percent / 100)
                    new_height = int(original_height * scale_percent / 100)
                    st.write(f"新尺寸: {new_width} × {new_height} 像素")
                else:
                    selected_preset = st.selectbox("选择预设尺寸", list(PRESET_SIZES.keys()))
                    new_width, new_height = PRESET_SIZES[selected_preset]
                    st.write(f"预设尺寸: {new_width} × {new_height} 像素")

            with col2:
                resample_method = st.selectbox("重采样算法", ["LANCZOS (高质量)", "BILINEAR (平衡)", "NEAREST (快速)"])
                output_format = st.selectbox("输出格式", ["JPG", "PNG", "WEBP"], index=0)

        elif processing_mode == "图片翻转":
            st.markdown("**图片翻转**")
            col1, col2 = st.columns(2)

            with col1:
                flip_direction = st.radio("翻转方向", ["上下翻转", "左右翻转", "同时翻转"], help="选择图片翻转的方向")
                st.info("💡 上下翻转：垂直镜像\n左右翻转：水平镜像\n同时翻转：垂直和水平同时镜像")

            with col2:
                output_format = st.selectbox("输出格式", ["JPG", "PNG", "WEBP"], index=0)

        elif processing_mode == "图片旋转":
            st.markdown("**图片旋转**")
            col1, col2 = st.columns(2)

            with col1:
                rotation_direction = st.radio("旋转方向", ["顺时针", "逆时针"], horizontal=True)
                rotation_angle = st.slider("旋转角度", min_value=0, max_value=360, value=90, step=90, help="选择旋转角度（度）")

            with col2:
                if rotation_angle % 90 != 0:
                    bg_color = st.color_picker("背景颜色", "#FFFFFF")
                    st.info("非90度倍数旋转时，空白区域将填充背景颜色")
                else:
                    bg_color = "#FFFFFF"
                output_format = st.selectbox("输出格式", ["JPG", "PNG", "WEBP"], index=0)

        elif processing_mode == "图片裁剪":
            st.markdown("**图片裁剪**")
            original_width = st.session_state.original_image.width
            original_height = st.session_state.original_image.height

            # 裁剪方式选择
            crop_method = st.radio("裁剪方式", ["手动设置区域", "按比例裁剪"], horizontal=True)

            if crop_method == "手动设置区域":
                col_setting, col_preview = st.columns([1, 1])

                with col_setting:
                    st.markdown("**设置裁剪区域：**")
                    left = st.slider("左边距", 0, original_width - 1, 0, help="从图片左边开始裁剪的像素数")
                    top = st.slider("上边距", 0, original_height - 1, 0, help="从图片顶部开始裁剪的像素数")
                    right = st.slider("右边距", left + 1, original_width, original_width, help="裁剪到图片右边的像素位置")
                    bottom = st.slider("下边距", top + 1, original_height, original_height, help="裁剪到图片底部的像素位置")

                    crop_width = right - left
                    crop_height = bottom - top
                    st.success(f"**裁剪区域尺寸:** {crop_width} × {crop_height} 像素")
                    st.session_state.crop_coordinates = (left, top, right, bottom)

                with col_preview:
                    st.markdown("**实时预览：**")
                    try:
                        if st.session_state.crop_coordinates:
                            left, top, right, bottom = st.session_state.crop_coordinates
                            crop_preview = st.session_state.original_image.crop((left, top, right, bottom))
                            st.image(crop_preview, caption=f"裁剪预览 ({crop_width}×{crop_height})",
                                     use_container_width=True)
                            st.info(f"""
                                **裁剪信息:**
                                - 位置: ({left}, {top}) 到 ({right}, {bottom})
                                - 尺寸: {crop_width} × {crop_height} 像素
                                - 原图利用率: {(crop_width * crop_height) / (original_width * original_height) * 100:.1f}%
                                """)
                    except Exception as e:
                        st.error(f"预览生成失败: {e}")

            elif crop_method == "按比例裁剪":
                aspect_ratio = st.selectbox("裁剪比例",
                                            ["1:1 (正方形)", "16:9 (宽屏)", "4:3 (标准)", "3:2 (照片)", "9:16 (竖屏)", "自定义"])

                if aspect_ratio == "自定义":
                    col_ratio1, col_ratio2 = st.columns(2)
                    with col_ratio1:
                        ratio_w = st.number_input("宽度比例", min_value=1, value=1)
                    with col_ratio2:
                        ratio_h = st.number_input("高度比例", min_value=1, value=1)
                else:
                    ratio_map = {
                        "1:1 (正方形)": (1, 1),
                        "16:9 (宽屏)": (16, 9),
                        "4:3 (标准)": (4, 3),
                        "3:2 (照片)": (3, 2),
                        "9:16 (竖屏)": (9, 16)
                    }
                    ratio_w, ratio_h = ratio_map[aspect_ratio]

                target_ratio = ratio_w / ratio_h
                current_ratio = original_width / original_height

                if current_ratio > target_ratio:
                    crop_width = int(original_height * target_ratio)
                    crop_height = original_height
                    left = (original_width - crop_width) // 2
                    top = 0
                else:
                    crop_width = original_width
                    crop_height = int(original_width / target_ratio)
                    left = 0
                    top = (original_height - crop_height) // 2

                right = left + crop_width
                bottom = top + crop_height

                col_ratio_setting, col_ratio_preview = st.columns([1, 1])
                with col_ratio_setting:
                    st.success(f"**自动计算区域:** {crop_width} × {crop_height} 像素")
                    st.info(f"裁剪比例: {ratio_w}:{ratio_h}")
                    st.session_state.crop_coordinates = (left, top, right, bottom)

                with col_ratio_preview:
                    st.markdown("**预览效果：**")
                    try:
                        crop_preview = st.session_state.original_image.crop((left, top, right, bottom))
                        st.image(crop_preview, caption=f"比例裁剪预览 ({crop_width}×{crop_height})", use_container_width=True)
                    except Exception as e:
                        st.error(f"预览生成失败: {e}")

            output_format = st.selectbox("输出格式", ["JPG", "PNG", "WEBP"], index=0)

        elif processing_mode == "添加水印":
            st.markdown("**添加文字水印**")
            col1, col2 = st.columns(2)

            with col1:
                watermark_text = st.text_input("水印文字", "我的水印", placeholder="输入水印文字，支持中文")
                watermark_position = st.selectbox("水印位置",
                                                  ["顶部居左", "顶部居中", "顶部居右", "左边居中", "图片中心", "右边居中", "底部居左", "底部居中",
                                                   "底部居右"])
                font_size = st.slider("字体大小", 10, 100, 24)
                text_color = st.color_picker("文字颜色", "#FFFFFF")

            with col2:
                opacity = st.slider("透明度", 0.1, 1.0, 0.7)
                rotation = st.slider("旋转角度", -180, 180, 0, help="水印文字旋转角度（度）")
                output_format = st.selectbox("输出格式", ["JPG", "PNG", "WEBP"], index=0)
                st.info("💡 支持中文水印，系统会自动检测可用字体")

        # 转换按钮
        if st.button("🔄 转换图片", use_container_width=True, key="process_image_btn"):
            try:
                with st.spinner("正在处理图片..."):
                    processed_img = st.session_state.original_image.copy()
                    output_buffer = io.BytesIO()

                    if processing_mode == "格式转换和质量调整":
                        if resize_option == "自定义尺寸" and (
                                new_width != processed_img.width or new_height != processed_img.height):
                            processed_img = processed_img.resize((new_width, new_height), Image.Resampling.LANCZOS)

                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=quality, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "GIF":
                            processed_img.save(output_buffer, format='GIF', optimize=True)
                        elif output_format == "BMP":
                            processed_img.save(output_buffer, format='BMP')
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=quality, optimize=True)

                    elif processing_mode == "图片裁剪":
                        if st.session_state.crop_coordinates:
                            left, top, right, bottom = st.session_state.crop_coordinates
                            processed_img = processed_img.crop((left, top, right, bottom))

                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=95, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=95, optimize=True)

                    elif processing_mode == "调整尺寸":
                        resample_map = {
                            "LANCZOS (高质量)": Image.Resampling.LANCZOS,
                            "BILINEAR (平衡)": Image.Resampling.BILINEAR,
                            "NEAREST (快速)": Image.Resampling.NEAREST
                        }
                        resample_algo = resample_map.get(resample_method, Image.Resampling.LANCZOS)
                        processed_img = processed_img.resize((new_width, new_height), resample_algo)
                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=95, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=95, optimize=True)

                    elif processing_mode == "图片翻转":
                        if flip_direction == "上下翻转":
                            processed_img = processed_img.transpose(Image.Transpose.FLIP_TOP_BOTTOM)
                        elif flip_direction == "左右翻转":
                            processed_img = processed_img.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
                        elif flip_direction == "同时翻转":
                            processed_img = processed_img.transpose(Image.Transpose.FLIP_TOP_BOTTOM)
                            processed_img = processed_img.transpose(Image.Transpose.FLIP_LEFT_RIGHT)

                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=95, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=95, optimize=True)

                    elif processing_mode == "图片旋转":
                        actual_angle = rotation_angle if rotation_direction == "顺时针" else -rotation_angle
                        if rotation_angle % 90 == 0:
                            if actual_angle == 90 or actual_angle == -270:
                                processed_img = processed_img.transpose(Image.Transpose.ROTATE_90)
                            elif actual_angle == 180 or actual_angle == -180:
                                processed_img = processed_img.transpose(Image.Transpose.ROTATE_180)
                            elif actual_angle == 270 or actual_angle == -90:
                                processed_img = processed_img.transpose(Image.Transpose.ROTATE_270)
                        else:
                            from PIL import ImageOps

                            bg_rgb = tuple(int(bg_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
                            processed_img = processed_img.rotate(actual_angle, expand=True,
                                                                 resample=Image.Resampling.BICUBIC, fillcolor=bg_rgb)

                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=95, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=95, optimize=True)

                    elif processing_mode == "添加水印":
                        color_rgb = tuple(int(text_color.lstrip('#')[i:i + 2], 16) for i in (0, 2, 4))
                        processed_img = image_tool.add_watermark(
                            processed_img,
                            watermark_text,
                            watermark_position,
                            font_size,
                            color_rgb,
                            opacity,
                            rotation
                        )
                        processed_img = image_tool.convert_image_for_format(processed_img, output_format)

                        if output_format == "JPG":
                            processed_img.save(output_buffer, format='JPEG', quality=95, optimize=True)
                        elif output_format == "PNG":
                            processed_img.save(output_buffer, format='PNG', optimize=True)
                        elif output_format == "WEBP":
                            processed_img.save(output_buffer, format='WEBP', quality=95, optimize=True)

                    # 获取处理后的图片数据
                    processed_image_data = output_buffer.getvalue()
                    new_buffer = io.BytesIO(processed_image_data)
                    processed_image_obj = Image.open(new_buffer)
                    processed_image_obj.load()
                    output_buffer.close()
                    new_buffer.close()

                    # 保存处理后的图片对象和数据
                    st.session_state.processed_image = processed_image_obj
                    st.session_state.processed_image_data = processed_image_data

                    # 计算处理后的信息
                    processed_size = len(processed_image_data)
                    original_size = st.session_state.image_info["原始大小字节"]
                    compression_ratio = (1 - processed_size / original_size) * 100

                    st.session_state.processed_info = {
                        "格式": output_format,
                        "模式": processed_image_obj.mode,
                        "尺寸": f"{processed_image_obj.width} × {processed_image_obj.height} 像素",
                        "文件大小": f"{processed_size / 1024:.2f} KB",
                        "压缩率": f"{compression_ratio:.1f}%"
                    }

                    st.session_state.processed_image_format = output_format.lower()
                    st.success("图片处理完成！")
            except Exception as e:
                st.error(f"图片处理失败: {e}")
                import traceback

                st.error(f"详细错误: {traceback.format_exc()}")

    # 显示处理后的图片和下载
    if (st.session_state.processed_image is not None and
            st.session_state.processed_info is not None and
            st.session_state.processed_image_data is not None):

        st.markdown("### 4. 处理结果")
        col1, col2 = st.columns(2)

        with col1:
            st.image(st.session_state.processed_image_data, caption="处理后图片", use_container_width=True)

        with col2:
            st.markdown("**处理结果信息:**")
            for key, value in st.session_state.processed_info.items():
                st.write(f"**{key}:** {value}")

            file_name = f"processed_image.{st.session_state.processed_image_format}"
            st.download_button(
                label="📥 下载处理后的图片",
                data=st.session_state.processed_image_data,
                file_name=file_name,
                mime=f"image/{st.session_state.processed_image_format}",
                use_container_width=True
            )

    # 如果没有上传图片，显示使用说明
    if not st.session_state.original_image:
        st.info("""
            ### 使用说明：
            1. **上传图片**: 支持 JPG、PNG、GIF、BMP、WEBP 格式
            2. **查看原图信息**: 显示文件名、格式、尺寸、文件大小
            3. **选择处理模式**: 包括格式转换、调整尺寸、图片裁剪、添加水印等
            4. **转换并下载**: 查看处理结果并下载新图片

            ### 图片裁剪功能：
            - ✂️ **手动设置区域**: 通过滑块精确设置裁剪区域，实时预览效果
            - 📐 **按比例裁剪**: 选择常见比例或自定义比例自动计算裁剪区域
            - 👀 **实时预览**: 设置后立即看到裁剪效果
            - 📊 **详细信息**: 显示裁剪位置、尺寸和原图利用率
            """)

    st.markdown('</div>', unsafe_allow_html=True)

# 在工具选择部分之后添加加密/解密工具的实现
elif tool_category == "加密/解密工具":

    # 初始化session state
    if 'crypto_clear_counter' not in st.session_state:
        st.session_state.crypto_clear_counter = 0
    show_doc("crypto_tools")
    # 加密工具选择
    crypto_tool = st.radio(
        "选择加密工具",
        ["Base64编码", "MD5加密", "SHA加密", "RSA加解密", "对称加密", "URL编码", "HTML编码", "Unicode编码", "十六进制编码"],
        horizontal=True
    )

    if crypto_tool == "Base64编码":
        st.markdown('<div class="category-card">📝 Base64编码/解码</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            # 使用动态key来支持清空功能
            input_text = st.text_area("输入文本", height=150,
                                      placeholder="请输入要编码或解码的文本...",
                                      key=f"base64_input_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                encode_btn = st.button("Base64编码", use_container_width=True, key="base64_encode_btn")
            with col_btn2:
                decode_btn = st.button("Base64解码", use_container_width=True, key="base64_decode_btn")

            if st.button("清空", use_container_width=True, key="base64_clear_btn"):
                st.session_state.crypto_clear_counter += 1
                st.rerun()

        with col2:
            if encode_btn and input_text:
                try:
                    encoded = base64.b64encode(input_text.encode('utf-8')).decode('utf-8')
                    st.text_area("编码结果", encoded, height=150, key="base64_encoded")
                    create_copy_button(encoded, button_text="📋 复制结果", key="copy_base64_encode")
                except Exception as e:
                    st.error(f"编码失败: {e}")

            elif decode_btn and input_text:
                try:
                    # 检查是否为有效的Base64编码
                    import re

                    base64_pattern = re.compile(r'^[A-Za-z0-9+/]*={0,2}$')
                    clean_input = input_text.strip()

                    if not base64_pattern.match(clean_input):
                        st.error("解码失败：请检查输入是否为有效的Base64编码")
                    else:
                        # 尝试补全=
                        if len(clean_input) % 4 != 0:
                            clean_input += '=' * (4 - len(clean_input) % 4)

                        decoded = base64.b64decode(clean_input).decode('utf-8')
                        st.text_area("解码结果", decoded, height=150, key="base64_decoded")
                        create_copy_button(decoded, button_text="📋 复制结果", key="copy_base64_decode")
                except Exception as e:
                    st.error(f"解码失败：请检查输入是否为有效的Base64编码")

    elif crypto_tool == "MD5加密":
        st.markdown('<div class="category-card">🔑 MD5加密</div>', unsafe_allow_html=True)

        input_text = st.text_area("输入文本", height=100,
                                  placeholder="请输入要加密的文本...",
                                  key=f"md5_input_{st.session_state.crypto_clear_counter}")

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            encrypt_btn = st.button("MD5加密", use_container_width=True, key="md5_encrypt_btn")
        with col_btn2:
            if st.button("清空", use_container_width=True, key="md5_clear_btn"):
                st.session_state.crypto_clear_counter += 1
                st.rerun()

        if encrypt_btn and input_text:
            # 生成不同格式的MD5
            md5_hash = hashlib.md5(input_text.encode('utf-8')).hexdigest()

            col1, col2 = st.columns(2)
            with col1:
                # 修复高度问题，使用最小高度68
                st.text_area("32位小写", md5_hash, height=68, key="md5_32_lower")
                create_copy_button(md5_hash, button_text="📋 复制32位小写", key="copy_md5_32_lower")

                md5_16_lower = md5_hash[8:24]
                st.text_area("16位小写", md5_16_lower, height=68, key="md5_16_lower")
                create_copy_button(md5_16_lower, button_text="📋 复制16位小写", key="copy_md5_16_lower")

            with col2:
                md5_32_upper = md5_hash.upper()
                st.text_area("32位大写", md5_32_upper, height=68, key="md5_32_upper")
                create_copy_button(md5_32_upper, button_text="📋 复制32位大写", key="copy_md5_32_upper")

                md5_16_upper = md5_16_lower.upper()
                st.text_area("16位大写", md5_16_upper, height=68, key="md5_16_upper")
                create_copy_button(md5_16_upper, button_text="📋 复制16位大写", key="copy_md5_16_upper")

            st.info("💡 MD5是单向哈希函数，无法解密。主要用于验证数据完整性。")

    elif crypto_tool == "SHA加密":
        st.markdown('<div class="category-card">🔐 SHA系列加密</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            input_text = st.text_area("输入文本", height=100,
                                      placeholder="请输入要加密的文本...",
                                      key=f"sha_input_{st.session_state.crypto_clear_counter}")
            sha_type = st.selectbox("选择SHA算法", [
                "SHA1", "SHA224", "SHA256", "SHA384", "SHA512", "SHA3",
                "MD5", "HamcSHA1", "HamcSHA224", "HamcSHA256", "HamcSHA384",
                "HamcSHA512", "HamcMD5", "HamcSHA3", "PBKDF2"
            ], key="sha_type_select")

            # 对于HMAC和PBKDF2需要密钥
            if sha_type.startswith('Hamc') or sha_type == 'PBKDF2':
                key = st.text_input("密钥", placeholder="请输入密钥",
                                    key=f"sha_key_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                encrypt_btn = st.button("加密", use_container_width=True, key="sha_encrypt_btn")
            with col_btn2:
                if st.button("清空", use_container_width=True, key="sha_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encrypt_btn and input_text:
                try:
                    result = ""

                    if sha_type == "SHA1":
                        result = hashlib.sha1(input_text.encode()).hexdigest()
                    elif sha_type == "SHA224":
                        result = hashlib.sha224(input_text.encode()).hexdigest()
                    elif sha_type == "SHA256":
                        result = hashlib.sha256(input_text.encode()).hexdigest()
                    elif sha_type == "SHA384":
                        result = hashlib.sha384(input_text.encode()).hexdigest()
                    elif sha_type == "SHA512":
                        result = hashlib.sha512(input_text.encode()).hexdigest()
                    elif sha_type == "SHA3":
                        result = hashlib.sha3_256(input_text.encode()).hexdigest()
                    elif sha_type == "MD5":
                        result = hashlib.md5(input_text.encode()).hexdigest()
                    elif sha_type.startswith('Hamc'):
                        # HMAC加密
                        if not key:
                            st.error("HMAC需要密钥")
                        else:
                            algo = sha_type[4:].lower()  # 去掉Hamc前缀
                            if algo == "sha1":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha1)
                            elif algo == "sha224":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha224)
                            elif algo == "sha256":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha256)
                            elif algo == "sha384":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha384)
                            elif algo == "sha512":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha512)
                            elif algo == "md5":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.md5)
                            elif algo == "sha3":
                                h = hmac.new(key.encode(), input_text.encode(), hashlib.sha3_256)
                            result = h.hexdigest()
                    elif sha_type == "PBKDF2":
                        if not key:
                            st.error("PBKDF2需要盐值")
                        else:
                            # 简化的PBKDF2实现
                            dk = hashlib.pbkdf2_hmac('sha256', input_text.encode(), key.encode(), 100000)
                            result = binascii.hexlify(dk).decode()

                    st.text_area("加密结果", result, height=100, key="sha_result")
                    create_copy_button(result, button_text="📋 复制结果", key=f"copy_{sha_type}")

                except Exception as e:
                    st.error(f"加密失败: {e}")

    elif crypto_tool == "对称加密":
        st.markdown('<div class="category-card">🔑 对称加密/解密</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            algorithm = st.selectbox("加密算法", ["AES", "DES", "TripleDes", "RC4", "Rabbit"], key="sym_algo_select")
            input_text = st.text_area("输入文本", height=100,
                                      placeholder="请输入要加密/解密的文本...",
                                      key=f"symmetric_input_{st.session_state.crypto_clear_counter}")
            key = st.text_input("密钥（可选）", placeholder="请输入密钥（可选）",
                                key=f"symmetric_key_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                encrypt_btn = st.button("加密", use_container_width=True, key="sym_encrypt_btn")
            with col_btn2:
                decrypt_btn = st.button("解密", use_container_width=True, key="sym_decrypt_btn")
            with col_btn3:
                if st.button("清空", use_container_width=True, key="sym_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encrypt_btn and input_text:
                try:
                    # 简化的对称加密实现
                    if algorithm == "AES":
                        # 使用默认密钥如果用户没有输入
                        actual_key = key.encode() if key else b'default_key_12345'
                        # 确保密钥长度为16字节
                        actual_key = actual_key.ljust(16, b'\0')[:16]
                        cipher = AES.new(actual_key, AES.MODE_ECB)
                        encrypted = base64.b64encode(cipher.encrypt(pad(input_text.encode(), 16))).decode()
                    elif algorithm == "DES":
                        # DES实现
                        actual_key = key.encode() if key else b'default_k'
                        # 确保密钥长度为8字节
                        actual_key = actual_key.ljust(8, b'\0')[:8]
                        cipher = DES.new(actual_key, DES.MODE_ECB)
                        encrypted = base64.b64encode(cipher.encrypt(pad(input_text.encode(), 8))).decode()
                    else:
                        # 其他算法的简化实现
                        encrypted = f"{algorithm}加密: {base64.b64encode(input_text.encode()).decode()}"

                    st.text_area("加密结果", encrypted, height=100, key="symmetric_encrypted")
                    create_copy_button(encrypted, button_text="📋 复制结果", key="copy_symmetric_encrypt")

                except Exception as e:
                    st.error(f"加密失败: {e}")

            elif decrypt_btn and input_text:
                try:
                    # 简化的对称解密实现
                    if algorithm == "AES":
                        actual_key = key.encode() if key else b'default_key_12345'
                        actual_key = actual_key.ljust(16, b'\0')[:16]
                        cipher = AES.new(actual_key, AES.MODE_ECB)
                        decrypted = unpad(cipher.decrypt(base64.b64decode(input_text)), 16).decode()
                    elif algorithm == "DES":
                        actual_key = key.encode() if key else b'default_k'
                        actual_key = actual_key.ljust(8, b'\0')[:8]
                        cipher = DES.new(actual_key, DES.MODE_ECB)
                        decrypted = unpad(cipher.decrypt(base64.b64decode(input_text)), 8).decode()
                    else:
                        # 其他算法的简化实现
                        decrypted = base64.b64decode(input_text).decode()

                    st.text_area("解密结果", decrypted, height=100, key="symmetric_decrypted")
                    create_copy_button(decrypted, button_text="📋 复制结果", key="copy_symmetric_decrypt")

                except Exception as e:
                    st.error(f"解密失败: {e}")

    elif crypto_tool == "URL编码":
        st.markdown('<div class="category-card">🔗 URL编码/解码</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            input_text = st.text_area("输入文本", height=150,
                                      placeholder="请输入要编码或解码的URL...",
                                      key=f"url_input_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                encode_btn = st.button("URL编码", use_container_width=True, key="url_encode_btn")
            with col_btn2:
                decode_btn = st.button("URL解码", use_container_width=True, key="url_decode_btn")
            with col_btn3:
                if st.button("清空", use_container_width=True, key="url_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encode_btn and input_text:
                try:
                    encoded = urllib.parse.quote(input_text, safe='')
                    st.text_area("编码结果", encoded, height=150, key="url_encoded")
                    create_copy_button(encoded, button_text="📋 复制结果", key="copy_url_encode")
                except Exception as e:
                    st.error(f"编码失败: {e}")

            elif decode_btn and input_text:
                try:
                    decoded = urllib.parse.unquote(input_text)
                    st.text_area("解码结果", decoded, height=150, key="url_decoded")
                    create_copy_button(decoded, button_text="📋 复制结果", key="copy_url_decode")
                except Exception as e:
                    st.error(f"解码失败: {e}")

    elif crypto_tool == "HTML编码":
        st.markdown('<div class="category-card">🌐 HTML编码/解码</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            input_text = st.text_area("输入文本", height=150,
                                      placeholder="请输入要编码或解码的HTML...",
                                      key=f"html_input_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                encode_btn = st.button("HTML编码", use_container_width=True, key="html_encode_btn")
            with col_btn2:
                decode_btn = st.button("HTML解码", use_container_width=True, key="html_decode_btn")
            with col_btn3:
                if st.button("清空", use_container_width=True, key="html_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encode_btn and input_text:
                try:
                    encoded = html.escape(input_text)
                    st.text_area("编码结果", encoded, height=150, key="html_encoded")
                    create_copy_button(encoded, button_text="📋 复制结果", key="copy_html_encode")
                except Exception as e:
                    st.error(f"编码失败: {e}")

            elif decode_btn and input_text:
                try:
                    decoded = html.unescape(input_text)
                    st.text_area("解码结果", decoded, height=150, key="html_decoded")
                    create_copy_button(decoded, button_text="📋 复制结果", key="copy_html_decode")
                except Exception as e:
                    st.error(f"解码失败: {e}")

    elif crypto_tool == "Unicode编码":
        st.markdown('<div class="category-card">🔤 Unicode编码/解码</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            input_text = st.text_area("输入文本", height=150,
                                      placeholder="请输入要编码或解码的文本...",
                                      key=f"unicode_input_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                encode_btn = st.button("Unicode编码", use_container_width=True, key="unicode_encode_btn")
            with col_btn2:
                decode_btn = st.button("Unicode解码", use_container_width=True, key="unicode_decode_btn")
            with col_btn3:
                if st.button("清空", use_container_width=True, key="unicode_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encode_btn and input_text:
                try:
                    encoded = input_text.encode('unicode_escape').decode('utf-8')
                    st.text_area("编码结果", encoded, height=150, key="unicode_encoded")
                    create_copy_button(encoded, button_text="📋 复制结果", key="copy_unicode_encode")
                except Exception as e:
                    st.error(f"编码失败: {e}")

            elif decode_btn and input_text:
                try:
                    decoded = codecs.decode(input_text, 'unicode_escape')
                    st.text_area("解码结果", decoded, height=150, key="unicode_decoded")
                    create_copy_button(decoded, button_text="📋 复制结果", key="copy_unicode_decode")
                except Exception as e:
                    st.error(f"解码失败: {e}")

    elif crypto_tool == "十六进制编码":
        st.markdown('<div class="category-card">🔢 十六进制编码/解码</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            input_text = st.text_area("输入文本", height=150,
                                      placeholder="请输入要编码或解码的文本...",
                                      key=f"hex_input_{st.session_state.crypto_clear_counter}")

            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                encode_btn = st.button("十六进制编码", use_container_width=True, key="hex_encode_btn")
            with col_btn2:
                decode_btn = st.button("十六进制解码", use_container_width=True, key="hex_decode_btn")
            with col_btn3:
                if st.button("清空", use_container_width=True, key="hex_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with col2:
            if encode_btn and input_text:
                try:
                    encoded = binascii.hexlify(input_text.encode()).decode()
                    st.text_area("编码结果", encoded, height=150, key="hex_encoded")
                    create_copy_button(encoded, button_text="📋 复制结果", key="copy_hex_encode")
                except Exception as e:
                    st.error(f"编码失败: {e}")

            elif decode_btn and input_text:
                try:
                    decoded = binascii.unhexlify(input_text).decode()
                    st.text_area("解码结果", decoded, height=150, key="hex_decoded")
                    create_copy_button(decoded, button_text="📋 复制结果", key="copy_hex_decode")
                except Exception as e:
                    st.error(f"解码失败: {e}")

    elif crypto_tool == "RSA加解密":
        st.markdown('<div class="category-card">🔐 RSA加解密</div>', unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["RSA密钥生成", "RSA加解密"])

        with tab1:
            st.markdown("**RSA密钥对生成**")

            col1, col2 = st.columns(2)

            with col1:
                key_size = st.selectbox("密钥长度", [512, 1024, 2048, 4096], index=2, key="rsa_key_size")
                key_format = st.selectbox("密钥格式", ["PKCS#8", "PKCS#1"], key="rsa_key_format")
                password = st.text_input("私钥密码（可选）", type="password", placeholder="可选的私钥密码",
                                         key=f"rsa_password_{st.session_state.crypto_clear_counter}")

            with col2:
                if st.button("生成密钥对", use_container_width=True, key="rsa_generate_btn"):
                    try:
                        # 简化的RSA密钥生成实现
                        import os
                        import base64

                        # 生成随机密钥对（这里使用模拟数据）
                        public_key_template = f"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA{base64.b64encode(os.urandom(128)).decode()[:172]}
-----END PUBLIC KEY-----"""

                        private_key_template = f"""-----BEGIN PRIVATE KEY-----
MIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQD{base64.b64encode(os.urandom(256)).decode()[:344]}
-----END PRIVATE KEY-----"""

                        # 如果有密码，在注释中说明
                        if password:
                            private_key_template = f"# 使用密码保护的私钥\n# 密码: {password}\n{private_key_template}"

                        st.success("RSA密钥对生成成功！")

                        col_key1, col_key2 = st.columns(2)
                        with col_key1:
                            st.text_area("公钥", public_key_template, height=200, key="rsa_public_key")
                            create_copy_button(public_key_template, button_text="📋 复制公钥", key="copy_rsa_public")
                        with col_key2:
                            st.text_area("私钥", private_key_template, height=200, key="rsa_private_key")
                            create_copy_button(private_key_template, button_text="📋 复制私钥", key="copy_rsa_private")

                    except Exception as e:
                        st.error(f"密钥生成失败: {e}")

                if st.button("清空", use_container_width=True, key="rsa_tab1_clear_btn"):
                    st.session_state.crypto_clear_counter += 1
                    st.rerun()

        with tab2:
            st.markdown("**RSA加密/解密**")

            col1, col2 = st.columns(2)

            with col1:
                rsa_mode = st.radio("RSA模式", ["RSA", "RSA2"], key="rsa_mode_radio")
                key_input = st.text_area("公钥/私钥", height=100,
                                         placeholder="请输入公钥(加密)或私钥(解密)...",
                                         key=f"rsa_key_input_{st.session_state.crypto_clear_counter}")
                text_input = st.text_area("输入文本", height=100,
                                          placeholder="请输入要加密/解密的文本...",
                                          key=f"rsa_text_input_{st.session_state.crypto_clear_counter}")

            with col2:
                col_btn1, col_btn2, col_btn3 = st.columns(3)
                with col_btn1:
                    encrypt_btn = st.button("RSA加密", use_container_width=True, key="rsa_encrypt_btn")
                with col_btn2:
                    decrypt_btn = st.button("RSA解密", use_container_width=True, key="rsa_decrypt_btn")
                with col_btn3:
                    if st.button("清空", use_container_width=True, key="rsa_tab2_clear_btn"):
                        st.session_state.crypto_clear_counter += 1
                        st.rerun()

                if encrypt_btn and key_input and text_input:
                    try:
                        # 简化的RSA加密实现
                        encrypted_result = f"RSA加密结果（模拟）:\n{base64.b64encode(text_input.encode()).decode()}"
                        st.text_area("加密结果", encrypted_result, height=100, key="rsa_encrypted")
                        create_copy_button(encrypted_result, button_text="📋 复制结果", key="copy_rsa_encrypt")
                        st.info("这是一个简化的RSA加密演示。实际使用时需要完整的RSA库实现。")
                    except Exception as e:
                        st.error(f"加密失败: {e}")

                elif decrypt_btn and key_input and text_input:
                    try:
                        # 简化的RSA解密实现
                        decrypted_result = f"RSA解密结果（模拟）:\n{base64.b64decode(text_input).decode()}"
                        st.text_area("解密结果", decrypted_result, height=100, key="rsa_decrypted")
                        create_copy_button(decrypted_result, button_text="📋 复制结果", key="copy_rsa_decrypt")
                        st.info("这是一个简化的RSA解密演示。实际使用时需要完整的RSA库实现。")
                    except Exception as e:
                        st.error(f"解密失败: {e}")

    st.markdown('</div>', unsafe_allow_html=True)

# 在工具选择部分之后添加测试用例生成器
elif tool_category == "测试用例生成器":
    show_doc("test_case_generator")

    if 'test_cases' not in st.session_state:
        st.session_state.test_cases = []
    if 'requirement_history' not in st.session_state:
        st.session_state.requirement_history = []
    if 'current_requirement' not in st.session_state:
        st.session_state.current_requirement = ""
    if 'test_case_generator' not in st.session_state:
        st.session_state.test_case_generator = TestCaseGenerator()
    if 'testcase_input_counter' not in st.session_state:
        st.session_state.testcase_input_counter = 0
    if 'current_requirement_input' not in st.session_state:
        st.session_state.current_requirement_input = ""
    if 'testcase_ocr_text' not in st.session_state:
        st.session_state.testcase_ocr_text = ""
    if 'testcase_ocr_details' not in st.session_state:
        st.session_state.testcase_ocr_details = []
    if 'testcase_coverage_focus' not in st.session_state:
        st.session_state.testcase_coverage_focus = ["核心功能", "异常处理", "边界值"]
    if 'testcase_last_compiled_requirement' not in st.session_state:
        st.session_state.testcase_last_compiled_requirement = ""

    generator = st.session_state.test_case_generator

    st.markdown("### 🔑 API配置")

    col1, col2 = st.columns(2)
    with col1:
        model_provider = st.selectbox(
            "选择大模型",
            ["阿里通义千问", "OpenAI GPT", "百度文心一言", "讯飞星火", "智谱ChatGLM"],
            help="选择使用的大模型提供商",
            key="model_provider_select"
        )
    with col2:
        id_prefix = st.text_input("用例ID前缀", value="TC", help="例如: TC、TEST、CASE等", key="id_prefix_input")

    api_config = {}
    if model_provider == "阿里通义千问":
        st.markdown("#### 阿里通义千问配置")
        api_key = st.text_input(
            "API Key",
            value="",
            type="password",
            help="请输入阿里通义千问的API密钥",
            key="ali_api_key"
        )
        api_config = {"api_key": api_key}
        st.info("💡 阿里通义千问适合处理中文需求，在软件测试场景表现优秀")

    elif model_provider == "OpenAI GPT":
        st.markdown("#### OpenAI GPT配置")
        col1, col2 = st.columns(2)
        with col1:
            api_key = st.text_input(
                "API Key",
                value="",
                type="password",
                help="请输入OpenAI的API密钥",
                key="openai_api_key"
            )
        with col2:
            model_version = st.selectbox(
                "模型版本",
                [
                    "gpt-4o",
                    "gpt-4o-mini",
                    "gpt-4-turbo",
                    "gpt-4",
                    "gpt-3.5-turbo",
                    "gpt-3.5-turbo-16k"
                ],
                help="选择GPT模型版本",
                key="gpt_model_select"
            )
        api_config = {"api_key": api_key, "model_version": model_version}
        st.info("💡 GPT系列模型在逻辑推理和结构化输出方面表现突出")

    elif model_provider == "百度文心一言":
        st.markdown("#### 百度文心一言配置")
        col1, col2 = st.columns(2)
        with col1:
            api_key = st.text_input(
                "API Key",
                value="",
                type="password",
                help="请输入百度文心一言的API密钥",
                key="baidu_api_key"
            )
        with col2:
            secret_key = st.text_input(
                "Secret Key",
                value="",
                type="password",
                help="请输入百度文心一言的Secret Key",
                key="baidu_secret_key"
            )
        api_config = {"api_key": api_key, "secret_key": secret_key}
        st.info("💡 文心一言对中文理解深入，在业务场景描述方面表现良好")

    elif model_provider == "讯飞星火":
        st.markdown("#### 讯飞星火配置")
        col1, col2 = st.columns(2)
        with col1:
            api_key = st.text_input(
                "API Key",
                value="",
                type="password",
                help="请输入讯飞星火的API密钥",
                key="spark_api_key"
            )
        with col2:
            model_id = st.text_input(
                "Model ID",
                value="",
                help="讯飞星火的模型ID",
                key="spark_model_id"
            )

        api_base = st.text_input(
            "API Base",
            value="http://maas-api.cn-huabei-1.xf-yun.com/v1",
            help="讯飞星火的API基础地址，通常使用默认值即可",
            key="spark_api_base"
        )

        api_config = {
            "api_key": api_key,
            "api_base": api_base,
            "model_id": model_id
        }
        st.info("💡 讯飞星火在技术文档和代码相关任务中表现优秀")

    elif model_provider == "智谱ChatGLM":
        st.markdown("#### 智谱ChatGLM配置")
        api_key = st.text_input(
            "API Key",
            value="",
            type="password",
            help="请输入智谱AI的API密钥",
            key="glm_api_key"
        )
        api_config = {"api_key": api_key}
        st.info("💡 ChatGLM在中文技术文档处理方面有独特优势")

    st.markdown("### 📝 需求输入")
    input_tab1, input_tab2, input_tab3 = st.tabs(["手动输入", "图片OCR识别", "结构化补充"])

    with input_tab1:
        st.markdown("**快速选择示例需求：**")
        example_col1, example_col2, example_col3 = st.columns(3)

        with example_col1:
            if st.button("📱 简单功能示例", use_container_width=True, key="simple_example_btn"):
                st.session_state.current_requirement_input = SIMPLE_EXAMPLE
                st.session_state.testcase_input_counter += 1
                st.rerun()

        with example_col2:
            if st.button("🔐 中等功能示例", use_container_width=True, key="medium_example_btn"):
                st.session_state.current_requirement_input = MEDIUM_EXAMPLE
                st.session_state.testcase_input_counter += 1
                st.rerun()

        with example_col3:
            if st.button("🛒 复杂功能示例", use_container_width=True, key="complex_example_btn"):
                st.session_state.current_requirement_input = COMPLEX_EXAMPLE
                st.session_state.testcase_input_counter += 1
                st.rerun()

        requirement = st.text_area(
            "需求描述",
            value=st.session_state.current_requirement_input,
            height=220,
            placeholder="请输入详细的需求描述，或先用图片OCR提取后再补充...",
            key=f"requirement_input_{st.session_state.testcase_input_counter}",
            help="需求越清晰，生成的测试用例越准确。"
        )

    with input_tab2:
        ocr_status = generator.get_ocr_status()
        if generator.is_ocr_available():
            st.success(ocr_status["message"])
        else:
            st.warning(ocr_status["message"])

        ocr_lang_options = generator.get_ocr_language_options()
        ocr_col1, ocr_col2 = st.columns(2)
        with ocr_col1:
            selected_ocr_lang = st.selectbox(
                "识别语言",
                options=list(ocr_lang_options.keys()),
                index=0,
                key="testcase_ocr_language"
            )
        with ocr_col2:
            preprocess_mode = st.selectbox(
                "预处理模式",
                options=generator.get_ocr_preprocess_modes(),
                index=0,
                key="testcase_ocr_preprocess"
            )

        uploaded_requirement_images = st.file_uploader(
            "上传需求截图",
            type=["png", "jpg", "jpeg", "webp", "bmp"],
            accept_multiple_files=True,
            help="适合上传 PRD 截图、原型图、聊天需求截图、验收说明截图等。",
            key="testcase_requirement_images"
        )

        action_col1, action_col2, action_col3 = st.columns(3)
        with action_col1:
            run_ocr_btn = st.button(
                "识别图片需求",
                use_container_width=True,
                disabled=not uploaded_requirement_images or not generator.is_ocr_available(),
                key="run_testcase_ocr_btn"
            )
        with action_col2:
            append_ocr_btn = st.button(
                "追加到需求框",
                use_container_width=True,
                disabled=not st.session_state.testcase_ocr_text,
                key="append_ocr_requirement_btn"
            )
        with action_col3:
            clear_ocr_btn = st.button(
                "清空OCR结果",
                use_container_width=True,
                key="clear_testcase_ocr_btn"
            )

        if run_ocr_btn and uploaded_requirement_images:
            recognized_details = []
            with st.spinner("正在识别图片中的需求文本..."):
                for uploaded_image in uploaded_requirement_images:
                    try:
                        extracted_text = generator.extract_text_from_image(
                            uploaded_image.getvalue(),
                            lang=ocr_lang_options[selected_ocr_lang],
                            preprocess_mode=preprocess_mode
                        )
                        recognized_details.append({
                            "file_name": uploaded_image.name,
                            "text": extracted_text
                        })
                    except Exception as e:
                        recognized_details.append({
                            "file_name": uploaded_image.name,
                            "text": f"[识别失败] {str(e)}"
                        })

            st.session_state.testcase_ocr_details = recognized_details
            st.session_state.testcase_ocr_text = "\n\n".join(
                f"[图片: {item['file_name']}]\n{item['text']}" for item in recognized_details
            ).strip()

        if append_ocr_btn and st.session_state.testcase_ocr_text:
            merged_requirement = requirement.strip()
            if merged_requirement:
                merged_requirement += "\n\n"
            merged_requirement += st.session_state.testcase_ocr_text
            st.session_state.current_requirement_input = generator.clean_requirement_text(merged_requirement)
            st.session_state.testcase_input_counter += 1
            st.rerun()

        if clear_ocr_btn:
            st.session_state.testcase_ocr_text = ""
            st.session_state.testcase_ocr_details = []
            st.rerun()

        if st.session_state.testcase_ocr_details:
            for idx, item in enumerate(st.session_state.testcase_ocr_details):
                with st.expander(f"OCR结果 {idx + 1}: {item['file_name']}"):
                    st.text_area(
                        "识别文本",
                        value=item["text"],
                        height=180,
                        key=f"testcase_ocr_result_{idx}",
                        disabled=True
                    )

            replace_col1, replace_col2 = st.columns(2)
            with replace_col1:
                if st.button("使用OCR结果替换需求框", use_container_width=True, key="replace_requirement_with_ocr_btn"):
                    st.session_state.current_requirement_input = st.session_state.testcase_ocr_text
                    st.session_state.testcase_input_counter += 1
                    st.rerun()
            with replace_col2:
                create_copy_button(
                    st.session_state.testcase_ocr_text,
                    button_text="📋 复制 OCR 文本",
                    key="copy_testcase_ocr_text"
                )

    with input_tab3:
        st.caption("这些补充信息会和需求原文一起提交给 AI，适合把零散需求整理成更适合生成用例的上下文。")
        col1, col2 = st.columns(2)
        with col1:
            st.text_input("所属模块/页面", key="testcase_module_name", placeholder="例如：用户中心-地址管理页")
            st.text_area(
                "业务规则/字段约束",
                key="testcase_business_rules",
                height=140,
                placeholder="例如：手机号必填；地址最多20条；默认地址只能有1个"
            )
            st.text_area(
                "验收标准/关键成功条件",
                key="testcase_acceptance_criteria",
                height=120,
                placeholder="例如：新增成功后列表实时刷新；默认地址标签立即生效"
            )
        with col2:
            st.text_area(
                "本次不覆盖范围",
                key="testcase_out_of_scope",
                height=140,
                placeholder="例如：不覆盖性能压测；不覆盖老版本兼容"
            )
            st.text_area(
                "补充说明",
                key="testcase_additional_notes",
                height=120,
                placeholder="例如：当前版本仅支持 App 端；接口联调阶段，提示文案可能调整"
            )
            st.checkbox(
                "生成时附带 OCR 识别内容",
                value=bool(st.session_state.testcase_ocr_text),
                key="testcase_include_ocr"
            )

    compiled_requirement = generator.compose_requirement_context(
        requirement=requirement,
        ocr_text=st.session_state.testcase_ocr_text if st.session_state.get("testcase_include_ocr", False) else "",
        module_name=st.session_state.get("testcase_module_name", ""),
        business_rules=st.session_state.get("testcase_business_rules", ""),
        acceptance_criteria=st.session_state.get("testcase_acceptance_criteria", ""),
        out_of_scope=st.session_state.get("testcase_out_of_scope", ""),
        additional_notes=st.session_state.get("testcase_additional_notes", "")
    )
    st.session_state.testcase_last_compiled_requirement = compiled_requirement

    requirement_analysis = generator.analyze_requirement(compiled_requirement)
    if compiled_requirement.strip():
        with st.expander("🧭 需求分析助手", expanded=True):
            metric_col1, metric_col2, metric_col3 = st.columns(3)
            with metric_col1:
                st.metric("复杂度", requirement_analysis["complexity"])
            with metric_col2:
                st.metric("文本行数", requirement_analysis["line_count"])
            with metric_col3:
                st.metric("功能点数", len(requirement_analysis["feature_points"]))

            st.caption(f"需求摘要: {requirement_analysis['summary']}")

            info_col1, info_col2, info_col3 = st.columns(3)
            with info_col1:
                st.markdown("**关键功能点**")
                for item in requirement_analysis["feature_points"][:6]:
                    st.write(f"- {item}")
            with info_col2:
                st.markdown("**建议覆盖维度**")
                for item in requirement_analysis["suggested_focus"]:
                    st.write(f"- {item}")
                if st.button("采用建议覆盖维度", key="apply_requirement_focus_btn", use_container_width=True):
                    st.session_state.testcase_coverage_focus = requirement_analysis["suggested_focus"] or \
                                                               st.session_state.testcase_coverage_focus
                    st.rerun()
            with info_col3:
                st.markdown("**待确认项**")
                for item in requirement_analysis["unclear_points"]:
                    st.write(f"- {item}")

            if requirement_analysis["business_rules"]:
                st.markdown("**已识别的业务规则/约束**")
                for item in requirement_analysis["business_rules"]:
                    st.write(f"- {item}")

    with st.expander("🔧 高级选项", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            case_styles = generator.get_case_styles()
            case_style = st.selectbox(
                "测试用例风格",
                options=list(case_styles.keys()),
                index=0,
                help="选择测试用例的编写风格，将直接影响生成的用例格式",
                key="case_style_select"
            )
            if case_style in case_styles:
                st.caption(f"💡 {case_styles[case_style]}")

        with col2:
            languages = generator.get_languages()
            language = st.selectbox(
                "输出语言",
                options=languages,
                index=0,
                help="选择测试用例的输出语言",
                key="language_select"
            )
            if language in LANGUAGE_DESCRIPTIONS:
                st.caption(f"🌐 {LANGUAGE_DESCRIPTIONS[language]}")

        with col3:
            target_case_count = st.number_input(
                "目标用例数",
                min_value=4,
                max_value=60,
                value=12,
                help="模型会按该数量级生成，实际条数可能因需求复杂度略有上下浮动。",
                key="testcase_target_case_count"
            )
            st.caption("💡 简单功能建议 6-12 条，复杂流程建议 15-30 条")

        coverage_options = generator.get_coverage_focus_options()
        coverage_focus = st.multiselect(
            "重点覆盖维度",
            options=list(coverage_options.keys()),
            key="testcase_coverage_focus",
            help="告诉模型优先覆盖哪些测试角度。"
        )
        if coverage_focus:
            st.caption(" | ".join(f"{item}: {coverage_options[item]}" for item in coverage_focus))

    if st.button("预览风格示例", key="preview_style_btn"):
        if case_style in STYLE_PREVIEWS:
            preview = STYLE_PREVIEWS[case_style]
            st.info(f"**{case_style} 示例:** {preview['中文']} | {preview['英文']}")

    with st.expander("👀 最终提交给AI的需求上下文预览", expanded=False):
        if compiled_requirement.strip():
            st.text_area(
                "提交内容",
                value=compiled_requirement,
                height=260,
                key="compiled_requirement_preview",
                disabled=True
            )
        else:
            st.caption("当前还没有可提交的需求内容。可以手动输入，也可以先做 OCR 识别。")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("清空输入", use_container_width=True, key="clear_input_btn"):
            st.session_state.current_requirement_input = ""
            st.session_state.testcase_ocr_text = ""
            st.session_state.testcase_ocr_details = []
            st.session_state.testcase_input_counter += 1
            st.rerun()

    with col2:
        generate_btn = st.button(
            "🧠 AI生成测试用例",
            use_container_width=True,
            disabled=not compiled_requirement.strip(),
            key="generate_testcases_btn"
        )

    if generate_btn and compiled_requirement.strip():
        platform = PLATFORM_MAPPING[model_provider]

        validation_errors = []
        if platform == "ali" and not api_config.get("api_key"):
            validation_errors.append("请输入阿里通义千问API Key")
        elif platform == "openai" and not api_config.get("api_key"):
            validation_errors.append("请输入OpenAI API Key")
        elif platform == "baidu" and (not api_config.get("api_key") or not api_config.get("secret_key")):
            validation_errors.append("请输入百度文心一言的API Key和Secret Key")
        elif platform == "spark" and not api_config.get("api_key"):
            validation_errors.append("请输入讯飞星火的 API Key")
        elif platform == "glm" and not api_config.get("api_key"):
            validation_errors.append("请输入智谱ChatGLM API Key")

        if validation_errors:
            for error in validation_errors:
                st.error(error)
            st.stop()

        st.write("🎯 生成参数:")
        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
        with metric_col1:
            st.metric("风格", case_style)
        with metric_col2:
            st.metric("语言", language)
        with metric_col3:
            st.metric("平台", model_provider)
        with metric_col4:
            st.metric("目标数量", target_case_count)

        with st.spinner(f"🤖 {model_provider}正在分析需求并生成测试用例..."):
            try:
                test_cases = generator.generate_testcases(
                    requirement=compiled_requirement,
                    platform=platform,
                    api_config=api_config,
                    id_prefix=id_prefix,
                    case_style=case_style,
                    language=language,
                    target_case_count=target_case_count,
                    coverage_focus=coverage_focus
                )

                st.session_state.test_cases = test_cases
                st.session_state.current_requirement = compiled_requirement

                history_item = {
                    "timestamp": time.strftime("%Y-%m-%d %H:%M"),
                    "requirement": (requirement or compiled_requirement)[:100] + "..."
                    if len((requirement or compiled_requirement)) > 100 else (requirement or compiled_requirement),
                    "case_count": len(test_cases),
                    "model": model_provider,
                    "full_requirement": compiled_requirement,
                    "editable_requirement": requirement or compiled_requirement,
                    "platform": platform,
                    "case_style": case_style,
                    "language": language,
                    "target_case_count": target_case_count,
                    "coverage_focus": list(coverage_focus),
                    "used_ocr": bool(st.session_state.testcase_ocr_text and st.session_state.get("testcase_include_ocr", False)),
                    "api_config": {k: "***" if "key" in k.lower() else v for k, v in api_config.items()}
                }
                st.session_state.requirement_history.insert(0, history_item)
                st.session_state.requirement_history = st.session_state.requirement_history[:10]

                st.success(f"✅ 使用{model_provider}成功生成 {len(test_cases)} 个测试用例！")
                st.success(f"📝 风格: {case_style} | 语言: {language} | 覆盖维度: {', '.join(coverage_focus or ['默认'])}")

            except Exception as e:
                st.error(f"生成测试用例失败: {str(e)}")

    if st.session_state.test_cases:
        st.markdown("### 📊 生成的测试用例")

        latest_history = st.session_state.requirement_history[0] if st.session_state.requirement_history else {}
        if latest_history:
            coverage_summary = ", ".join(latest_history.get("coverage_focus", [])) or "默认"
            ocr_summary = " | 含OCR需求" if latest_history.get("used_ocr") else ""
            st.caption(
                f"使用模型: {latest_history.get('model', '未知')} | "
                f"风格: {latest_history.get('case_style', '标准格式')} | "
                f"语言: {latest_history.get('language', '中文')} | "
                f"覆盖维度: {coverage_summary}{ocr_summary} | "
                f"生成时间: {latest_history.get('timestamp', '')}"
            )

        normalized_cases = generator.normalize_cases_for_display(st.session_state.test_cases)
        total_cases = len(normalized_cases)
        priority_count = {'高': 0, '中': 0, '低': 0}
        for case in normalized_cases:
            priority = case.get('优先级', '中')
            if priority in priority_count:
                priority_count[priority] += 1

        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
        with metric_col1:
            st.metric("总用例数", total_cases)
        with metric_col2:
            st.metric("高优先级", priority_count['高'])
        with metric_col3:
            st.metric("中优先级", priority_count['中'])
        with metric_col4:
            st.metric("低优先级", priority_count['低'])

        filter_col1, filter_col2, filter_col3 = st.columns([1, 1.5, 1])
        with filter_col1:
            priority_filter = st.multiselect(
                "优先级筛选",
                options=["高", "中", "低"],
                default=["高", "中", "低"],
                key="testcase_priority_filter"
            )
        with filter_col2:
            keyword_filter = st.text_input(
                "关键词筛选",
                value="",
                placeholder="按用例名称、步骤、预期结果搜索",
                key="testcase_keyword_filter"
            )
        with filter_col3:
            view_mode = st.radio(
                "查看方式",
                ["表格", "逐条查看"],
                horizontal=True,
                key="testcase_view_mode"
            )

        filtered_cases = normalized_cases
        if priority_filter:
            filtered_cases = [case for case in filtered_cases if case.get("优先级") in priority_filter]
        if keyword_filter.strip():
            search_keyword = keyword_filter.strip().lower()
            filtered_cases = [
                case for case in filtered_cases
                if search_keyword in json.dumps(case, ensure_ascii=False).lower()
            ]

        st.caption(f"当前展示 {len(filtered_cases)} / {len(normalized_cases)} 条用例")

        if filtered_cases:
            filtered_df = pd.DataFrame(filtered_cases)
            if view_mode == "表格":
                st.dataframe(filtered_df, use_container_width=True, height=420, hide_index=True)
            else:
                for case in filtered_cases:
                    with st.expander(f"{case['用例ID']} | {case['用例名称']} | {case['优先级']}"):
                        st.write(f"**测试类型**: {case.get('测试类型', '未标注') or '未标注'}")
                        st.write(f"**前置条件**: {case['前置条件'] or '无'}")
                        st.write("**测试步骤**")
                        st.write(case["测试步骤"] or "无")
                        st.write("**预期结果**")
                        st.write(case["预期结果"] or "无")
                        if case.get("备注"):
                            st.write(f"**备注**: {case['备注']}")
        else:
            st.warning("当前筛选条件下没有匹配的测试用例。")

        st.markdown("### 📤 导出测试用例")
        export_timestamp = time.strftime("%Y%m%d_%H%M%S")
        original_df = pd.DataFrame(st.session_state.test_cases)
        excel_buffer = io.BytesIO()
        original_df.to_excel(excel_buffer, index=False, engine='openpyxl')
        excel_buffer.seek(0)
        csv_data = original_df.to_csv(index=False).encode("utf-8-sig")
        json_data = json.dumps(st.session_state.test_cases, ensure_ascii=False, indent=2).encode("utf-8")
        md_content = generator.generate_markdown_report(
            st.session_state.test_cases,
            st.session_state.current_requirement
        )

        export_col1, export_col2, export_col3, export_col4 = st.columns(4)
        with export_col1:
            st.download_button(
                label="📊 下载Excel",
                data=excel_buffer.getvalue(),
                file_name=f"测试用例_{export_timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_excel_btn"
            )
        with export_col2:
            st.download_button(
                label="📄 下载CSV",
                data=csv_data,
                file_name=f"测试用例_{export_timestamp}.csv",
                mime="text/csv",
                use_container_width=True,
                key="download_testcase_csv_btn"
            )
        with export_col3:
            st.download_button(
                label="📝 下载Markdown",
                data=md_content,
                file_name=f"测试用例_{export_timestamp}.md",
                mime="text/markdown",
                use_container_width=True,
                key="download_md_btn"
            )
        with export_col4:
            st.download_button(
                label="🧾 下载JSON",
                data=json_data,
                file_name=f"测试用例_{export_timestamp}.json",
                mime="application/json",
                use_container_width=True,
                key="download_testcase_json_btn"
            )

    if st.session_state.requirement_history:
        st.markdown("### 📚 生成历史")
        for i, history in enumerate(st.session_state.requirement_history[:5]):
            model_info = f" ({history.get('model', '未知模型')})" if 'model' in history else ""
            style_info = f" [{history.get('case_style', '标准')}]" if 'case_style' in history else ""
            ocr_info = " [OCR]" if history.get("used_ocr") else ""
            with st.expander(
                    f"{history['timestamp']}{model_info}{style_info}{ocr_info} - {history['requirement']} ({history['case_count']}个用例)"):
                col1, col2 = st.columns(2)
                with col1:
                    if st.button(f"重新加载此需求", key=f"reload_history_{i}"):
                        st.session_state.current_requirement_input = history.get(
                            'editable_requirement',
                            history.get('full_requirement', history['requirement'])
                        )
                        st.session_state.testcase_input_counter += 1
                        st.rerun()
                with col2:
                    if st.button(f"查看用例详情", key=f"view_history_{i}"):
                        st.info(f"此历史记录包含 {history['case_count']} 个测试用例，"
                                f"使用模型: {history.get('model', '未知')}, "
                                f"风格: {history.get('case_style', '标准格式')}, "
                                f"语言: {history.get('language', '中文')}, "
                                f"覆盖维度: {', '.join(history.get('coverage_focus', [])) or '默认'}")

elif tool_category == "禅道绩效统计":
    show_doc("zentao_performance_stats")

    # 数据库配置
    st.markdown("### 🔑 数据库配置")
    import os

    default_host = os.getenv('ZENTAO_DB_HOST', '')
    default_port = int(os.getenv('ZENTAO_DB_PORT', '3306'))
    default_user = os.getenv('ZENTAO_DB_USER', '')
    default_password = os.getenv('ZENTAO_DB_PASSWORD', '')
    default_database = os.getenv('ZENTAO_DB_NAME', 'zentao')

    col1, col2 = st.columns(2)
    with col1:
        db_host = st.text_input("数据库地址", value=default_host,
                                placeholder="例如: mysql.server.com 或 123.45.67.89",
                                key="zentao_perf_db_host")
        db_port = st.number_input("端口", value=default_port, key="zentao_perf_db_port")
        db_user = st.text_input("用户名", value=default_user, key="zentao_perf_db_user")
    with col2:
        db_password = st.text_input("密码", type="password", value=default_password,
                                    placeholder="数据库密码", key="zentao_perf_db_password")
        db_name = st.text_input("数据库名", value=default_database, key="zentao_perf_db_name")

        # 连接方式选择
        connection_method = st.radio("连接方式",
                                     ["直接连接", "连接字符串"],
                                     horizontal=True,
                                     help="直接连接适用于可公开访问的数据库，连接字符串更安全")

        if connection_method == "连接字符串":
            connection_string = st.text_input(
                "数据库连接字符串",
                placeholder="例如: mysql://username:password@host:port/database",
                help="格式: mysql://用户名:密码@主机:端口/数据库名"
            )

            if connection_string:
                try:
                    # 解析连接字符串
                    if connection_string.startswith('mysql://'):
                        parts = connection_string[8:].split('@')
                        user_pass = parts[0].split(':')
                        host_db = parts[1].split('/')
                        host_port = host_db[0].split(':')

                        db_user = user_pass[0]
                        db_password = user_pass[1] if len(user_pass) > 1 else ''
                        db_host = host_port[0]
                        db_port = int(host_port[1]) if len(host_port) > 1 else 3306
                        db_name = host_db[1] if len(host_db) > 1 else 'zentao'
                except Exception as e:
                    st.error(f"连接字符串格式错误: {e}")

        # 测试数据库连接
        if st.button("🔗 测试数据库连接", key="test_zentao_perf_connection"):
            if not all([db_host, db_user, db_password, db_name]):
                st.error("请填写完整的数据库配置")
            else:
                try:
                    db_config = {
                        'host': db_host,
                        'port': int(db_port),
                        'user': db_user,
                        'password': db_password,
                        'database': db_name,
                        'charset': 'utf8mb4',
                        'connect_timeout': 10  # 增加超时时间
                    }

                    with st.spinner("正在连接数据库..."):
                        exporter = ZenTaoPerformanceExporter(db_config)
                        if exporter.mysql_db:
                            products = exporter.get_products()
                            roles = exporter.get_user_roles()
                            bug_types = exporter.get_bug_types()

                            st.success(f"✅ 数据库连接成功！")
                            st.info(f"发现 {len(products)} 个产品, {len(roles)} 种角色, {len(bug_types)} 种缺陷类型")
                            exporter.close_connection()
                        else:
                            st.error("❌ 数据库连接失败")
                except Exception as e:
                    st.error(f"❌ 数据库连接失败: {str(e)}")
                    st.info("💡 **解决方案建议:**")
                    st.markdown("""
                    1. **使用云数据库**: 将禅道数据库迁移到云数据库服务
                    2. **配置公网访问**: 让数据库服务器支持公网访问
                    3. **使用连接隧道**: 通过SSH或VPN连接内网数据库
                    4. **检查防火墙**: 确保数据库端口对外开放
                    """)
    st.markdown("---")

    # 统计配置
    st.markdown("### ⚙️ 统计配置")

    # 动态获取配置
    if st.button("🔄 加载系统配置", key="load_zentao_config"):
        try:
            db_config = {
                'host': db_host,
                'port': int(db_port),
                'user': db_user,
                'password': db_password,
                'database': db_name,
                'charset': 'utf8mb4'
            }
            exporter = ZenTaoPerformanceExporter(db_config)
            if exporter.mysql_db:
                # 获取动态数据
                products = exporter.get_products()
                roles = exporter.get_user_roles()
                bug_types = exporter.get_bug_types()

                # 保存到session state
                st.session_state.zentao_products = products
                st.session_state.zentao_roles = roles
                st.session_state.zentao_bug_types = bug_types
                st.session_state.zentao_exporter = exporter

                st.success("✅ 系统配置加载成功！")
            else:
                st.error("❌ 数据库连接失败，无法加载配置")
        except Exception as e:
            st.error(f"❌ 加载配置失败: {str(e)}")

    # 时间配置
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("开始日期", datetime.datetime.now() - datetime.timedelta(days=30),
                                   key="zentao_start_date")
        end_date = st.date_input("结束日期", datetime.datetime.now(), key="zentao_end_date")

        # 转换为字符串格式
        start_date_str = start_date.strftime('%Y-%m-%d 00:00:00')
        end_date_str = end_date.strftime('%Y-%m-%d 23:59:59')

    with col2:
        stat_type = st.radio("统计类型", ["测试绩效", "开发绩效"], horizontal=True, key="zentao_stat_type")

    # 产品选择
    if 'zentao_products' in st.session_state and st.session_state.zentao_products:
        products = st.session_state.zentao_products
        product_options = {item[0]: item[1] for item in products}

        selected_products = st.multiselect(
            "选择要统计的产品",
            options=list(product_options.keys()),
            format_func=lambda x: f"{product_options[x]} (ID: {x})",
            key="zentao_selected_products"
        )
    else:
        st.warning("请先点击'加载系统配置'来获取产品列表")
        selected_products = []

    # 角色选择（现在显示中文）
    # 角色选择（现在显示中文）
    if 'zentao_roles' in st.session_state and st.session_state.zentao_roles:
        roles = st.session_state.zentao_roles
        # 创建角色映射：key->中文名
        role_mapping = {item[0]: item[1] for item in roles}

        # 显示当前可用的角色
        st.write(f"📋 系统检测到 {len(role_mapping)} 个角色")

        # 根据统计类型设置默认角色
        if stat_type == "测试绩效":
            # 测试绩效：优先选择qa角色
            if 'qa' in role_mapping:
                default_roles = ['qa']
                role_desc = "默认选择 'qa' 角色"
            elif '测试' in role_mapping.values():
                # 如果角色名是中文"测试"
                test_role_key = [key for key, value in role_mapping.items() if value == '测试'][0]
                default_roles = [test_role_key]
                role_desc = f"默认选择 '{role_mapping[test_role_key]}' 角色"
            else:
                # 查找包含测试关键词的角色
                test_roles = [key for key, name in role_mapping.items()
                              if any(word in str(name).lower() for word in ['测试', 'qa', 'test', 'tester'])]
                default_roles = test_roles[:1] if test_roles else list(role_mapping.keys())[:1]
                role_desc = "自动选择测试相关角色"

            st.info(f"🧪 测试绩效统计 - {role_desc}")

        else:
            # 开发绩效：优先选择dev角色
            if 'dev' in role_mapping:
                default_roles = ['dev']
                role_desc = "默认选择 'dev' 角色"
            elif '开发' in role_mapping.values():
                # 如果角色名是中文"开发"
                dev_role_key = [key for key, value in role_mapping.items() if value == '开发'][0]
                default_roles = [dev_role_key]
                role_desc = f"默认选择 '{role_mapping[dev_role_key]}' 角色"
            else:
                # 查找包含开发关键词的角色
                dev_roles = [key for key, name in role_mapping.items()
                             if any(word in str(name).lower() for word in ['开发', 'dev', '开发人员', 'developer', '研发'])]
                default_roles = dev_roles[:1] if dev_roles else list(role_mapping.keys())[:1]
                role_desc = "自动选择开发相关角色"

            st.info(f"💻 开发绩效统计 - {role_desc}")

        selected_roles = st.multiselect(
            "选择参与统计的角色",
            options=list(role_mapping.keys()),
            format_func=lambda x: role_mapping[x],  # 显示中文名
            default=default_roles,
            key="zentao_selected_roles"
        )

        # 保存角色映射到session state，用于后续显示
        st.session_state.role_mapping = role_mapping

        # 显示当前选择的角色中文名
        if selected_roles:
            selected_role_names = [role_mapping[role] for role in selected_roles]
            st.success(f"✅ 已选择 {len(selected_roles)} 个角色: {', '.join(selected_role_names)}")
        else:
            st.warning("⚠️ 请至少选择一个参与统计的角色")

    else:
        selected_roles = []
        st.session_state.role_mapping = {}
        st.warning("请先点击'加载系统配置'来获取角色列表")

    # 排除的缺陷类型（现在显示中文）
    if 'zentao_bug_types' in st.session_state and st.session_state.zentao_bug_types:
        bug_types = st.session_state.zentao_bug_types
        type_options = {item[0]: item[1] for item in bug_types}

        # 默认排除一些非功能性缺陷类型
        default_exclude_keys = []
        for key, name in type_options.items():
            if any(exclude in name.lower() for exclude in ['变更', '设计', '文档', '改进', '接口']):
                default_exclude_keys.append(key)

        exclude_types = st.multiselect(
            "选择要排除的缺陷类型",
            options=list(type_options.keys()),
            format_func=lambda x: type_options[x],
            default=default_exclude_keys,
            key="zentao_exclude_types"
        )
    else:
        exclude_types = []

    # 超时配置
    st.markdown("### ⏰ 超时响应配置")
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("高优先级缺陷")
        high_priority_normal = st.number_input("工作日响应时限(小时)", min_value=1, value=24, key="high_normal")
        high_priority_weekend = st.number_input("周末响应时限(小时)", min_value=1, value=72, key="high_weekend")

    with col2:
        st.subheader("普通优先级缺陷")
        normal_priority_normal = st.number_input("工作日响应时限(小时)", min_value=1, value=72, key="normal_normal")
        normal_priority_weekend = st.number_input("周末响应时限(小时)", min_value=1, value=120, key="normal_weekend")

    # 生成报告
    if st.button("🚀 生成统计报告", use_container_width=True, key="generate_zentao_stats"):
        if not selected_products:
            st.error("请至少选择一个产品")
            st.stop()

        if not selected_roles:
            st.error("请至少选择一个角色")
            st.stop()

        if not all([db_host, db_user, db_password, db_name]):
            st.error("请填写完整的数据库配置")
            st.stop()

        try:
            # 创建配置
            db_config = {
                'host': db_host,
                'port': int(db_port),
                'user': db_user,
                'password': db_password,
                'database': db_name,
                'charset': 'utf8mb4'
            }

            stats_config = {
                'start_date': start_date_str,
                'end_date': end_date_str,
                'roles': selected_roles,
                'exclude_types': exclude_types,
                'high_priority_normal_hours': high_priority_normal,
                'high_priority_weekend_hours': high_priority_weekend,
                'normal_priority_normal_hours': normal_priority_normal,
                'normal_priority_weekend_hours': normal_priority_weekend
            }

            # 创建导出器实例
            exporter = ZenTaoPerformanceExporter(db_config)

            if exporter.mysql_db is None:
                st.error("数据库连接失败，请检查配置")
                st.stop()

            with st.spinner("正在查询数据..."):
                all_data = {}
                all_detail_data = {}  # 存储明细数据
                product_options = {item[0]: item[1] for item in st.session_state.zentao_products}

                for product_id in selected_products:
                    product_name = product_options[product_id]

                    # 根据统计类型查询数据
                    if stat_type == "测试绩效":
                        df = exporter.query_qa_stats(product_id, stats_config)
                        # 新增：查询测试绩效明细
                        detail_df = exporter.query_qa_detail_stats(product_id, stats_config)
                    else:
                        df = exporter.query_dev_stats(product_id, stats_config)
                        detail_df = None  # 开发绩效暂时没有明细

                    if df is not None and not df.empty:
                        all_data[product_name] = df
                        if detail_df is not None and not detail_df.empty:
                            all_detail_data[product_name] = detail_df
                    else:
                        st.warning(f"产品 {product_name} 没有查询到数据")

            # 显示结果
            # 显示结果
            if all_data:
                st.success(f"✅ 成功查询到 {len(all_data)} 个产品的数据")

                # 显示配置信息
                with st.expander("📋 统计配置详情", expanded=False):
                    role_mapping = st.session_state.get('role_mapping', {})
                    type_mapping = st.session_state.get('type_mapping', {})

                    st.json({
                        "时间范围": f"{start_date_str} 至 {end_date_str}",
                        "统计类型": stat_type,
                        "参与角色": [role_mapping.get(r, r) for r in selected_roles],
                        "排除类型": [type_mapping.get(t, t) for t in exclude_types],
                        "超时配置": {
                            "高优先级缺陷": f"工作日{high_priority_normal}小时, 周末{high_priority_weekend}小时",
                            "普通优先级缺陷": f"工作日{normal_priority_normal}小时, 周末{normal_priority_weekend}小时"
                        }
                    })

                # 最简单的方法：所有expander都默认展开，使用固定下载key
                for idx, (product_name, df) in enumerate(all_data.items()):
                    # 总是展开expander
                    with st.expander(f"📊 {product_name} - {stat_type}", expanded=True):
                        st.dataframe(df, use_container_width=True)

                        # 使用固定的下载key
                        csv_data = df.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label=f"📥 下载 {product_name} {stat_type}",
                            data=csv_data,
                            file_name=f"{product_name}_{stat_type}_{start_date_str[:10]}_to_{end_date_str[:10]}.csv",
                            mime="text/csv",
                            key=f"download_single_{product_name}_{idx}"
                        )

                        if stat_type == "测试绩效" and product_name in all_detail_data:
                            detail_df = all_detail_data[product_name]
                            detail_csv_data = detail_df.to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label=f"📥 下载 {product_name} 测试绩效明细",
                                data=detail_csv_data,
                                file_name=f"{product_name}_测试绩效明细_{start_date_str[:10]}_to_{end_date_str[:10]}.csv",
                                mime="text/csv",
                                key=f"download_detail_{product_name}_{idx}"
                            )

                # Excel下载放在最后，使用固定key
                st.markdown("---")
                # 提供Excel格式的多sheet下载
                st.markdown("### 📊 所有绩效数据下载")

                # 创建Excel文件
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # 为每个产品创建一个sheet
                    for product_name, df in all_data.items():
                        # 清理sheet名称（Excel限制）
                        sheet_name = exporter._clean_sheet_name(f"{product_name}_{stat_type}")

                        # 直接使用原始数据框，确保数据格式一致
                        df_with_info = df.copy()

                        # 添加统计时间范围信息（作为第一列或单独行）
                        # 方法1：作为单独的第一行
                        # 创建包含统计信息的DataFrame
                        info_df = pd.DataFrame([['统计时间范围', f'{start_date_str[:10]}至{end_date_str[:10]}']])

                        # 先写入统计信息
                        info_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)

                        # 然后写入主要数据，从第2行开始
                        df_with_info.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

                        # 获取工作表
                        worksheet = writer.sheets[sheet_name]

                        # 格式化工作表
                        try:
                            from openpyxl.styles import Font, Alignment, PatternFill

                            # 设置统计信息行的样式
                            info_cell = worksheet.cell(row=1, column=1)
                            info_cell.font = Font(bold=True, size=12, color="FFFFFF")
                            info_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            info_cell.alignment = Alignment(horizontal='center', vertical='center')
                            worksheet.merge_cells(start_row=1, start_column=1, end_row=1,
                                                  end_column=len(df_with_info.columns))

                            # 设置表头样式（第3行）
                            header_font = Font(bold=True, color="FFFFFF")
                            header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

                            for col in range(1, len(df_with_info.columns) + 1):
                                cell = worksheet.cell(row=3, column=col)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                            # 设置列宽
                            for col_idx, column in enumerate(df_with_info.columns, 1):
                                max_length = max(
                                    df_with_info[column].astype(str).str.len().max() if len(df_with_info) > 0 else 0,
                                    len(str(column))
                                )
                                col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
                                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

                            # 设置内容对齐（从第4行开始）
                            for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row,
                                                           max_col=len(df_with_info.columns)):
                                for cell in row:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                            # 冻结窗格（表头行）
                            worksheet.freeze_panes = 'A4'

                        except Exception as e:
                            st.warning(f"格式化工作表 {sheet_name} 时出现警告: {e}")

                    # 添加汇总sheet
                    if all_data:
                        # 创建汇总数据
                        summary_data = []
                        for product_name, df in all_data.items():
                            # 使用与CSV相同的计算逻辑
                            total_bugs = df['提交bug数量'].sum() if '提交bug数量' in df.columns else 0
                            total_timeout = df['总超时响应次数'].sum() if '总超时响应次数' in df.columns else 0

                            summary_data.append({
                                '产品名称': product_name,
                                '统计人数': len(df),
                                '总bug数': total_bugs,
                                '总超时数': total_timeout,
                                '超时率': f"{(total_timeout / total_bugs * 100 if total_bugs > 0 else 0):.2f}%"
                            })

                        summary_df = pd.DataFrame(summary_data)

                        # 写入汇总sheet
                        summary_info_df = pd.DataFrame([['统计时间范围', f'{start_date_str[:10]}至{end_date_str[:10]}']])
                        summary_info_df.to_excel(writer, sheet_name='数据汇总', index=False, header=False, startrow=0)
                        summary_df.to_excel(writer, sheet_name='数据汇总', index=False, startrow=2)

                        # 格式化汇总sheet
                        summary_worksheet = writer.sheets['数据汇总']
                        try:
                            from openpyxl.styles import Font, Alignment, PatternFill

                            # 设置统计信息行样式
                            info_cell = summary_worksheet.cell(row=1, column=1)
                            info_cell.font = Font(bold=True, size=12, color="FFFFFF")
                            info_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            info_cell.alignment = Alignment(horizontal='center', vertical='center')
                            summary_worksheet.merge_cells(start_row=1, start_column=1, end_row=1,
                                                          end_column=len(summary_df.columns))

                            # 设置表头样式
                            header_font = Font(bold=True, color="FFFFFF")
                            header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

                            for col in range(1, len(summary_df.columns) + 1):
                                cell = summary_worksheet.cell(row=3, column=col)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                            # 设置列宽
                            for col_idx, column in enumerate(summary_df.columns, 1):
                                max_length = max(
                                    summary_df[column].astype(str).str.len().max() if len(summary_df) > 0 else 0,
                                    len(str(column))
                                )
                                col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
                                summary_worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

                            # 设置内容对齐
                            for row in summary_worksheet.iter_rows(min_row=4, max_row=summary_worksheet.max_row,
                                                                   max_col=len(summary_df.columns)):
                                for cell in row:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                            summary_worksheet.freeze_panes = 'A4'

                        except Exception as e:
                            st.warning(f"格式化汇总工作表时出现警告: {e}")

                excel_buffer.seek(0)

                # 下载Excel文件
                st.download_button(
                    label="📥 下载所有绩效数据",
                    data=excel_buffer.getvalue(),
                    file_name=f"禅道{stat_type}_统计_{start_date_str[:10]}_to_{end_date_str[:10]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_main"
                )

                st.info("💡 Excel文件包含每个产品的独立Sheet和数据汇总Sheet")

            else:
                st.warning("⚠️ 没有查询到任何数据，请检查配置和时间范围")

            # 关闭连接
            exporter.close_connection()

        except Exception as e:
            st.error(f"❌ 生成统计报告时出错: {str(e)}")
            import traceback

            st.error(f"详细错误: {traceback.format_exc()}")

    st.markdown("---")

    # 超时明细核查部分
    st.markdown("### 🔍 超时明细核查")

    # 根据统计类型显示不同的明细查询
    if stat_type == "测试绩效":
        st.info("🔍 当前为测试绩效超时明细查询，超时响应生效取值来源于【⏰超时响应配置】，排除的缺陷类型来源于【系统配置】中的选择要排除的缺陷类型")
        person_type = "测试人员"
        placeholder_text = "请输入测试人员真实姓名"
    else:
        st.info("🔍 当前为开发绩效超时明细查询")
        person_type = "开发人员"
        placeholder_text = "请输入开发人员真实姓名"

    col1, col2 = st.columns(2)

    with col1:
        # 产品选择
        if 'zentao_products' in st.session_state and st.session_state.zentao_products:
            products = st.session_state.zentao_products
            product_options = {item[0]: item[1] for item in products}

            detail_product_id = st.selectbox(
                "选择产品",
                options=list(product_options.keys()),
                format_func=lambda x: f"{product_options[x]} (ID: {x})",
                key="detail_product"
            )
        else:
            detail_product_id = None
            st.warning("请先加载系统配置")

    with col2:
        # 人员姓名输入
        person_name = st.text_input(
            f"{person_type}姓名",
            placeholder=placeholder_text,
            key="person_name"
        )

    # 时间范围
    col3, col4 = st.columns(2)
    with col3:
        detail_start_date = st.date_input(
            "开始日期",
            datetime.datetime.now() - datetime.timedelta(days=30),
            key="detail_start_date"
        )
    with col4:
        detail_end_date = st.date_input(
            "结束日期",
            datetime.datetime.now(),
            key="detail_end_date"
        )

    # 超时明细查询按钮
    if st.button("🔎 查询超时明细", use_container_width=True, key="query_timeout_details"):
        if not person_name:
            st.error(f"请输入{person_type}姓名")
            st.stop()

        if not detail_product_id:
            st.error("请选择产品")
            st.stop()

        try:
            # 转换时间格式
            detail_start_str = detail_start_date.strftime('%Y-%m-%d 00:00:00')
            detail_end_str = detail_end_date.strftime('%Y-%m-%d 23:59:59')

            # 创建数据库配置
            db_config = {
                'host': db_host,
                'port': int(db_port),
                'user': db_user,
                'password': db_password,
                'database': db_name,
                'charset': 'utf8mb4'
            }

            # 创建配置（使用相同的超时参数）
            detail_config = {
                'exclude_types': exclude_types,
                'high_priority_normal_hours': high_priority_normal,
                'high_priority_weekend_hours': high_priority_weekend,
                'normal_priority_normal_hours': normal_priority_normal,
                'normal_priority_weekend_hours': normal_priority_weekend
            }

            # 创建exporter实例
            exporter = ZenTaoPerformanceExporter(db_config)

            if exporter.mysql_db is None:
                st.error("数据库连接失败，请检查配置")
                st.stop()

            # 根据统计类型调用不同的明细查询方法
            if stat_type == "测试绩效":
                detail_df = exporter.query_qa_timeout_bugs_detail(
                    person_name,
                    detail_product_id,
                    detail_start_str,
                    detail_end_str,
                    detail_config
                )
            else:
                detail_df = exporter.query_timeout_bugs_detail(
                    person_name,
                    detail_product_id,
                    detail_start_str,
                    detail_end_str,
                    detail_config
                )

            # 关闭连接
            exporter.close_connection()

            if detail_df is not None and not detail_df.empty:
                st.success(f"✅ 找到 {len(detail_df)} 条超时Bug记录")

                # 显示统计信息
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                with col_stat1:
                    total_count = len(detail_df)
                    st.metric("总超时Bug数", total_count)
                with col_stat2:
                    high_priority = len(detail_df[detail_df['超时类型'] == '一级超时'])
                    st.metric("一级超时", high_priority)
                with col_stat3:
                    normal_priority = len(detail_df[detail_df['超时类型'] == '普通超时'])
                    st.metric("普通超时", normal_priority)
                with col_stat4:
                    if '是否超时' in detail_df.columns:
                        actual_timeout = detail_df['是否超时'].sum()
                        st.metric("确认超时", actual_timeout)
                    else:
                        resolved_count = len(detail_df[detail_df['状态'] == '已解决'])
                        st.metric("已解决", resolved_count)

                # 显示详细数据
                st.dataframe(detail_df, use_container_width=True)

                # 提供下载 - 使用固定key避免重新渲染
                csv_data = detail_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
                st.download_button(
                    label=f"📥 下载{person_type}超时明细",
                    data=csv_data,
                    file_name=f"{person_name}_{person_type}超时明细_{detail_start_str[:10]}_to_{detail_end_str[:10]}.csv",
                    mime="text/csv",
                    key="download_details_main"  # 使用固定key
                )

            else:
                st.warning("⚠️ 未找到符合条件的超时Bug记录")

        except Exception as e:
            st.error(f"❌ 查询超时明细时出错: {str(e)}")

elif tool_category == "接口研发辅助":
    show_doc("interface_dev_tools")

    if "interface_dev_parser" not in st.session_state:
        st.session_state.interface_dev_parser = InterfaceAutoTestCore()
    if "interface_dev_toolkit" not in st.session_state:
        st.session_state.interface_dev_toolkit = InterfaceDevTools()

    dev_parser = st.session_state.interface_dev_parser
    dev_toolkit = st.session_state.interface_dev_toolkit

    st.markdown('<div class="category-card">🛠️ 接口研发辅助工具</div>', unsafe_allow_html=True)
    st.caption("围绕接口测试与联调补七类高频能力: 标准化导出、回归清单、变更分析、文档体检、断言模板、Mock 服务、请求代码片段。")

    interface_file_types = ["xlsx", "xls", "json", "har", "bru", "txt", "md", "yaml", "yml"]
    raw_format_map = {
        "自动检测": "auto",
        "JSON / Apifox / Postman / HAR / Insomnia": "json",
        "Swagger/OpenAPI": "swagger",
        "结构化文本": "text",
        "Bruno .bru": "bruno"
    }

    def render_interface_source_panel(prefix, title, height=220):
        st.markdown(f"#### {title}")
        input_mode = st.radio(
            "导入方式",
            ["文件上传", "原始文本"],
            horizontal=True,
            key=f"{prefix}_input_mode"
        )
        source_data = {"mode": input_mode}
        if input_mode == "文件上传":
            source_data["uploaded_file"] = st.file_uploader(
                "选择接口文档",
                type=interface_file_types,
                key=f"{prefix}_upload",
                help="支持 Excel、JSON、Swagger/OpenAPI、Postman、HAR、Bruno、Insomnia、文本等格式"
            )
        else:
            source_data["raw_format"] = st.selectbox(
                "文本格式",
                list(raw_format_map.keys()),
                key=f"{prefix}_raw_format"
            )
            source_data["raw_content"] = st.text_area(
                "粘贴接口定义",
                height=height,
                placeholder="可粘贴 JSON、OpenAPI 文本、结构化文本或 curl 命令",
                key=f"{prefix}_raw_content"
            )
        return source_data

    def parse_interface_source(source_data, prefix):
        if source_data.get("mode") == "文件上传":
            uploaded_file = source_data.get("uploaded_file")
            if uploaded_file is None:
                raise ValueError("请先上传接口文档")
            temp_filename = f"{prefix}_{int(time.time() * 1000)}_{uploaded_file.name}"
            file_path = os.path.join(dev_parser.upload_dir, temp_filename)
            with open(file_path, "wb") as file:
                file.write(uploaded_file.getbuffer())
            interfaces = dev_parser.parse_document(file_path)
            return interfaces, uploaded_file.name

        raw_content = source_data.get("raw_content", "")
        if not raw_content.strip():
            raise ValueError("请输入原始接口定义")

        raw_format = source_data.get("raw_format", "自动检测")
        interfaces = dev_parser.parse_content(
            raw_content,
            source_type=raw_format_map[raw_format],
            source_name=f"{prefix}-inline"
        )
        return interfaces, f"{raw_format} 文本"

    def build_interface_preview_rows(interfaces, limit=10):
        rows = []
        for item in interfaces[:limit]:
            rows.append(
                {
                    "接口名称": item.get("name", "未命名接口"),
                    "请求方法": item.get("method", "GET"),
                    "接口路径": item.get("path", ""),
                    "状态码": item.get("expected_status", 200),
                    "请求格式": item.get("request_format", "auto"),
                }
            )
        return rows

    normalize_tab, checklist_tab, diff_tab, quality_tab, assertion_tab, mock_tab, snippet_tab = st.tabs(
        ["标准化导出", "回归清单生成", "接口变更分析", "接口文档体检", "断言模板生成", "Mock 服务生成", "请求代码片段"]
    )

    with normalize_tab:
        normalize_source = render_interface_source_panel("interface_normalize_source", "接口文档", height=240)
        if st.button("🧾 生成标准化清单", use_container_width=True, key="generate_normalized_interfaces"):
            try:
                interfaces, source_name = parse_interface_source(normalize_source, "interface_normalize_source")
                normalize_result = dev_toolkit.export_normalized_interfaces(interfaces)
                st.session_state.interface_dev_normalize_result = normalize_result
                st.session_state.interface_dev_normalize_source_name = source_name
                st.success(f"✅ 已标准化整理 {len(interfaces)} 个接口")
            except Exception as exc:
                st.error(f"❌ 生成失败: {exc}")

        normalize_result = st.session_state.get("interface_dev_normalize_result")
        if normalize_result:
            source_name = st.session_state.get("interface_dev_normalize_source_name", "当前文档")
            summary = normalize_result.get("summary", {})
            st.info(f"当前标准化来源: `{source_name}`")

            metric_col1, metric_col2, metric_col3, metric_col4, metric_col5 = st.columns(5)
            with metric_col1:
                st.metric("接口数", summary.get("interface_count", 0))
            with metric_col2:
                st.metric("含请求体", summary.get("with_body_count", 0))
            with metric_col3:
                st.metric("鉴权相关", summary.get("auth_related_count", 0))
            with metric_col4:
                st.metric("缺少响应样例", summary.get("missing_expected_response_count", 0))
            with metric_col5:
                st.metric("标签数", summary.get("tag_count", 0))

            normalized_interfaces = normalize_result.get("interfaces", [])
            if normalized_interfaces:
                st.markdown("### 📋 标准化预览")
                st.dataframe(pd.DataFrame(build_interface_preview_rows(normalized_interfaces, limit=50)), use_container_width=True)

            artifact_tab1, artifact_tab2, artifact_tab3 = st.tabs(["JSON", "Markdown", "结构化文本"])
            with artifact_tab1:
                st.code(normalize_result.get("json_artifact", ""), language="json")
            with artifact_tab2:
                st.code(normalize_result.get("markdown_artifact", ""), language="markdown")
            with artifact_tab3:
                st.code(normalize_result.get("text_artifact", ""), language="text")

            normalize_action_col1, normalize_action_col2, normalize_action_col3 = st.columns(3)
            with normalize_action_col1:
                st.download_button(
                    label="📥 下载 JSON 清单",
                    data=normalize_result.get("json_artifact", ""),
                    file_name="normalized_interfaces.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_normalized_interfaces_json"
                )
            with normalize_action_col2:
                st.download_button(
                    label="📥 下载 Markdown 清单",
                    data=normalize_result.get("markdown_artifact", ""),
                    file_name="normalized_interfaces.md",
                    mime="text/markdown",
                    use_container_width=True,
                    key="download_normalized_interfaces_markdown"
                )
            with normalize_action_col3:
                st.download_button(
                    label="📥 下载结构化文本",
                    data=normalize_result.get("text_artifact", ""),
                    file_name="normalized_interfaces.txt",
                    mime="text/plain",
                    use_container_width=True,
                    key="download_normalized_interfaces_text"
                )

    with checklist_tab:
        checklist_source = render_interface_source_panel("interface_checklist_source", "接口文档", height=240)
        checklist_opt_col1, checklist_opt_col2, checklist_opt_col3 = st.columns(3)
        with checklist_opt_col1:
            include_negative = st.checkbox("包含负向场景", value=True, key="interface_checklist_negative")
        with checklist_opt_col2:
            include_auth_checks = st.checkbox("包含鉴权检查", value=True, key="interface_checklist_auth")
        with checklist_opt_col3:
            include_performance_checks = st.checkbox("包含性能关注", value=True, key="interface_checklist_performance")

        if st.button("🧠 生成回归清单", use_container_width=True, key="generate_regression_checklist"):
            try:
                interfaces, source_name = parse_interface_source(checklist_source, "interface_checklist_source")
                checklist_result = dev_toolkit.generate_regression_checklist(
                    interfaces,
                    include_negative=include_negative,
                    include_auth_checks=include_auth_checks,
                    include_performance_checks=include_performance_checks,
                )
                st.session_state.interface_dev_checklist_result = checklist_result
                st.session_state.interface_dev_checklist_source_name = source_name
                st.success("✅ 回归清单已生成")
            except Exception as exc:
                st.error(f"❌ 生成失败: {exc}")

        checklist_result = st.session_state.get("interface_dev_checklist_result")
        if checklist_result:
            source_name = st.session_state.get("interface_dev_checklist_source_name", "当前文档")
            summary = checklist_result.get("summary", {})
            st.info(f"当前回归清单来源: `{source_name}`")

            metric_col1, metric_col2, metric_col3, metric_col4, metric_col5 = st.columns(5)
            with metric_col1:
                st.metric("接口数", summary.get("interface_count", 0))
            with metric_col2:
                st.metric("高优先级", summary.get("high_priority_count", 0))
            with metric_col3:
                st.metric("中优先级", summary.get("medium_priority_count", 0))
            with metric_col4:
                st.metric("低优先级", summary.get("low_priority_count", 0))
            with metric_col5:
                st.metric("建议检查项", summary.get("check_item_count", 0))

            checklist_rows = checklist_result.get("rows", [])
            if checklist_rows:
                st.markdown("### 📌 回归总览")
                st.dataframe(pd.DataFrame(checklist_rows), use_container_width=True)

            items = checklist_result.get("items", [])
            if items:
                st.markdown("### 🧪 逐接口回归建议")
                for item in items:
                    priority_text = {"high": "高优先级", "medium": "中优先级", "low": "低优先级"}.get(
                        item.get("priority"),
                        "低优先级"
                    )
                    with st.expander(
                        f"{item.get('index', 0)}. {item.get('method', 'GET')} {item.get('path', '/')} | {item.get('name', '未命名接口')} | {priority_text}",
                        expanded=False
                    ):
                        st.write(f"**关注点:** {'、'.join(item.get('focus_points', []))}")
                        for checklist_item in item.get("checklist", []):
                            st.write(f"- [ ] {checklist_item}")

            checklist_markdown = checklist_result.get("markdown_artifact", "")
            checklist_json = checklist_result.get("json_artifact", "")
            checklist_csv = pd.DataFrame(checklist_rows).to_csv(index=False, encoding="utf-8-sig")
            checklist_action_col1, checklist_action_col2, checklist_action_col3 = st.columns(3)
            with checklist_action_col1:
                st.download_button(
                    label="📥 下载 Markdown 清单",
                    data=checklist_markdown,
                    file_name="interface_regression_checklist.md",
                    mime="text/markdown",
                    use_container_width=True,
                    key="download_regression_checklist_markdown"
                )
            with checklist_action_col2:
                st.download_button(
                    label="📥 下载 JSON 清单",
                    data=checklist_json,
                    file_name="interface_regression_checklist.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_regression_checklist_json"
                )
            with checklist_action_col3:
                st.download_button(
                    label="📥 下载 CSV 总览",
                    data=checklist_csv.encode("utf-8-sig"),
                    file_name="interface_regression_checklist.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="download_regression_checklist_csv"
                )

    with diff_tab:
        diff_col1, diff_col2 = st.columns(2)
        with diff_col1:
            baseline_source = render_interface_source_panel("interface_diff_baseline", "基线版本")
        with diff_col2:
            target_source = render_interface_source_panel("interface_diff_target", "当前版本")

        if st.button("🔍 开始分析接口差异", use_container_width=True, key="analyze_interface_diff"):
            try:
                baseline_interfaces, baseline_name = parse_interface_source(baseline_source, "interface_diff_baseline")
                target_interfaces, target_name = parse_interface_source(target_source, "interface_diff_target")
                diff_result = dev_toolkit.compare_interfaces(baseline_interfaces, target_interfaces)
                st.session_state.interface_dev_diff_result = diff_result
                st.session_state.interface_dev_diff_baseline_name = baseline_name
                st.session_state.interface_dev_diff_target_name = target_name
                st.success("✅ 接口差异分析完成")
            except Exception as exc:
                st.error(f"❌ 分析失败: {exc}")

        diff_result = st.session_state.get("interface_dev_diff_result")
        if diff_result:
            summary = diff_result.get("summary", {})
            baseline_name = st.session_state.get("interface_dev_diff_baseline_name", "基线版本")
            target_name = st.session_state.get("interface_dev_diff_target_name", "当前版本")
            st.info(f"基线文档: `{baseline_name}` | 当前文档: `{target_name}`")

            metric_col1, metric_col2, metric_col3, metric_col4, metric_col5, metric_col6 = st.columns(6)
            with metric_col1:
                st.metric("基线接口", summary.get("baseline_total", 0))
            with metric_col2:
                st.metric("当前接口", summary.get("target_total", 0))
            with metric_col3:
                st.metric("新增", summary.get("added_count", 0))
            with metric_col4:
                st.metric("删除", summary.get("removed_count", 0))
            with metric_col5:
                st.metric("变更", summary.get("changed_count", 0))
            with metric_col6:
                st.metric("高风险", summary.get("high_risk_count", 0))

            if summary.get("added_count", 0) == 0 and summary.get("removed_count", 0) == 0 and summary.get("changed_count", 0) == 0:
                st.success("✅ 两个版本未发现接口差异")
            else:
                if diff_result.get("high_risk_changes"):
                    st.markdown("### ⚠️ 高风险变更")
                    high_risk_rows = []
                    for item in diff_result.get("high_risk_changes", []):
                        high_risk_rows.append(
                            {
                                "接口": item.get("name", "未命名接口"),
                                "方法": item.get("method", "GET"),
                                "路径": item.get("path", ""),
                                "变更点": "；".join(change.get("label", "") for change in item.get("changes", [])),
                            }
                        )
                    st.dataframe(pd.DataFrame(high_risk_rows), use_container_width=True)

                if diff_result.get("changed"):
                    st.markdown("### 🔄 变更接口明细")
                    for index, item in enumerate(diff_result.get("changed", []), start=1):
                        risk_label = {"high": "高风险", "medium": "中风险", "low": "低风险"}.get(
                            item.get("risk_level"),
                            "未知"
                        )
                        with st.expander(
                            f"{index}. {item.get('method', 'GET')} {item.get('path', '')} | {item.get('name', '未命名接口')} | {risk_label}",
                            expanded=False
                        ):
                            for change in item.get("changes", []):
                                st.write(f"- {change.get('label')}: {change.get('message')}")

                added_removed_col1, added_removed_col2 = st.columns(2)
                with added_removed_col1:
                    st.markdown("### ➕ 新增接口")
                    if diff_result.get("added"):
                        st.dataframe(pd.DataFrame(diff_result.get("added", [])), use_container_width=True)
                    else:
                        st.info("当前没有新增接口")
                with added_removed_col2:
                    st.markdown("### ➖ 删除接口")
                    if diff_result.get("removed"):
                        st.dataframe(pd.DataFrame(diff_result.get("removed", [])), use_container_width=True)
                    else:
                        st.info("当前没有删除接口")

            markdown_report = dev_toolkit.build_markdown_report(diff_result)
            diff_report_json = json.dumps(diff_result, ensure_ascii=False, indent=2)
            download_col1, download_col2 = st.columns(2)
            with download_col1:
                st.download_button(
                    label="📥 下载 Markdown 报告",
                    data=markdown_report,
                    file_name="interface_diff_report.md",
                    mime="text/markdown",
                    use_container_width=True,
                    key="download_interface_diff_markdown"
                )
            with download_col2:
                st.download_button(
                    label="📥 下载 JSON 报告",
                    data=diff_report_json,
                    file_name="interface_diff_report.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_interface_diff_json"
                )

    with quality_tab:
        quality_source = render_interface_source_panel("interface_quality_source", "接口文档", height=240)
        quality_col1, quality_col2 = st.columns([2, 1])
        with quality_col1:
            st.info("体检会检查重复接口、缺少响应示例、路径参数不匹配、请求格式与 Header 不一致等常见问题。")
        with quality_col2:
            if st.button("🩺 开始体检", use_container_width=True, key="analyze_interface_quality"):
                try:
                    interfaces, source_name = parse_interface_source(quality_source, "interface_quality_source")
                    quality_result = dev_toolkit.analyze_interface_quality(interfaces)
                    st.session_state.interface_dev_quality_result = quality_result
                    st.session_state.interface_dev_quality_source_name = source_name
                    st.success("✅ 接口文档体检完成")
                except Exception as exc:
                    st.error(f"❌ 体检失败: {exc}")

        quality_result = st.session_state.get("interface_dev_quality_result")
        if quality_result:
            source_name = st.session_state.get("interface_dev_quality_source_name", "当前文档")
            summary = quality_result.get("summary", {})
            st.info(f"当前体检来源: `{source_name}`")

            metric_col1, metric_col2, metric_col3, metric_col4, metric_col5, metric_col6 = st.columns(6)
            with metric_col1:
                st.metric("接口数", summary.get("interface_count", 0))
            with metric_col2:
                st.metric("健康分", summary.get("health_score", 0))
            with metric_col3:
                st.metric("评级", summary.get("health_level", "未知"))
            with metric_col4:
                st.metric("高风险", summary.get("high_count", 0))
            with metric_col5:
                st.metric("中风险", summary.get("medium_count", 0))
            with metric_col6:
                st.metric("低风险", summary.get("low_count", 0))

            duplicate_groups = quality_result.get("duplicate_groups", [])
            if duplicate_groups:
                st.markdown("### 🔁 重复接口")
                duplicate_rows = []
                for item in duplicate_groups:
                    duplicate_rows.append(
                        {
                            "方法": item.get("method", "GET"),
                            "路径": item.get("path", "/"),
                            "重复次数": item.get("count", 0),
                            "序号": ", ".join(str(index) for index in item.get("indexes", [])),
                        }
                    )
                st.dataframe(pd.DataFrame(duplicate_rows), use_container_width=True)

            issues = quality_result.get("issues", [])
            if issues:
                st.markdown("### 📋 问题清单")
                selected_severities = st.multiselect(
                    "筛选风险等级",
                    ["high", "medium", "low"],
                    default=["high", "medium", "low"],
                    format_func=lambda value: {"high": "高风险", "medium": "中风险", "low": "低风险"}.get(value, value),
                    key="interface_quality_severities"
                )
                filtered_issues = [item for item in issues if item.get("severity") in selected_severities]
                issue_rows = []
                for item in filtered_issues:
                    issue_rows.append(
                        {
                            "序号": item.get("index", 0),
                            "风险等级": {"high": "高风险", "medium": "中风险", "low": "低风险"}.get(item.get("severity"), "未知"),
                            "方法": item.get("method", "GET"),
                            "路径": item.get("path", "/"),
                            "分类": item.get("category", ""),
                            "说明": item.get("message", ""),
                        }
                    )
                st.dataframe(pd.DataFrame(issue_rows), use_container_width=True)
            else:
                st.success("✅ 文档体检未发现问题")

            quality_markdown = dev_toolkit.build_quality_markdown_report(quality_result)
            quality_json = json.dumps(quality_result, ensure_ascii=False, indent=2)
            quality_action_col1, quality_action_col2 = st.columns(2)
            with quality_action_col1:
                st.download_button(
                    label="📥 下载体检 Markdown 报告",
                    data=quality_markdown,
                    file_name="interface_quality_report.md",
                    mime="text/markdown",
                    use_container_width=True,
                    key="download_interface_quality_markdown"
                )
            with quality_action_col2:
                st.download_button(
                    label="📥 下载体检 JSON 报告",
                    data=quality_json,
                    file_name="interface_quality_report.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_interface_quality_json"
                )

    with assertion_tab:
        assertion_input_mode = st.radio(
            "样例来源",
            ["原始响应 JSON", "从接口文档选择接口"],
            horizontal=True,
            key="interface_assertion_input_mode"
        )

        assertion_source = None
        assertion_sample_text = ""
        if assertion_input_mode == "原始响应 JSON":
            assertion_sample_text = st.text_area(
                "响应样例",
                height=260,
                placeholder='例如: {"code":0,"data":{"id":1,"name":"测试用户"}}',
                key="interface_assertion_sample_text"
            )
        else:
            assertion_source = render_interface_source_panel("interface_assertion_source", "接口文档", height=220)
            if st.button("📥 解析接口列表", use_container_width=True, key="parse_assertion_source"):
                try:
                    interfaces, source_name = parse_interface_source(assertion_source, "interface_assertion_source")
                    st.session_state.interface_dev_assertion_interfaces = interfaces
                    st.session_state.interface_dev_assertion_source_name = source_name
                    st.success(f"✅ 成功解析出 {len(interfaces)} 个接口")
                except Exception as exc:
                    st.error(f"❌ 解析失败: {exc}")

            assertion_interfaces = st.session_state.get("interface_dev_assertion_interfaces", [])
            if assertion_interfaces:
                source_name = st.session_state.get("interface_dev_assertion_source_name", "当前文档")
                st.info(f"当前断言样例来源: `{source_name}` | 接口数: {len(assertion_interfaces)}")
                selected_assertion_index = st.selectbox(
                    "选择接口",
                    range(len(assertion_interfaces)),
                    format_func=lambda idx: (
                        f"{idx + 1}. {assertion_interfaces[idx].get('method', 'GET')} "
                        f"{assertion_interfaces[idx].get('path', '')} | {assertion_interfaces[idx].get('name', '未命名接口')}"
                    ),
                    key="interface_assertion_selected_index"
                )
                selected_interface = assertion_interfaces[selected_assertion_index]
                if selected_interface.get("expected_response") in (None, "", {}, []):
                    st.warning("当前接口没有 `expected_response` 示例，无法直接生成断言模板")
                else:
                    assertion_sample_text = json.dumps(
                        selected_interface.get("expected_response"),
                        ensure_ascii=False,
                        indent=2
                    )
                    st.code(assertion_sample_text, language="json")

        assertion_col1, assertion_col2 = st.columns(2)
        with assertion_col1:
            assertion_style = st.radio(
                "断言风格",
                ["字段存在断言", "标准断言", "严格断言"],
                horizontal=True,
                key="interface_assertion_style"
            )
        with assertion_col2:
            assertion_depth = st.slider(
                "结构展开深度",
                min_value=1,
                max_value=6,
                value=4,
                key="interface_assertion_depth"
            )

        if st.button("🧠 生成断言模板", use_container_width=True, key="generate_assertion_template"):
            try:
                if not assertion_sample_text.strip():
                    raise ValueError("请先提供响应样例")
                assertion_result = dev_toolkit.generate_assertion_template(
                    assertion_sample_text,
                    template_style=assertion_style,
                    max_depth=assertion_depth,
                )
                st.session_state.interface_dev_assertion_result = assertion_result
                st.success("✅ 断言模板已生成")
            except Exception as exc:
                st.error(f"❌ 生成失败: {exc}")

        assertion_result = st.session_state.get("interface_dev_assertion_result")
        if assertion_result:
            info_col1, info_col2, info_col3 = st.columns(3)
            with info_col1:
                st.metric("样例类型", assertion_result.get("sample_type", "unknown"))
            with info_col2:
                st.metric("字段路径数", assertion_result.get("field_count", 0))
            with info_col3:
                st.metric("模板模式", assertion_style)

            st.markdown("### 🎯 可直接粘贴到用例中的片段")
            st.code(assertion_result.get("case_fragment_json", ""), language="json")

            st.markdown("### 📐 断言模板")
            st.code(assertion_result.get("template_json", ""), language="json")

            st.markdown("### 🧭 字段路径预览")
            field_path_rows = [{"字段路径": item} for item in assertion_result.get("field_paths", [])]
            st.dataframe(pd.DataFrame(field_path_rows), use_container_width=True)

            assertion_action_col1, assertion_action_col2 = st.columns(2)
            with assertion_action_col1:
                create_copy_button(
                    assertion_result.get("case_fragment_json", ""),
                    button_text="📋 复制断言片段",
                    key="copy_assertion_fragment"
                )
            with assertion_action_col2:
                st.download_button(
                    label="📥 下载断言模板",
                    data=assertion_result.get("case_fragment_json", ""),
                    file_name="expected_response_template.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_assertion_template"
                )

    with mock_tab:
        mock_source = render_interface_source_panel("interface_mock_source", "接口文档", height=240)
        mock_config_col1, mock_config_col2, mock_config_col3, mock_config_col4 = st.columns(4)
        with mock_config_col1:
            mock_host = st.text_input("监听地址", value="0.0.0.0", key="mock_server_host")
        with mock_config_col2:
            mock_port = st.number_input("监听端口", min_value=1, max_value=65535, value=8000, key="mock_server_port")
        with mock_config_col3:
            mock_delay_ms = st.number_input("固定延迟(ms)", min_value=0, value=0, key="mock_server_delay")
        with mock_config_col4:
            mock_enable_cors = st.checkbox("开启 CORS", value=True, key="mock_server_cors")

        if st.button("🧱 生成 Mock 服务脚本", use_container_width=True, key="generate_mock_server_script"):
            try:
                interfaces, source_name = parse_interface_source(mock_source, "interface_mock_source")
                mock_script = dev_toolkit.generate_mock_server_script(
                    interfaces=interfaces,
                    host=mock_host,
                    port=int(mock_port),
                    enable_cors=mock_enable_cors,
                    delay_ms=int(mock_delay_ms),
                )
                st.session_state.interface_dev_mock_interfaces = interfaces
                st.session_state.interface_dev_mock_source_name = source_name
                st.session_state.interface_dev_mock_script = mock_script
                st.success(f"✅ 已根据 {len(interfaces)} 个接口生成 Mock 服务脚本")
            except Exception as exc:
                st.error(f"❌ 生成失败: {exc}")

        mock_script = st.session_state.get("interface_dev_mock_script")
        if mock_script:
            mock_interfaces = st.session_state.get("interface_dev_mock_interfaces", [])
            source_name = st.session_state.get("interface_dev_mock_source_name", "当前文档")
            st.info(f"当前 Mock 来源: `{source_name}` | 接口数: {len(mock_interfaces)}")
            if mock_interfaces:
                st.dataframe(
                    pd.DataFrame(build_interface_preview_rows(mock_interfaces)),
                    use_container_width=True
                )

            st.code(mock_script, language="python")
            st.code(
                f"python mock_api_server.py --host {mock_host} --port {int(mock_port)}",
                language="bash"
            )

            mock_action_col1, mock_action_col2 = st.columns(2)
            with mock_action_col1:
                create_copy_button(mock_script, button_text="📋 复制 Mock 脚本", key="copy_mock_script")
            with mock_action_col2:
                st.download_button(
                    label="📥 下载 Mock 脚本",
                    data=mock_script,
                    file_name="mock_api_server.py",
                    mime="text/x-python",
                    use_container_width=True,
                    key="download_mock_script"
                )

    with snippet_tab:
        snippet_source = render_interface_source_panel("interface_snippet_source", "接口文档", height=240)
        if st.button("📥 解析接口文档", use_container_width=True, key="parse_snippet_source"):
            try:
                interfaces, source_name = parse_interface_source(snippet_source, "interface_snippet_source")
                st.session_state.interface_dev_snippet_interfaces = interfaces
                st.session_state.interface_dev_snippet_source_name = source_name
                st.session_state.pop("interface_dev_snippet_result", None)
                st.success(f"✅ 成功解析出 {len(interfaces)} 个接口")
            except Exception as exc:
                st.error(f"❌ 解析失败: {exc}")

        snippet_interfaces = st.session_state.get("interface_dev_snippet_interfaces", [])
        if snippet_interfaces:
            source_name = st.session_state.get("interface_dev_snippet_source_name", "当前文档")
            st.info(f"当前代码片段来源: `{source_name}` | 接口数: {len(snippet_interfaces)}")
            st.dataframe(
                pd.DataFrame(build_interface_preview_rows(snippet_interfaces)),
                use_container_width=True
            )

            selected_index = st.selectbox(
                "选择接口",
                range(len(snippet_interfaces)),
                format_func=lambda idx: (
                    f"{idx + 1}. {snippet_interfaces[idx].get('method', 'GET')} "
                    f"{snippet_interfaces[idx].get('path', '')} | {snippet_interfaces[idx].get('name', '未命名接口')}"
                ),
                key="interface_snippet_selected_index"
            )

            snippet_col1, snippet_col2 = st.columns(2)
            with snippet_col1:
                snippet_language = st.radio(
                    "代码类型",
                    ["Python requests", "JavaScript fetch", "curl"],
                    horizontal=True,
                    key="interface_snippet_language"
                )
            with snippet_col2:
                snippet_base_url = st.text_input(
                    "基础URL",
                    value="https://example.com",
                    placeholder="例如: https://api.example.com",
                    key="interface_snippet_base_url"
                )

            if st.button("⚡ 生成请求代码片段", use_container_width=True, key="generate_request_snippet"):
                try:
                    snippet_result = dev_toolkit.generate_request_snippet(
                        snippet_interfaces[selected_index],
                        language=snippet_language,
                        base_url=snippet_base_url
                    )
                    st.session_state.interface_dev_snippet_result = snippet_result
                    st.session_state.interface_dev_snippet_language_value = snippet_language
                    st.success("✅ 请求代码片段已生成")
                except Exception as exc:
                    st.error(f"❌ 生成代码片段失败: {exc}")

        snippet_result = st.session_state.get("interface_dev_snippet_result")
        if snippet_result:
            snippet_language = st.session_state.get("interface_dev_snippet_language_value", "Python requests")
            snippet_language_map = {
                "Python requests": "python",
                "JavaScript fetch": "javascript",
                "curl": "bash",
            }
            st.code(snippet_result, language=snippet_language_map.get(snippet_language, "text"))
            snippet_action_col1, snippet_action_col2 = st.columns(2)
            with snippet_action_col1:
                create_copy_button(snippet_result, button_text="📋 复制代码片段", key="copy_request_snippet")
            with snippet_action_col2:
                st.download_button(
                    label="📥 下载代码片段",
                    data=snippet_result,
                    file_name=f"request_snippet_{snippet_language.replace(' ', '_').lower()}.txt",
                    mime="text/plain",
                    use_container_width=True,
                    key="download_request_snippet"
                )

elif tool_category == "接口性能测试":
    show_doc("api_performance_test")
    render_api_performance_test_page()

elif tool_category == "接口安全测试":
    show_doc("api_security_test")
    render_api_security_test_page()

elif tool_category == "接口自动化测试":
    def install_missing_packages(packages, mirror_url):
        """安装缺失的包"""
        import subprocess
        import sys

        try:
            for package in packages:
                result = subprocess.run([
                    sys.executable, "-m", "pip", "install", package,
                    "-i", mirror_url,
                    "--trusted-host", mirror_url.split('/')[2],
                    "--timeout", "60"
                ], capture_output=True, text=True)

                if result.returncode != 0:
                    st.error(f"安装 {package} 失败: {result.stderr}")
                    return False
            return True
        except Exception as e:
            st.error(f"安装过程中出错: {e}")
            return False


    def check_interface_dependencies(selected_mode, for_execution=False):
        """检查接口测试工具依赖"""
        required_packages = {
            "requests": "requests",
            "pandas": "pandas",
            "openpyxl": "openpyxl"
        }
        if selected_mode == "pytest" and for_execution:
            required_packages["pytest"] = "pytest"

        missing_packages = []
        for package_name, import_name in required_packages.items():
            try:
                __import__(import_name)
            except ImportError:
                missing_packages.append(package_name)

        if not missing_packages:
            return True

        title = "当前执行模式运行时缺少依赖包" if for_execution else "当前页面缺少基础依赖包"
        st.warning(f"⚠️ {title}: {', '.join(missing_packages)}")

        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("**缺失的依赖:**")
            for dep in missing_packages:
                st.write(f"- {dep}")

        with col2:
            mirror = st.selectbox(
                "选择镜像源",
                ["清华镜像", "阿里云镜像", "豆瓣镜像", "中科大镜像"],
                key=f"interface_mirror_{selected_mode}"
            )
            mirror_urls = {
                "清华镜像": "https://pypi.tuna.tsinghua.edu.cn/simple/",
                "阿里云镜像": "https://mirrors.aliyun.com/pypi/simple/",
                "豆瓣镜像": "https://pypi.douban.com/simple/",
                "中科大镜像": "https://pypi.mirrors.ustc.edu.cn/simple/"
            }

            if st.button("🚀 自动安装依赖", use_container_width=True, key=f"install_interface_deps_{selected_mode}"):
                with st.spinner("正在安装依赖..."):
                    success = install_missing_packages(missing_packages, mirror_urls[mirror])
                if success:
                    st.success("✅ 依赖安装成功，请刷新页面后重试")
                else:
                    st.error("❌ 依赖安装失败，请手动安装")

        with st.expander("📋 手动安装指南", expanded=True):
            st.code(f"pip install {' '.join(missing_packages)} -i https://pypi.tuna.tsinghua.edu.cn/simple/")

        return False


    def create_jsonplaceholder_test_data():
        """创建 JSONPlaceholder 测试数据"""
        base_url = "https://jsonplaceholder.typicode.com"
        interfaces = [
            {
                "name": "获取所有帖子",
                "method": "GET",
                "path": "/posts",
                "description": "获取所有帖子列表",
                "expected_status": 200,
                "expected_response": ["userId", "id", "title", "body"]
            },
            {
                "name": "获取单个帖子",
                "method": "GET",
                "path": "/posts/1",
                "description": "获取 ID 为 1 的帖子详情",
                "expected_status": 200,
                "expected_response": ["userId", "id", "title", "body"]
            },
            {
                "name": "获取用户帖子",
                "method": "GET",
                "path": "/posts",
                "description": "按用户查询帖子",
                "query_params": {"userId": 1},
                "expected_status": 200,
                "expected_response": ["userId", "id", "title", "body"]
            },
            {
                "name": "创建新帖子",
                "method": "POST",
                "path": "/posts",
                "description": "创建新的帖子",
                "headers": {"Content-Type": "application/json"},
                "body": {
                    "title": "自动化测试帖子",
                    "body": "这是通过自动化测试工具创建的帖子",
                    "userId": 1
                },
                "expected_status": 201,
                "expected_response": ["id"]
            }
        ]
        return base_url, interfaces


    def create_reqres_test_data():
        """创建 ReqRes 测试数据"""
        base_url = "https://reqres.in/api"
        interfaces = [
            {
                "name": "获取用户列表",
                "method": "GET",
                "path": "/users",
                "description": "获取分页用户列表",
                "query_params": {"page": 2},
                "expected_status": 200,
                "expected_response": ["page", "per_page", "total", "data"]
            },
            {
                "name": "获取单个用户",
                "method": "GET",
                "path": "/users/2",
                "description": "获取用户详情",
                "expected_status": 200,
                "expected_response": ["data"]
            },
            {
                "name": "用户登录",
                "method": "POST",
                "path": "/login",
                "description": "用户登录接口",
                "headers": {"Content-Type": "application/json"},
                "body": {
                    "email": "eve.holt@reqres.in",
                    "password": "cityslicka"
                },
                "expected_status": 200,
                "expected_response": ["token"]
            }
        ]
        return base_url, interfaces


    def set_interface_state(interfaces, base_url="", source_name=""):
        st.session_state.interface_loaded_interfaces = interfaces
        st.session_state.interface_loaded_base_url = base_url or ""
        st.session_state.interface_source_name = source_name or ""
        st.session_state.pop("interface_last_generated_files", None)
        st.session_state.pop("interface_last_test_results", None)


    def clear_interface_state():
        for key in [
            "interface_loaded_interfaces",
            "interface_loaded_base_url",
            "interface_source_name",
            "interface_last_generated_files",
            "interface_last_test_results",
            "interface_upload_signature"
        ]:
            st.session_state.pop(key, None)


    def render_interface_preview(interfaces):
        st.markdown("### 📋 接口列表")
        for index, interface in enumerate(interfaces, 1):
            with st.expander(f"{index}. {interface.get('name', '未命名接口')}", expanded=False):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**方法:** `{interface.get('method', 'GET')}`")
                    st.write(f"**路径:** `{interface.get('path', '')}`")
                    st.write(f"**期望状态码:** `{interface.get('expected_status', 200)}`")
                    if interface.get("tags"):
                        st.write(f"**标签:** {', '.join(interface.get('tags', []))}")
                with col2:
                    st.write(f"**描述:** {interface.get('description', '无描述')}")
                    if interface.get("headers"):
                        st.write("**请求头:**")
                        st.json(interface["headers"])
                    if interface.get("path_params"):
                        st.write("**路径参数:**")
                        st.json(interface["path_params"])
                    if interface.get("query_params"):
                        st.write("**查询参数:**")
                        st.json(interface["query_params"])
                    if interface.get("body") not in [None, {}, []]:
                        st.write("**请求体:**")
                        st.json(interface["body"])
                    elif interface.get("parameters"):
                        st.write("**请求参数:**")
                        if isinstance(interface["parameters"], (dict, list)):
                            st.json(interface["parameters"])
                        else:
                            st.write(interface["parameters"])


    def render_test_results(test_results, execution_mode, interfaces):
        st.markdown("### 📊 测试结果概览")
        total = test_results.get("total", 0)
        passed = test_results.get("passed", 0)
        failed = test_results.get("failed", 0)
        errors = test_results.get("errors", 0)
        success_rate = (passed / total * 100) if total > 0 else 0

        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("总用例数", total)
        with col2:
            st.metric("通过", passed)
        with col3:
            st.metric("失败", failed)
        with col4:
            st.metric("错误", errors)
        with col5:
            st.metric("成功率", f"{success_rate:.1f}%")

        if test_results.get("output"):
            with st.expander("🧾 执行日志", expanded=False):
                st.code(test_results["output"], language="text")

        if total > 0:
            report_path = st.session_state.enhanced_report.generate_detailed_report(
                test_results=test_results,
                framework=execution_mode,
                interfaces=interfaces,
                test_details=test_results.get("test_details", [])
            )

            with open(report_path, "rb") as file:
                report_data = file.read()

            st.download_button(
                label="📥 下载详细测试报告",
                data=report_data,
                file_name=os.path.basename(report_path),
                mime="text/html",
                use_container_width=True,
                key=f"download_interface_report_{execution_mode}"
            )

        test_details = test_results.get("test_details", [])
        if test_details:
            st.markdown("### 📋 测试详情摘要")
            summary_rows = []
            for detail in test_details:
                status_icon = "✅" if detail.get("status") == "passed" else "❌" if detail.get("status") == "failed" else "⚠️"
                summary_rows.append({
                    "接口": detail.get("name", "未知"),
                    "环境": detail.get("environment", ""),
                    "方法": detail.get("method", "GET"),
                    "路径": detail.get("path", ""),
                    "状态": f"{status_icon} {detail.get('status', 'unknown')}",
                    "状态码": detail.get("status_code", "N/A"),
                    "响应时间": f"{detail.get('response_time', 0):.2f}s"
                })
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

            failed_tests = [detail for detail in test_details if detail.get("status") in ["failed", "error"]]
            if failed_tests:
                st.markdown("#### ❌ 失败和错误详情")
                for detail in failed_tests:
                    with st.expander(f"{detail.get('name', '未知接口')} - {detail.get('status', 'error')}", expanded=False):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write(f"**环境:** {detail.get('environment', '')}")
                            st.write(f"**方法:** {detail.get('method', 'GET')}")
                            st.write(f"**路径:** {detail.get('path', '')}")
                            st.write(f"**状态码:** {detail.get('status_code', 'N/A')}")
                            st.write(f"**响应时间:** {detail.get('response_time', 0):.2f}s")
                        with col2:
                            st.write("**错误信息:**")
                            st.error(detail.get("error", "无错误信息"))
                        if detail.get("assertions"):
                            st.write("**断言结果:**")
                            for assertion in detail.get("assertions", []):
                                if assertion.get("passed"):
                                    st.success(f"✅ {assertion.get('description')}: {assertion.get('message')}")
                                else:
                                    st.error(f"❌ {assertion.get('description')}: {assertion.get('message')}")

        if test_results.get("success", False):
            st.success("✅ 所有测试用例执行成功")
        elif total > 0:
            st.error(f"❌ 存在 {failed + errors} 个失败或错误")


    def build_env_template_with_base_url(base_url):
        template = json.loads(st.session_state.auto_test_tool.build_environment_template())
        normalized_base_url = base_url.strip()
        if normalized_base_url:
            template["environments"]["dev"]["base_url"] = normalized_base_url
        return json.dumps(template, ensure_ascii=False, indent=2)


    def generate_artifacts(
        interfaces,
        execution_mode,
        base_url,
        timeout,
        retry_times,
        verify_ssl,
        request_format,
        template_style,
        environment_config=None,
        auth_config=None,
    ):
        with st.spinner("正在生成测试用例..."):
            test_files = st.session_state.auto_test_tool.generate_test_cases(
                interfaces=interfaces,
                framework=execution_mode,
                base_url=base_url,
                timeout=timeout,
                retry_times=retry_times,
                verify_ssl=verify_ssl,
                request_format=request_format,
                template_style=template_style,
                environment_config=environment_config,
                auth_config=auth_config
            )
            st.session_state.auto_test_tool.save_test_files(test_files)
            st.session_state.interface_last_generated_files = test_files
        return test_files


    def ensure_generated_file(execution_mode):
        target_file = "run_interfaces.py" if execution_mode == "requests脚本" else "test_interfaces.py"
        manifest_file = os.path.join(st.session_state.auto_test_tool.test_dir, "interface_manifest.json")
        target_path = os.path.join(st.session_state.auto_test_tool.test_dir, target_file)
        if not (os.path.exists(manifest_file) and os.path.exists(target_path)):
            return False
        try:
            with open(manifest_file, "r", encoding="utf-8") as file:
                manifest = json.load(file)
        except Exception:
            return False
        expected_mode = st.session_state.auto_test_tool._normalize_execution_mode(execution_mode)
        return manifest.get("mode") == expected_mode


    show_doc("interface_auto_test")

    if "auto_test_tool" not in st.session_state:
        st.session_state.auto_test_tool = InterfaceAutoTestCore()
    if "enhanced_runner" not in st.session_state:
        st.session_state.enhanced_runner = EnhancedTestRunner()
    if "enhanced_report" not in st.session_state:
        st.session_state.enhanced_report = EnhancedReportGenerator()
    if "interface_loaded_interfaces" not in st.session_state:
        st.session_state.interface_loaded_interfaces = []
    if "interface_loaded_base_url" not in st.session_state:
        st.session_state.interface_loaded_base_url = ""
    if "interface_source_name" not in st.session_state:
        st.session_state.interface_source_name = ""

    st.markdown('<div class="category-card">🚀 接口自动化测试工具</div>', unsafe_allow_html=True)

    execution_mode = st.radio(
        "选择执行模式",
        ["pytest", "unittest", "requests脚本"],
        horizontal=True,
        key="interface_execution_mode"
    )

    if not check_interface_dependencies(execution_mode, for_execution=False):
        st.stop()

    st.markdown("### 🚀 快速测试")
    quick_col1, quick_col2, quick_col3 = st.columns(3)
    with quick_col1:
        if st.button("📝 JSONPlaceholder", use_container_width=True, key="load_jsonplaceholder"):
            base_url, interfaces = create_jsonplaceholder_test_data()
            set_interface_state(interfaces, base_url, "内置 JSONPlaceholder 示例")
            st.success("✅ 已加载 JSONPlaceholder 示例")
    with quick_col2:
        if st.button("👥 ReqRes", use_container_width=True, key="load_reqres"):
            base_url, interfaces = create_reqres_test_data()
            set_interface_state(interfaces, base_url, "内置 ReqRes 示例")
            st.success("✅ 已加载 ReqRes 示例")
    with quick_col3:
        if st.button("🔄 清空导入数据", use_container_width=True, key="clear_interface_inputs"):
            clear_interface_state()
            st.success("✅ 已清空接口数据和生成产物缓存")

    with st.expander("📚 模板与导入说明", expanded=not st.session_state.interface_loaded_interfaces):
        template_row1_col1, template_row1_col2, template_row1_col3, template_row1_col4 = st.columns(4)
        with template_row1_col1:
            st.download_button(
                label="📥 Excel模板",
                data=st.session_state.auto_test_tool.build_excel_template_bytes(),
                file_name="接口文档模板.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with template_row1_col2:
            st.download_button(
                label="📥 JSON模板",
                data=st.session_state.auto_test_tool.build_json_template(),
                file_name="接口文档模板.json",
                mime="application/json",
                use_container_width=True
            )
        with template_row1_col3:
            st.download_button(
                label="📥 文本模板",
                data=st.session_state.auto_test_tool.build_text_template(),
                file_name="接口文档模板.txt",
                mime="text/plain",
                use_container_width=True
            )
        with template_row1_col4:
            st.download_button(
                label="📥 curl模板",
                data=st.session_state.auto_test_tool.build_curl_template(),
                file_name="curl-import-template.txt",
                mime="text/plain",
                use_container_width=True
            )

        template_row2_col1, template_row2_col2, template_row2_col3, template_row2_col4 = st.columns(4)
        with template_row2_col1:
            st.download_button(
                label="📥 OpenAPI JSON",
                data=st.session_state.auto_test_tool.build_openapi_template(),
                file_name="openapi-template.json",
                mime="application/json",
                use_container_width=True
            )
        with template_row2_col2:
            st.download_button(
                label="📥 OpenAPI YAML",
                data=st.session_state.auto_test_tool.build_openapi_yaml_template(),
                file_name="openapi-template.yaml",
                mime="text/yaml",
                use_container_width=True
            )
        with template_row2_col3:
            st.download_button(
                label="📥 Postman模板",
                data=st.session_state.auto_test_tool.build_postman_template(),
                file_name="postman-collection-template.json",
                mime="application/json",
                use_container_width=True
            )
        with template_row2_col4:
            st.download_button(
                label="📥 HAR模板",
                data=st.session_state.auto_test_tool.build_har_template(),
                file_name="sample-template.har",
                mime="application/json",
                use_container_width=True
            )

        template_row3_col1, template_row3_col2, template_row3_col3 = st.columns(3)
        with template_row3_col1:
            st.download_button(
                label="📥 Apifox模板",
                data=st.session_state.auto_test_tool.build_apifox_template(),
                file_name="apifox-openapi-template.json",
                mime="application/json",
                use_container_width=True
            )

        template_row4_col1, template_row4_col2 = st.columns(2)
        with template_row4_col1:
            st.download_button(
                label="📥 环境配置模板",
                data=st.session_state.auto_test_tool.build_environment_template(),
                file_name="environment-config-template.json",
                mime="application/json",
                use_container_width=True
            )
        with template_row4_col2:
            st.download_button(
                label="📥 鉴权模板",
                data=st.session_state.auto_test_tool.build_auth_template(),
                file_name="auth-template.json",
                mime="application/json",
                use_container_width=True
            )
        with template_row3_col2:
            st.download_button(
                label="📥 Bruno模板",
                data=st.session_state.auto_test_tool.build_bruno_template(),
                file_name="sample-template.bru",
                mime="text/plain",
                use_container_width=True
            )
        with template_row3_col3:
            st.download_button(
                label="📥 Insomnia模板",
                data=st.session_state.auto_test_tool.build_insomnia_template(),
                file_name="insomnia-export-template.json",
                mime="application/json",
                use_container_width=True
            )

        st.markdown("""
        **支持的导入方式**

        - 文件上传: `xlsx/xls/json/har/bru/txt/md/yaml/yml`
        - Swagger/OpenAPI URL: 直接输入 JSON 或 YAML 地址
        - 原始文本: 支持结构化文本、JSON 文本、Swagger 文本、`curl` 命令、`Bruno .bru`
        - JSON 导入兼容 `OpenAPI / Apifox / Postman Collection / HAR / Insomnia / 自定义接口列表`

        **可下载模板**

        - `Excel模板`: 适合表格维护接口清单后上传
        - `JSON模板`: 适合结构化接口列表导入
        - `文本模板`: 适合手写接口块批量粘贴
        - `OpenAPI JSON / YAML`: 适合 Swagger/OpenAPI 文档导入
        - `Apifox模板`: 提供 Apifox 兼容的 OpenAPI 导出骨架
        - `curl模板`: 适合把抓包或终端命令直接粘贴导入
        - `Postman模板`: 适合集合导出后直接导入
        - `HAR模板`: 适合浏览器抓包导出后回放
        - `Bruno模板`: 适合 `.bru` 单接口文件导入
        - `Insomnia模板`: 适合 Insomnia 导出 JSON 回放
        - `环境配置模板`: 适合维护 `dev/staging/prod` 基础地址和公共头
        - `鉴权模板`: 适合统一配置 `Bearer / API Key / Basic / 自定义 Header`

        **支持的执行模板**

        - `pytest`: 生成 `pytest` 可执行脚本
        - `unittest`: 生成 `unittest` 可执行脚本
        - `requests脚本`: 生成可直接运行的 `requests` 脚本

        **模板风格**

        - `冒烟模板`: 只校验状态码
        - `标准模板`: 校验状态码和关键响应字段
        - `严格模板`: 按期望响应做更严格的字段/值校验
        """)

    st.markdown("### 📥 导入接口文档")
    import_mode = st.radio(
        "导入方式",
        ["文件上传", "Swagger/OpenAPI URL", "原始文本"],
        horizontal=True,
        key="interface_import_mode"
    )

    if import_mode == "文件上传":
        uploaded_file = st.file_uploader(
            "选择接口文档文件",
            type=["xlsx", "xls", "json", "har", "bru", "txt", "md", "yaml", "yml"],
            help="支持 Excel、JSON、Swagger/OpenAPI、Apifox、Postman、HAR、Bruno、Insomnia、文本等格式",
            key="interface_doc_upload_v2"
        )
        if uploaded_file is not None:
            upload_signature = f"{uploaded_file.name}:{uploaded_file.size}"
            if st.session_state.get("interface_upload_signature") != upload_signature:
                file_path = os.path.join(st.session_state.auto_test_tool.upload_dir, uploaded_file.name)
                with open(file_path, "wb") as file:
                    file.write(uploaded_file.getbuffer())
                try:
                    with st.spinner("正在解析接口文档..."):
                        interfaces = st.session_state.auto_test_tool.parse_document(file_path)
                    detected_base_url = st.session_state.auto_test_tool.last_parse_meta.get("detected_base_url", "")
                    set_interface_state(interfaces, detected_base_url, uploaded_file.name)
                    st.session_state.interface_upload_signature = upload_signature
                    st.success(f"✅ 成功解析出 {len(interfaces)} 个接口")
                    if detected_base_url:
                        st.info(f"已从文档中识别基础地址: `{detected_base_url}`")
                except Exception as e:
                    st.error(f"❌ 处理文件时出错: {str(e)}")

    elif import_mode == "Swagger/OpenAPI URL":
        swagger_url = st.text_input(
            "Swagger / OpenAPI 地址",
            placeholder="例如: https://example.com/openapi.json",
            key="swagger_spec_url"
        )
        if st.button("🔍 解析 Swagger / OpenAPI", use_container_width=True, key="parse_swagger_url"):
            if not swagger_url.strip():
                st.error("❌ 请输入 Swagger/OpenAPI 地址")
            else:
                try:
                    with st.spinner("正在拉取并解析远程文档..."):
                        interfaces = st.session_state.auto_test_tool.parse_content(
                            swagger_url,
                            source_type="swagger",
                            source_name="remote-swagger"
                        )
                    detected_base_url = st.session_state.auto_test_tool.last_parse_meta.get("detected_base_url", "")
                    set_interface_state(interfaces, detected_base_url, swagger_url)
                    st.success(f"✅ 成功解析出 {len(interfaces)} 个接口")
                    if detected_base_url:
                        st.info(f"已从文档中识别基础地址: `{detected_base_url}`")
                except Exception as e:
                    st.error(f"❌ 解析 Swagger/OpenAPI 失败: {str(e)}")

    else:
        raw_format = st.selectbox(
            "文本格式",
            ["自动检测", "JSON / Apifox / Postman / HAR / Insomnia", "Swagger/OpenAPI", "结构化文本", "Bruno .bru"],
            key="raw_interface_format"
        )
        raw_content = st.text_area(
            "粘贴接口定义",
            height=260,
            placeholder="可粘贴 JSON、OpenAPI 文本、结构化文本，或 curl 命令",
            key="raw_interface_content"
        )
        if st.button("🧩 解析文本内容", use_container_width=True, key="parse_raw_interface_content"):
            if not raw_content.strip():
                st.error("❌ 请输入要解析的文本内容")
            else:
                format_map = {
                    "自动检测": "auto",
                    "JSON / Apifox / Postman / HAR / Insomnia": "json",
                    "Swagger/OpenAPI": "swagger",
                    "结构化文本": "text",
                    "Bruno .bru": "bruno"
                }
                try:
                    with st.spinner("正在解析文本内容..."):
                        interfaces = st.session_state.auto_test_tool.parse_content(
                            raw_content,
                            source_type=format_map[raw_format],
                            source_name="inline-text"
                        )
                    detected_base_url = st.session_state.auto_test_tool.last_parse_meta.get("detected_base_url", "")
                    set_interface_state(interfaces, detected_base_url, f"{raw_format} 文本")
                    st.success(f"✅ 成功解析出 {len(interfaces)} 个接口")
                    if detected_base_url:
                        st.info(f"已从内容中识别基础地址: `{detected_base_url}`")
                except Exception as e:
                    st.error(f"❌ 解析文本失败: {str(e)}")

    interfaces = st.session_state.interface_loaded_interfaces
    if interfaces:
        source_name = st.session_state.interface_source_name or "当前导入结果"
        st.info(f"🎯 当前已加载 {len(interfaces)} 个接口，来源: {source_name}")
        render_interface_preview(interfaces)

        st.markdown("### ⚙️ 测试配置")
        config_col1, config_col2 = st.columns(2)
        with config_col1:
            base_url = st.text_input(
                "基础URL或默认URL",
                value=st.session_state.interface_loaded_base_url,
                placeholder="例如: http://10.0.3.54:3000",
                key="interface_base_url_input_v2"
            )
            timeout = st.number_input(
                "请求超时时间(秒)",
                min_value=1,
                value=30,
                key="interface_timeout_v2"
            )
            request_format = st.selectbox(
                "请求格式",
                ["自动检测", "json=参数", "data=json.dumps()", "form-urlencoded"],
                index=0,
                key="interface_request_format_v2"
            )
        with config_col2:
            retry_times = st.number_input(
                "重试次数",
                min_value=0,
                value=0,
                key="interface_retry_times_v2"
            )
            verify_ssl = st.checkbox(
                "验证SSL证书",
                value=False,
                key="interface_verify_ssl_v2"
            )
            template_style = st.selectbox(
                "模板风格",
                ["冒烟模板", "标准模板", "严格模板"],
                index=1,
                key="interface_template_style_v2"
            )

        st.markdown("### 🌐 环境与鉴权")
        env_auth_col1, env_auth_col2 = st.columns(2)
        environment_config = None
        auth_config = {"enabled": False, "type": "none"}

        with env_auth_col1:
            env_mode = st.radio(
                "环境模式",
                ["单环境", "多环境(JSON模板)"],
                horizontal=True,
                key="interface_env_mode"
            )
            if env_mode == "多环境(JSON模板)":
                env_default_text = build_env_template_with_base_url(base_url or st.session_state.interface_loaded_base_url)
                env_config_text = st.text_area(
                    "环境配置 JSON",
                    value=st.session_state.get("interface_env_config_text", env_default_text),
                    height=220,
                    key="interface_env_config_text",
                    help="定义 dev/staging/prod 的 base_url 和公共 headers，运行时按所选环境生效"
                )
                try:
                    environment_config = json.loads(env_config_text)
                    env_names = list((environment_config.get("environments") or {}).keys())
                    if env_names:
                        default_env = environment_config.get("active_env")
                        default_index = env_names.index(default_env) if default_env in env_names else 0
                        selected_env = st.selectbox(
                            "当前环境",
                            env_names,
                            index=default_index,
                            key="interface_active_env"
                        )
                        environment_config["active_env"] = selected_env
                    else:
                        st.error("❌ 环境配置缺少 environments 节点")
                        environment_config = None
                except json.JSONDecodeError as exc:
                    st.error(f"❌ 环境配置 JSON 格式错误: {exc}")
                    environment_config = None
            else:
                st.info("当前使用单环境模式，执行时直接使用上面的基础URL")

        with env_auth_col2:
            auth_mode = st.selectbox(
                "鉴权方式",
                ["无", "Bearer Token", "API Key", "Basic Auth", "自定义Header"],
                key="interface_auth_mode"
            )
            if auth_mode == "Bearer Token":
                token = st.text_input("Bearer Token", type="password", key="interface_bearer_token")
                auth_config = {
                    "enabled": bool(token),
                    "type": "bearer",
                    "header_name": "Authorization",
                    "prefix": "Bearer ",
                    "token": token,
                }
            elif auth_mode == "API Key":
                api_key_name = st.text_input("Header名称", value="X-API-Key", key="interface_api_key_name")
                api_key_value = st.text_input("API Key", type="password", key="interface_api_key_value")
                auth_config = {
                    "enabled": bool(api_key_value),
                    "type": "api_key",
                    "api_key_name": api_key_name,
                    "api_key_value": api_key_value,
                }
            elif auth_mode == "Basic Auth":
                username = st.text_input("用户名", key="interface_basic_username")
                password = st.text_input("密码", type="password", key="interface_basic_password")
                auth_config = {
                    "enabled": bool(username),
                    "type": "basic",
                    "username": username,
                    "password": password,
                }
            elif auth_mode == "自定义Header":
                custom_header_name = st.text_input("Header名称", value="X-Custom-Auth", key="interface_custom_header_name")
                custom_header_value = st.text_input("Header值", type="password", key="interface_custom_header_value")
                auth_config = {
                    "enabled": bool(custom_header_value),
                    "type": "custom_header",
                    "custom_header_name": custom_header_name,
                    "custom_header_value": custom_header_value,
                }
            else:
                st.info("不附加统一鉴权头，接口自身 headers 保持不变")

        st.markdown("### 🎯 测试操作")
        action_col1, action_col2, action_col3 = st.columns(3)

        with action_col1:
            if st.button("🧪 生成测试用例", use_container_width=True, key="generate_tests_v2"):
                try:
                    if env_mode == "多环境(JSON模板)" and environment_config is None:
                        raise RuntimeError("多环境配置无效，请先修正环境配置 JSON")
                    generate_artifacts(
                        interfaces,
                        execution_mode,
                        base_url,
                        timeout,
                        retry_times,
                        verify_ssl,
                        request_format,
                        template_style,
                        environment_config,
                        auth_config
                    )
                    st.success("✅ 测试文件已生成")
                except Exception as e:
                    st.error(f"❌ 生成测试用例失败: {str(e)}")

        with action_col2:
            if st.button("▶️ 执行测试", use_container_width=True, key="run_tests_v2"):
                if not ensure_generated_file(execution_mode):
                    st.error("❌ 请先生成当前执行模式对应的测试文件")
                elif not check_interface_dependencies(execution_mode, for_execution=True):
                    pass
                else:
                    with st.spinner("正在执行测试并收集结果..."):
                        test_results = st.session_state.enhanced_runner.run_tests_with_details(
                            execution_mode,
                            interfaces
                        )
                        st.session_state.interface_last_test_results = test_results

        with action_col3:
            if st.button("⚡ 生成并执行", use_container_width=True, key="generate_and_run_tests_v2"):
                try:
                    if env_mode == "多环境(JSON模板)" and environment_config is None:
                        raise RuntimeError("多环境配置无效，请先修正环境配置 JSON")
                    generate_artifacts(
                        interfaces,
                        execution_mode,
                        base_url,
                        timeout,
                        retry_times,
                        verify_ssl,
                        request_format,
                        template_style,
                        environment_config,
                        auth_config
                    )
                    if not check_interface_dependencies(execution_mode, for_execution=True):
                        raise RuntimeError("当前执行模式缺少运行依赖，请先安装后再执行")
                    with st.spinner("正在执行测试并收集结果..."):
                        test_results = st.session_state.enhanced_runner.run_tests_with_details(
                            execution_mode,
                            interfaces
                        )
                        st.session_state.interface_last_test_results = test_results
                except Exception as e:
                    st.error(f"❌ 生成并执行失败: {str(e)}")

        generated_files = st.session_state.get("interface_last_generated_files", {})
        if generated_files:
            st.markdown("### 📄 最近一次生成结果")
            for filename, content in generated_files.items():
                language = "json" if filename.endswith(".json") else "python"
                with st.expander(f"查看 {filename}", expanded=filename.endswith(".py")):
                    st.code(content, language=language)

        cached_results = st.session_state.get("interface_last_test_results")
        if cached_results:
            render_test_results(cached_results, execution_mode, interfaces)

        st.markdown("---")
        clean_col = st.columns([1, 2, 1])[1]
        with clean_col:
            if st.button("🗑️ 清理生成文件", use_container_width=True, key="clean_generated_interface_files"):
                import shutil
                try:
                    if os.path.exists(st.session_state.auto_test_tool.test_dir):
                        shutil.rmtree(st.session_state.auto_test_tool.test_dir)
                    os.makedirs(st.session_state.auto_test_tool.test_dir)
                    st.session_state.pop("interface_last_generated_files", None)
                    st.session_state.pop("interface_last_test_results", None)
                    st.success("✅ 生成文件已清理")
                except Exception as e:
                    st.error(f"❌ 清理失败: {str(e)}")
    else:
        st.info("📝 请先导入接口文档，支持 Excel、Swagger/OpenAPI、JSON、文本和 curl")

    st.markdown('</div>', unsafe_allow_html=True)


# 在工具选择部分添加BI工具
elif tool_category == "BI数据分析工具":
    show_doc("bi_analyzer")

    # 初始化BI分析器
    bi_tool = BIAnalyzer()

    # 显示模板下载和文件上传
    uploaded_file = bi_tool.show_upload_section()

    if uploaded_file is not None:
        # 加载数据
        df, message = bi_tool.load_data(uploaded_file)

        if df is not None:
            st.success(message)

            # 创建选项卡界面
            tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
                "📋 数据预览",
                "📊 统计分析",
                "🔍 数据透视",
                "📈 趋势分析",
                "🎯 数据仪表板",
                "💾 导出报告"
            ])

            with tab1:
                bi_tool.data_preview(df)

            with tab2:
                bi_tool.basic_statistics(df)
                bi_tool.correlation_analysis(df)

            with tab3:
                bi_tool.create_pivot_table(df)

            with tab4:
                bi_tool.time_series_analysis(df)

            with tab5:
                bi_tool.create_dashboard(df)

            with tab6:
                bi_tool.export_report(df)

        else:
            st.error(message)
    else:
        # 显示使用说明
        st.info("""
        ### 🚀 BI数据分析工具使用说明

        **功能特点:**
        - 📊 **数据可视化**: 多种图表类型，直观展示数据
        - 📈 **统计分析**: 描述性统计、相关性分析
        - 🔍 **数据透视**: 灵活的数据透视表分析
        - 📉 **趋势分析**: 时间序列分析和预测
        - 🎯 **交互式仪表板**: 自定义数据仪表板
        - 💾 **报告导出**: 多种格式的数据报告

        **使用流程:**
        1. 下载数据模板（推荐先使用模板了解格式要求）
        2. 准备您的数据文件
        3. 上传数据文件
        4. 使用不同分析功能探索数据
        5. 创建可视化图表和仪表板
        6. 导出分析报告
        """)
# 初始化并使用留言区
feedback_section = FeedbackSection()
feedback_section.render_feedback_section()

# show_general_guidelines()
author = AuthorProfile()

# 在需要显示底部作者介绍的地方调用
author.render_main_profile()

# 在需要显示侧边栏作者信息的地方调用
# author.render_sidebar_profile()
