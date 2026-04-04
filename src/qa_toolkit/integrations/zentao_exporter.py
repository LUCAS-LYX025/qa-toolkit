import streamlit as st
import pymysql
import pandas as pd
from string import Template
from typing import Dict, List, Optional, Tuple


class MysqlDB(object):
    def __init__(self, host, port, user, passwd, db):
        """
        初始化数据库连接。

        参数:
        host: 数据库主机地址。
        port: 数据库监听端口。
        user: 连接数据库的用户名。
        passwd: 用户密码。
        db: 要连接的数据库名。
        """
        # 创建数据库连接
        self.conn = pymysql.connect(
            host=host,
            port=port,
            user=user,
            passwd=passwd,
            db=db,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor  # 使用字典游标方便处理
        )
        # 创建游标
        self.cur = self.conn.cursor()

    def __del__(self):
        """
        析构函数，用于在实例被删除时自动触发，确保资源被正确释放。
        """
        if hasattr(self, 'cur'):
            self.cur.close()  # 关闭当前游标
        if hasattr(self, 'conn'):
            self.conn.close()  # 关闭数据库连接

    def sql_query(self, sql_str):
        """
        执行SQL查询

        参数:
        - sql_str: 字符串，要执行的SQL查询语句

        返回值:
        - self.cur: 执行查询后的cursor对象
        """
        self.cur.execute(sql_str)
        return self.cur

    def fetchall_to_dict(self):
        """将查询结果转换为字典列表"""
        return self.cur.fetchall()

    def fetchone_to_dict(self):
        """获取单条查询结果"""
        return self.cur.fetchone()

    def robot_solve(self, solve_dic, data_dic):
        """
        使用给定的解决字典和数据字典，通过查询SQL来解决某些问题并将结果收集到一个新的字典中。

        :param solve_dic: 包含SQL模板字符串的字典，键是问题标识符，值是SQL查询模板。
        :param data_dic: 包含用于替换SQL模板字符串中占位符的数据的字典。
        :return: 一个字典，其中键是solve_dic的键，值是对应SQL查询结果的第0个元素。
        """
        dic = dict.fromkeys(list(solve_dic.keys()))

        for key in dic:
            self.sql_query(Template(solve_dic[key]).safe_substitute(data_dic))
            result = self.fetchone_to_dict()
            dic[key] = list(result.values())[0] if result else None

        return dic

    def report_solve(self, solve_dic, data_dic):
        """
        根据提供的解决字典和数据字典，生成报告解决方案。

        :param solve_dic: 包含模板字符串键值对的字典，其中键表示报告的字段，值是包含占位符的SQL查询模板。
        :param data_dic: 包含用于替换solve_dic模板中占位符的数据的字典。
        :return: 一个字典，其中键是solve_dic的键，值是通过将data_dic数据代入solve_dic相应模板并执行SQL查询得到的结果。
        """
        dic = dict.fromkeys(list(solve_dic.keys()))

        for key in dic:
            self.sql_query(Template(solve_dic[key]).safe_substitute(data_dic))
            dic[key] = self.fetchall_to_dict()

        return dic

    def execute_query_to_dataframe(self, sql_str):
        """
        执行SQL查询并返回pandas DataFrame

        参数:
        - sql_str: SQL查询语句

        返回值:
        - pd.DataFrame: 查询结果的数据框
        """
        self.sql_query(sql_str)
        results = self.fetchall_to_dict()
        if results:
            return pd.DataFrame(results)
        else:
            return pd.DataFrame()


class ZenTaoPerformanceExporter:
    def __init__(self, db_config):
        self.db_config = db_config
        self.mysql_db = self.create_mysql_db()

    def create_mysql_db(self):
        """创建 MysqlDB 实例"""
        try:
            return MysqlDB(
                host=self.db_config['host'],
                port=self.db_config['port'],
                user=self.db_config['user'],
                passwd=self.db_config['password'],
                db=self.db_config['database']
            )
        except Exception as e:
            st.error(f"创建数据库连接失败: {str(e)}")
            return None

    def get_products(self):
        """获取所有产品列表"""
        try:
            query = "SELECT id, name FROM zt_product WHERE deleted = '0' ORDER BY name"
            self.mysql_db.sql_query(query)
            results = self.mysql_db.fetchall_to_dict()
            return [(item['id'], item['name']) for item in results]
        except Exception as e:
            st.error(f"获取产品列表失败: {str(e)}")
            return []

    def get_user_roles(self):
        """获取所有用户角色（显示中文）"""
        try:
            # 先从zt_lang表获取角色中文名称
            role_query = """
            SELECT DISTINCT u.role as role_key, 
                   COALESCE(l.value, u.role) as role_name
            FROM zt_user u
            LEFT JOIN zt_lang l ON l.module = 'user' 
                AND l.section = 'roleList' 
                AND l.`key` = u.role 
                AND l.lang = 'zh-cn'
            WHERE u.role IS NOT NULL AND u.role != ''
            ORDER BY role_name
            """
            self.mysql_db.sql_query(role_query)
            results = self.mysql_db.fetchall_to_dict()

            # 返回角色键值对
            return [(item['role_key'], item['role_name']) for item in results]
        except Exception as e:
            st.error(f"获取用户角色失败: {str(e)}")
            # 如果获取中文失败，返回原始角色
            try:
                query = "SELECT DISTINCT role FROM zt_user WHERE role IS NOT NULL AND role != '' ORDER BY role"
                self.mysql_db.sql_query(query)
                results = self.mysql_db.fetchall_to_dict()
                return [(item['role'], item['role']) for item in results]
            except:
                return []

    def get_bug_types(self):
        """获取所有缺陷类型（显示中文）"""
        try:
            # 从zt_lang表获取缺陷类型中文名称
            type_query = """
              SELECT DISTINCT b.type as type_key,
                     COALESCE(l.value, b.type) as type_name
              FROM zt_bug b
              LEFT JOIN zt_lang l ON l.module = 'bug' 
                  AND l.section = 'typeList' 
                  AND l.`key` = b.type 
                  AND l.lang = 'zh-cn'
              WHERE b.type IS NOT NULL AND b.type != ''
              ORDER BY type_name
              """
            self.mysql_db.sql_query(type_query)
            results = self.mysql_db.fetchall_to_dict()

            # 返回类型键值对
            return [(item['type_key'], item['type_name']) for item in results]
        except Exception as e:
            st.error(f"获取缺陷类型失败: {str(e)}")
            # 如果获取中文失败，返回原始类型
            try:
                query = "SELECT DISTINCT type FROM zt_bug WHERE type IS NOT NULL AND type != '' ORDER BY type"
                self.mysql_db.sql_query(query)
                results = self.mysql_db.fetchall_to_dict()
                return [(item['type'], item['type']) for item in results]
            except:
                return []

    def get_bug_severities(self):
        """获取缺陷严重程度（显示中文）"""
        try:
            severity_query = """
            SELECT DISTINCT l.`key` as severity_key, 
                   l.value as severity_name
            FROM zt_lang l
            WHERE l.module = 'bug' 
                AND l.section = 'severityList' 
                AND l.lang = 'zh-cn'
            ORDER BY l.`key`
            """
            self.mysql_db.sql_query(severity_query)
            results = self.mysql_db.fetchall_to_dict()
            return [(item['severity_key'], item['severity_name']) for item in results]
        except Exception as e:
            st.error(f"获取缺陷严重程度失败: {str(e)}")
            return []

    def get_bug_priorities(self):
        """获取缺陷优先级（显示中文）"""
        try:
            priority_query = """
            SELECT DISTINCT l.`key` as priority_key, 
                   l.value as priority_name
            FROM zt_lang l
            WHERE l.module = 'bug' 
                AND l.section = 'priList' 
                AND l.lang = 'zh-cn'
            ORDER BY l.`key`
            """
            self.mysql_db.sql_query(priority_query)
            results = self.mysql_db.fetchall_to_dict()
            return [(item['priority_key'], item['priority_name']) for item in results]
        except Exception as e:
            st.error(f"获取缺陷优先级失败: {str(e)}")
            return []

    def build_qa_query(self, product_id: int, start_date: str, end_date: str, config: Dict) -> str:
        """构建测试人员SQL查询语句"""
        exclude_types_str = ", ".join(f"'{t}'" for t in config['exclude_types'])
        roles_str = ", ".join(f"'{r}'" for r in config['roles'])

        # 使用配置的超时参数
        high_priority_normal_hours = config['high_priority_normal_hours']
        high_priority_weekend_hours = config['high_priority_weekend_hours']
        normal_priority_normal_hours = config['normal_priority_normal_hours']
        normal_priority_weekend_hours = config['normal_priority_weekend_hours']

        return f"""
        SELECT 
            COALESCE(u.realname, combined_data.tester) AS 测试人员,
            COALESCE(l.value, u.role) AS 测试角色,
            SUM(submitted_bugs) AS 提交bug数量,
            SUM(CASE WHEN is_high_priority = 1 THEN bug_count ELSE 0 END) AS 一级超时响应次数,
            SUM(CASE WHEN is_high_priority = 0 THEN bug_count ELSE 0 END) AS 普通超时响应次数,
            SUM(CASE WHEN is_high_priority = 1 THEN bug_count ELSE 0 END) + 
            SUM(CASE WHEN is_high_priority = 0 THEN bug_count ELSE 0 END) AS 总超时响应次数,
            CONCAT(
                ROUND(
                    (SUM(CASE WHEN is_high_priority = 1 THEN bug_count ELSE 0 END) + 
                     SUM(CASE WHEN is_high_priority = 0 THEN bug_count ELSE 0 END)) / 
                    GREATEST(SUM(submitted_bugs), 1) * 100, 
                    2
                ), 
                '%'
            ) AS 超时响应率
        FROM (
            -- 已关闭的bug统计
            SELECT 
                closedBy AS tester,
                (severity = 1 OR pri = 1) AS is_high_priority,
                COUNT(*) AS bug_count,
                0 AS submitted_bugs
            FROM 
                zt_bug
            WHERE 
                product = {product_id}
                AND status = 'closed'
                AND deleted = '0'
                AND closedBy IS NOT NULL
                AND closedBy = openedBy
                AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                AND type NOT IN ({exclude_types_str})
                AND (
                    -- 高优先级：考虑周末顺延
                    ((severity = 1 OR pri = 1) AND 
                        CASE 
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN TIMESTAMPDIFF(HOUR, resolvedDate, closedDate) > {high_priority_weekend_hours}
                            ELSE TIMESTAMPDIFF(HOUR, resolvedDate, closedDate) > {high_priority_normal_hours}
                        END)
                    OR
                    -- 普通优先级：考虑周末顺延
                    ((severity != 1 AND pri != 1) AND 
                        CASE 
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN TIMESTAMPDIFF(HOUR, resolvedDate, closedDate) > {normal_priority_weekend_hours}
                            ELSE TIMESTAMPDIFF(HOUR, resolvedDate, closedDate) > {normal_priority_normal_hours}
                        END)
                )
            GROUP BY 
                closedBy, is_high_priority

            UNION ALL

            -- 已解决但未关闭的bug统计
            SELECT 
                assignedTo AS tester,
                (severity = 1 OR pri = 1) AS is_high_priority,
                COUNT(*) AS bug_count,
                0 AS submitted_bugs
            FROM 
                zt_bug
            WHERE 
                product = {product_id}
                AND status = 'resolved'
                AND deleted = '0'
                AND assignedTo IS NOT NULL
                AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                AND type NOT IN ({exclude_types_str})
                AND (
                    -- 高优先级：考虑周末顺延
                    ((severity = 1 OR pri = 1) AND 
                        CASE 
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN DATE_ADD(resolvedDate, INTERVAL {high_priority_weekend_hours} HOUR) < NOW()
                            ELSE DATE_ADD(resolvedDate, INTERVAL {high_priority_normal_hours} HOUR) < NOW()
                        END)
                    OR
                    -- 普通优先级：考虑周末顺延
                    ((severity != 1 AND pri != 1) AND 
                        CASE 
                            WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                            THEN DATE_ADD(resolvedDate, INTERVAL {normal_priority_weekend_hours} HOUR) < NOW()
                            ELSE DATE_ADD(resolvedDate, INTERVAL {normal_priority_normal_hours} HOUR) < NOW()
                        END)
                )
            GROUP BY 
                assignedTo, is_high_priority

            UNION ALL

            -- 每个人员提交的bug数量统计
            SELECT 
                openedBy AS tester,
                0 AS is_high_priority,
                0 AS bug_count,
                COUNT(*) AS submitted_bugs
            FROM 
                zt_bug
            WHERE 
                product = {product_id}
                AND deleted = '0'
                AND openedDate BETWEEN '{start_date}' AND '{end_date}'
            GROUP BY 
                openedBy
        ) AS combined_data
        LEFT JOIN zt_user u ON combined_data.tester = u.account
        LEFT JOIN zt_lang l ON l.module = 'user' 
            AND l.section = 'role' 
            AND l.`key` = u.role 
            AND l.lang = 'zh-cn'
        WHERE 
            combined_data.tester IS NOT NULL
            AND u.role IN ({roles_str})
        GROUP BY 
            combined_data.tester, u.realname, u.role
        HAVING 
            SUM(submitted_bugs) > 0
        ORDER BY 
            提交bug数量 DESC;
        """

    def build_dev_query(self, product_id: int, start_date: str, end_date: str, config: Dict) -> str:
        """构建开发人员SQL查询语句"""
        exclude_types_str = ", ".join(f"'{t}'" for t in config['exclude_types'])
        roles_str = ", ".join(f"'{r}'" for r in config['roles'])

        # 使用配置的超时参数
        high_priority_normal_hours = config['high_priority_normal_hours']
        high_priority_weekend_hours = config['high_priority_weekend_hours']
        normal_priority_normal_hours = config['normal_priority_normal_hours']
        normal_priority_weekend_hours = config['normal_priority_weekend_hours']

        return f"""
         SELECT 
             CASE 
                 WHEN u.role = 'qdev' THEN '前端开发'
                 WHEN u.role = 'hdev' THEN '后端开发'
                 WHEN u.role = 'adev' THEN 'App开发'
                 ELSE COALESCE(l.value, u.role)
             END AS 开发角色,
             u.realname AS 开发人员,
             COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 1 THEN combined_data.bug_count ELSE 0 END), 0) AS 一级超时响应次数,
             COALESCE(SUM(CASE WHEN combined_data.is_high_priority = 0 THEN combined_data.bug_count ELSE 0 END), 0) AS 普通超时响应次数,
             COALESCE(SUM(combined_data.bug_count), 0) AS 总超时响应次数,
             COALESCE(total_bugs.bug_total, 0) AS 开发总bug数,
             CASE 
                 WHEN COALESCE(total_bugs.bug_total, 0) = 0 THEN '0%'
                 ELSE CONCAT(ROUND(COALESCE(SUM(combined_data.bug_count), 0) / COALESCE(total_bugs.bug_total, 1) * 100, 2), '%')
             END AS 超时响应率
         FROM 
             zt_user u
         LEFT JOIN zt_lang l ON l.module = 'user' 
             AND l.section = 'role' 
             AND l.`key` = u.role 
             AND l.lang = 'zh-cn'
         LEFT JOIN (
             -- 已解决的bug统计（按resolvedBy）
             SELECT 
                 resolvedBy AS developer,
                 (severity = 1 OR pri = 1) AS is_high_priority,
                 COUNT(*) AS bug_count
             FROM 
                 zt_bug
             WHERE 
                 product = {product_id}
                 AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                 AND resolvedDate IS NOT NULL
                 AND deleted = '0'
                 AND type NOT IN ({exclude_types_str})
                 AND (
                     -- 高优先级：考虑周末顺延
                     ((severity = 1 OR pri = 1) AND 
                         CASE 
                             WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                             THEN TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > {high_priority_weekend_hours}
                             ELSE TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > {high_priority_normal_hours}
                         END)
                     OR
                     -- 普通优先级：考虑周末顺延
                     ((severity != 1 AND pri != 1) AND 
                         CASE 
                             WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                             THEN TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > {normal_priority_weekend_hours}
                             ELSE TIMESTAMPDIFF(HOUR, openedDate, resolvedDate) > {normal_priority_normal_hours}
                         END)
                 )
             GROUP BY 
                 resolvedBy, is_high_priority

             UNION ALL

             -- 未解决的bug统计（按assignedTo）
             SELECT 
                 assignedTo AS developer,
                 (severity = 1 OR pri = 1) AS is_high_priority,
                 COUNT(*) AS bug_count
             FROM 
                 zt_bug
             WHERE 
                 product = {product_id}
                 AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                 AND status = 'active'
                 AND deleted = '0'
                 AND type NOT IN ({exclude_types_str})
                 AND (
                     -- 高优先级：考虑周末顺延
                     ((severity = 1 OR pri = 1) AND 
                         CASE 
                             WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                             THEN DATE_ADD(openedDate, INTERVAL {high_priority_weekend_hours} HOUR) < NOW()
                             ELSE DATE_ADD(openedDate, INTERVAL {high_priority_normal_hours} HOUR) < NOW()
                         END)
                     OR
                     -- 普通优先级：考虑周末顺延
                     ((severity != 1 AND pri != 1) AND 
                         CASE 
                             WHEN DAYOFWEEK(openedDate) IN (1, 6, 7) 
                             THEN DATE_ADD(openedDate, INTERVAL {normal_priority_weekend_hours} HOUR) < NOW()
                             ELSE DATE_ADD(openedDate, INTERVAL {normal_priority_normal_hours} HOUR) < NOW()
                         END)
                 )
             GROUP BY 
                 assignedTo, is_high_priority
         ) AS combined_data ON u.account = combined_data.developer
         LEFT JOIN (
             -- 计算每个开发人员的总bug数
             SELECT 
                 developer,
                 SUM(bug_count) AS bug_total
             FROM (
                 SELECT 
                     resolvedBy AS developer,
                     COUNT(*) AS bug_count
                 FROM 
                     zt_bug
                 WHERE 
                     product = {product_id}
                     AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                     AND deleted = '0'
                     AND type NOT IN ({exclude_types_str})
                 GROUP BY 
                     resolvedBy

                 UNION ALL

                 SELECT 
                     assignedTo AS developer,
                     COUNT(*) AS bug_count
                 FROM 
                     zt_bug
                 WHERE 
                     product = {product_id}
                     AND openedDate BETWEEN '{start_date}' AND '{end_date}'
                     AND status = 'active'
                     AND deleted = '0'
                     AND type NOT IN ({exclude_types_str})
                 GROUP BY 
                     assignedTo
             ) AS all_bugs
             GROUP BY 
                 developer
         ) AS total_bugs ON u.account = total_bugs.developer
         WHERE 
             u.role IN ({roles_str})
             AND COALESCE(total_bugs.bug_total, 0) > 0
         GROUP BY 
             开发角色, u.realname, u.account, total_bugs.bug_total
         ORDER BY 
             开发角色,
             COALESCE(SUM(combined_data.bug_count), 0) DESC;
         """

    def build_qa_detail_query(self, product_id: int, start_date: str, end_date: str, config: Dict) -> str:
        """构建测试人员绩效明细查询语句"""
        exclude_types_str = ", ".join(f"'{t}'" for t in config['exclude_types'])
        roles_str = ", ".join(f"'{r}'" for r in config['roles'])

        # 使用配置的超时参数
        high_priority_normal_hours = config['high_priority_normal_hours']
        high_priority_weekend_hours = config['high_priority_weekend_hours']
        normal_priority_normal_hours = config['normal_priority_normal_hours']
        normal_priority_weekend_hours = config['normal_priority_weekend_hours']

        return f"""
        SELECT 
            u.realname AS 测试人员,
            COALESCE(l.value, u.role) AS 测试角色,
            b.id AS BugID,
            b.title AS Bug标题,
            b.severity AS 严重程度,
            b.pri AS 优先级,
            CASE 
                WHEN (b.severity = 1 OR b.pri = 1) THEN '一级超时'
                ELSE '普通超时'
            END AS 超时类型,
            b.openedDate AS 创建时间,
            b.resolvedDate AS 解决时间,
            b.closedDate AS 关闭时间,
            TIMESTAMPDIFF(HOUR, b.openedDate, COALESCE(b.resolvedDate, NOW())) AS 处理时长_小时,
            CASE 
                WHEN (b.severity = 1 OR b.pri = 1) THEN
                    CASE 
                        WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {high_priority_weekend_hours}
                        ELSE {high_priority_normal_hours}
                    END
                ELSE
                    CASE 
                        WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {normal_priority_weekend_hours}
                        ELSE {normal_priority_normal_hours}
                    END
            END AS 超时阈值_小时,
            b.status AS 状态,
            b.type AS Bug类型,
            p.name AS 产品名称
        FROM 
            zt_user u
        JOIN 
            zt_bug b ON u.account = b.openedBy
        JOIN 
            zt_product p ON b.product = p.id
        LEFT JOIN zt_lang l ON l.module = 'user' 
            AND l.section = 'role' 
            AND l.`key` = u.role 
            AND l.lang = 'zh-cn'
        WHERE 
            b.product = {product_id}
            AND b.openedDate BETWEEN '{start_date}' AND '{end_date}'
            AND b.deleted = '0'
            AND u.role IN ({roles_str})
            AND b.type NOT IN ({exclude_types_str})
        ORDER BY 
            u.realname, b.openedDate DESC;
        """

    def query_qa_stats(self, product_id: int, config: Dict) -> Optional[pd.DataFrame]:
        """查询测试人员统计数据"""
        try:
            query = self.build_qa_query(
                product_id=product_id,
                start_date=config['start_date'],
                end_date=config['end_date'],
                config=config
            )
            return self.mysql_db.execute_query_to_dataframe(query)
        except Exception as e:
            st.error(f"查询测试统计数据时出错: {str(e)}")
            return None

    def query_dev_stats(self, product_id: int, config: Dict) -> Optional[pd.DataFrame]:
        """查询开发人员统计数据"""
        try:
            query = self.build_dev_query(
                product_id=product_id,
                start_date=config['start_date'],
                end_date=config['end_date'],
                config=config
            )
            return self.mysql_db.execute_query_to_dataframe(query)
        except Exception as e:
            st.error(f"查询开发统计数据时出错: {str(e)}")
            return None

    def query_qa_detail_stats(self, product_id: int, config: Dict) -> Optional[pd.DataFrame]:
        """查询测试人员绩效明细数据"""
        try:
            query = self.build_qa_detail_query(
                product_id=product_id,
                start_date=config['start_date'],
                end_date=config['end_date'],
                config=config
            )
            return self.mysql_db.execute_query_to_dataframe(query)
        except Exception as e:
            st.error(f"查询测试绩效明细时出错: {str(e)}")
            return None

    def close_connection(self):
        """关闭数据库连接"""
        if hasattr(self, 'mysql_db'):
            del self.mysql_db

    def _clean_sheet_name(self, name):
        """清理sheet名称，确保符合Excel要求"""
        # Excel sheet名称限制：最大31字符，不能包含特殊字符
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        clean_name = name
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '')
        return clean_name[:31]

    def _format_excel_worksheet(self, worksheet, df, title):
        """格式化Excel工作表"""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill

            # 设置标题
            worksheet.cell(row=1, column=1, value=title)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

            # 标题样式
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            for col_idx, column in enumerate(df.columns, 1):
                max_length = max(
                    df[column].astype(str).str.len().max() if len(df) > 0 else 0,
                    len(str(column))
                )
                col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 设置内容对齐
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, max_col=len(df.columns)):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 冻结窗格
            worksheet.freeze_panes = 'A3'

        except Exception as e:
            # 如果格式化失败，不影响主要功能
            print(f"Excel格式化失败: {e}")

    def query_timeout_bugs_detail(self, developer_name, product_id, start_date, end_date, config):
        """
        查询指定开发人员在指定产品和时间范围内的超时Bug明细
        """
        try:
            # 使用配置的超时参数
            high_priority_normal_hours = config['high_priority_normal_hours']
            high_priority_weekend_hours = config['high_priority_weekend_hours']
            normal_priority_normal_hours = config['normal_priority_normal_hours']
            normal_priority_weekend_hours = config['normal_priority_weekend_hours']

            # SQL查询语句
            sql_query = f"""
            SELECT
                u.realname AS 开发人员,
                b.id AS BugID,
                b.title AS Bug标题,
                b.severity AS 严重程度,
                b.pri AS 优先级,
                CASE
                    WHEN (b.severity = 1 OR b.pri = 1) THEN '一级超时'
                    ELSE '普通超时'
                END AS 超时类型,
                b.openedDate AS 创建时间,
                b.resolvedDate AS 解决时间,
                TIMESTAMPDIFF(HOUR, b.openedDate, COALESCE(b.resolvedDate, NOW())) AS 处理时长_小时,
                CASE
                    WHEN (b.severity = 1 OR b.pri = 1) THEN
                        CASE
                            WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {high_priority_weekend_hours}
                            ELSE {high_priority_normal_hours}
                        END
                    ELSE
                        CASE
                            WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {normal_priority_weekend_hours}
                            ELSE {normal_priority_normal_hours}
                        END
                END AS 超时阈值_小时,
                CASE
                    WHEN b.resolvedDate IS NOT NULL THEN '已解决'
                    WHEN b.status = 'active' THEN '未解决'
                    ELSE b.status
                END AS 状态,
                b.type AS Bug类型,
                p.name AS 产品名称
            FROM
                zt_user u
            JOIN
                zt_bug b ON (u.account = b.resolvedBy OR u.account = b.assignedTo)
            JOIN
                zt_product p ON b.product = p.id
            WHERE
                u.realname = %s
                AND b.product = %s
                AND b.openedDate BETWEEN %s AND %s
                AND b.deleted = '0'
                AND (
                    -- 一级优先级超时条件
                    ((b.severity = 1 OR b.pri = 1) AND
                        (
                            -- 已解决的一级超时bug
                            (b.resolvedDate IS NOT NULL AND
                             TIMESTAMPDIFF(HOUR, b.openedDate, b.resolvedDate) >
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {high_priority_weekend_hours}
                                    ELSE {high_priority_normal_hours}
                                END)
                            OR
                            -- 未解决的一级超时bug
                            (b.status = 'active' AND
                             NOW() >
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN DATE_ADD(b.openedDate, INTERVAL {high_priority_weekend_hours} HOUR)
                                    ELSE DATE_ADD(b.openedDate, INTERVAL {high_priority_normal_hours} HOUR)
                                END)
                        ))
                    OR
                    -- 普通优先级超时条件
                    ((b.severity != 1 AND b.pri != 1) AND
                        (
                            -- 已解决的普通超时bug
                            (b.resolvedDate IS NOT NULL AND
                             TIMESTAMPDIFF(HOUR, b.openedDate, b.resolvedDate) >
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {normal_priority_weekend_hours}
                                    ELSE {normal_priority_normal_hours}
                                END)
                            OR
                            -- 未解决的普通超时bug
                            (b.status = 'active' AND
                             NOW() >
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN DATE_ADD(b.openedDate, INTERVAL {normal_priority_weekend_hours} HOUR)
                                    ELSE DATE_ADD(b.openedDate, INTERVAL {normal_priority_normal_hours} HOUR)
                                END)
                        ))
                )
            ORDER BY
                CASE WHEN (b.severity = 1 OR b.pri = 1) THEN 1 ELSE 2 END,
                b.openedDate DESC;
            """

            # 正确的方式：使用 execute 方法并传递参数
            self.mysql_db.cur.execute(sql_query, (developer_name, product_id, start_date, end_date))
            results = self.mysql_db.fetchall_to_dict()

            # 转换为DataFrame
            if results:
                df = pd.DataFrame(results)
                return df
            else:
                return pd.DataFrame()

        except Exception as e:
            st.error(f"查询开发超时明细失败: {str(e)}")
            return None

    def query_qa_timeout_bugs_detail(self, tester_name, product_id, start_date, end_date, config):
        """
        查询指定测试人员在指定产品和时间范围内的超时Bug明细
        基于你提供的测试绩效明细查询逻辑
        """
        try:
            # 使用配置的超时参数
            high_priority_normal_hours = config['high_priority_normal_hours']
            high_priority_weekend_hours = config['high_priority_weekend_hours']
            normal_priority_normal_hours = config['normal_priority_normal_hours']
            normal_priority_weekend_hours = config['normal_priority_weekend_hours']

            exclude_types_str = ", ".join(f"'{t}'" for t in config['exclude_types'])

            # 基于你提供的测试绩效明细查询SQL
            sql_query = f"""
            SELECT
                u.realname AS 测试人员,
                b.id AS BugID,
                b.title AS Bug标题,
                b.severity AS 严重程度,
                b.pri AS 优先级,
                CASE
                    WHEN (b.severity = 1 OR b.pri = 1) THEN '一级超时'
                    ELSE '普通超时'
                END AS 超时类型,
                b.openedDate AS 创建时间,
                b.openedBy AS 创建人,
                b.resolvedDate AS 解决时间,
                b.resolvedBy AS 解决人,
                b.closedDate AS 关闭时间,
                b.closedBy AS 关闭人,
                b.status AS 状态,
                b.type AS Bug类型,

                -- 精确的时间计算
                CASE
                    WHEN b.status = 'closed' AND b.closedBy = b.openedBy THEN 
                        TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate)
                    WHEN b.status = 'resolved' THEN 
                        TIMESTAMPDIFF(HOUR, b.resolvedDate, NOW())
                    ELSE NULL
                END AS 响应处理时长_小时,

                -- 精确的超时阈值计算（与统计逻辑完全一致）
                CASE
                    WHEN (b.severity = 1 OR b.pri = 1) THEN
                        CASE
                            WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {high_priority_weekend_hours}
                            ELSE {high_priority_normal_hours}
                        END
                    ELSE
                        CASE
                            WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN {normal_priority_weekend_hours}
                            ELSE {normal_priority_normal_hours}
                        END
                END AS 超时阈值_小时,

                -- 精确的超时判断（与统计逻辑完全一致）
                CASE
                    -- 已关闭的Bug超时判断
                    WHEN b.status = 'closed' AND b.closedBy = b.openedBy THEN
                        CASE
                            WHEN (b.severity = 1 OR b.pri = 1) THEN
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) AND 
                                         TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {high_priority_weekend_hours} 
                                        THEN 1
                                    WHEN DAYOFWEEK(b.openedDate) NOT IN (1, 6, 7) AND 
                                         TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {high_priority_normal_hours} 
                                        THEN 1
                                    ELSE 0
                                END
                            ELSE
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) AND 
                                         TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {normal_priority_weekend_hours} 
                                        THEN 1
                                    WHEN DAYOFWEEK(b.openedDate) NOT IN (1, 6, 7) AND 
                                         TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {normal_priority_normal_hours} 
                                        THEN 1
                                    ELSE 0
                                END
                        END
                    -- 已解决但未关闭的Bug超时判断
                    WHEN b.status = 'resolved' THEN
                        CASE
                            WHEN (b.severity = 1 OR b.pri = 1) THEN
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) AND 
                                         DATE_ADD(b.resolvedDate, INTERVAL {high_priority_weekend_hours} HOUR) < NOW()
                                        THEN 1
                                    WHEN DAYOFWEEK(b.openedDate) NOT IN (1, 6, 7) AND 
                                         DATE_ADD(b.resolvedDate, INTERVAL {high_priority_normal_hours} HOUR) < NOW()
                                        THEN 1
                                    ELSE 0
                                END
                            ELSE
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) AND 
                                         DATE_ADD(b.resolvedDate, INTERVAL {normal_priority_weekend_hours} HOUR) < NOW()
                                        THEN 1
                                    WHEN DAYOFWEEK(b.openedDate) NOT IN (1, 6, 7) AND 
                                         DATE_ADD(b.resolvedDate, INTERVAL {normal_priority_normal_hours} HOUR) < NOW()
                                        THEN 1
                                    ELSE 0
                                END
                        END
                    ELSE 0
                END AS 是否超时,

                -- 超时状态说明
                CASE
                    WHEN b.status = 'closed' AND b.closedBy = b.openedBy THEN '关闭超时'
                    WHEN b.status = 'resolved' THEN '验证超时'
                    ELSE '非超时状态'
                END AS 超时类别

            FROM
                zt_user u
            JOIN
                zt_bug b ON (u.account = b.closedBy OR u.account = b.assignedTo)
            WHERE
                u.realname = %s
                AND b.product = %s
                AND b.openedDate BETWEEN %s AND %s
                AND b.deleted = '0'
                AND b.type NOT IN ({exclude_types_str})
                AND (
                    -- 已关闭的bug超时条件
                    (b.status = 'closed' AND b.closedBy = b.openedBy AND
                        (
                            ((b.severity = 1 OR b.pri = 1) AND
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN 
                                        TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {high_priority_weekend_hours}
                                    ELSE 
                                        TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {high_priority_normal_hours}
                                END)
                            OR
                            ((b.severity != 1 AND b.pri != 1) AND
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN 
                                        TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {normal_priority_weekend_hours}
                                    ELSE 
                                        TIMESTAMPDIFF(HOUR, b.resolvedDate, b.closedDate) > {normal_priority_normal_hours}
                                END)
                        ))
                    OR
                    -- 已解决但未关闭的bug超时条件
                    (b.status = 'resolved' AND
                        (
                            ((b.severity = 1 OR b.pri = 1) AND
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN 
                                        DATE_ADD(b.resolvedDate, INTERVAL {high_priority_weekend_hours} HOUR) < NOW()
                                    ELSE 
                                        DATE_ADD(b.resolvedDate, INTERVAL {high_priority_normal_hours} HOUR) < NOW()
                                END)
                            OR
                            ((b.severity != 1 AND b.pri != 1) AND
                                CASE
                                    WHEN DAYOFWEEK(b.openedDate) IN (1, 6, 7) THEN 
                                        DATE_ADD(b.resolvedDate, INTERVAL {normal_priority_weekend_hours} HOUR) < NOW()
                                    ELSE 
                                        DATE_ADD(b.resolvedDate, INTERVAL {normal_priority_normal_hours} HOUR) < NOW()
                                END)
                        ))
                )
            ORDER BY
                b.openedDate DESC;
            """

            # 执行查询
            self.mysql_db.cur.execute(sql_query, (tester_name, product_id, start_date, end_date))
            results = self.mysql_db.fetchall_to_dict()

            # 转换为DataFrame
            if results:
                df = pd.DataFrame(results)
                return df
            else:
                return pd.DataFrame()

        except Exception as e:
            st.error(f"查询测试超时明细失败: {str(e)}")
            return None